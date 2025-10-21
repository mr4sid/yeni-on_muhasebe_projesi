# pencereler.py Dosyasının. Tamamım.
import os
from datetime import datetime, date, timedelta
import multiprocessing
import threading
import calendar
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill 
import locale 
import json
import requests
import shutil
import traceback
from PySide6.QtWidgets import (QTableWidget, QTableWidgetItem, QFormLayout,
    QApplication, QMessageBox, QFileDialog,
    QWidget, QDialog, QPushButton, QVBoxLayout,
    QHBoxLayout, QGridLayout, QLabel, QLineEdit, QComboBox,
    QTreeWidget, QTreeWidgetItem, QAbstractItemView, QHeaderView, QTextEdit,
    QCheckBox, QFrame, QGroupBox, QDialogButtonBox,
    QMenu, QTabWidget,QSizePolicy, QProgressBar)
from PySide6.QtGui import QFont, QPixmap, QDoubleValidator, QBrush, QColor, QPalette
from PySide6.QtCore import Qt, QTimer, Signal, QLocale, Slot, QThread, QObject
from veritabani import OnMuhasebe
from hizmetler import FaturaService, TopluIslemService, CariService
from yardimcilar import DatePickerDialog, normalize_turkish_chars, setup_locale, format_and_validate_numeric_input
from config import API_BASE_URL

# Logger kurulumu
logger = logging.getLogger(__name__)

if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)

def setup_numeric_entry(parent_app, entry_widget, allow_negative=False, decimal_places=2, max_value=None):
    validator = QDoubleValidator()

    # Yerel ayarı QLocale nesnesiyle ayarla
    current_locale = QLocale(locale.getlocale()[0])
    validator.setLocale(current_locale)

    validator.setBottom(0.0 if not allow_negative else -999999999.0)
    validator.setTop(999999999999.0 if max_value is None else float(max_value))
    validator.setDecimals(decimal_places)
    validator.setNotation(QDoubleValidator.StandardNotation)
    entry_widget.setValidator(validator)

    # SADECE ODAK KAYBINDA VEYA ENTER TUŞUNA BASILDIĞINDA FORMATLAMA YAP
    # Bu, kullanıcının serbestçe karakterleri silmesine olanak tanır.
    entry_widget.editingFinished.connect(lambda: format_and_validate_numeric_input(entry_widget, parent_app))

    # Başlangıçta 0'dan farklı bir değer varsa, onu formatla
    if entry_widget.text() and entry_widget.text().strip() != "0,00":
        format_and_validate_numeric_input(entry_widget, parent_app)
        

def setup_date_entry(parent_app, entry_widget):
    pass

class SiparisPenceresi(QDialog):
    # SiparisOlusturmaSayfasi'ndan dönen sinyalleri yakalamak için sinyaller tanımlanıyor
    saved_successfully = Signal()
    cancelled_successfully = Signal()

    def __init__(self, parent, db_manager, app_ref, siparis_tipi, siparis_id_duzenle=None, yenile_callback=None, initial_cari_id=None, initial_urunler=None, initial_data=None):
        super().__init__(parent)
        self.app = app_ref
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.siparis_id_duzenle = siparis_id_duzenle
        self.initial_cari_id = initial_cari_id
        self.initial_urunler = initial_urunler
        self.initial_data = initial_data

        title = "Yeni Sipariş"
        if siparis_id_duzenle:
            try:
                # DÜZELTME: db_manager metodunu kullanarak sipariş bilgisini al
                siparis_info = self.db.siparis_getir_by_id(siparis_id_duzenle)
                if siparis_info:
                    siparis_no_display = siparis_info.get('siparis_no', 'Bilinmiyor')
                    title = f"Sipariş Güncelleme: {siparis_no_display}"
                else:
                    title = "Sipariş Güncelleme: Hata"
                    QMessageBox.critical(self, "Hata", "Sipariş bilgisi yüklenirken hata oluştu.")
            except Exception as e:
                logging.error(f"Sipariş bilgisi çekilirken hata: {e}")
                QMessageBox.critical(self, "Hata", "Sipariş bilgisi yüklenirken hata oluştu.")
                title = "Sipariş Güncelleme: Hata"
        else:
            title = "Yeni Müşteri Siparişi" if siparis_tipi == self.db.SIPARIS_TIP_SATIS else "Yeni Tedarikçi Siparişi"

        self.setWindowTitle(title)
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)
        
        dialog_layout = QVBoxLayout(self)

        from arayuz import SiparisOlusturmaSayfasi
        self.siparis_form = SiparisOlusturmaSayfasi(
            self,
            self.db,
            self.app,
            islem_tipi=siparis_tipi, # DÜZELTME: siparis_tipi parametresi islem_tipi olarak adlandırıldı
            duzenleme_id=siparis_id_duzenle,
            yenile_callback=self._siparis_kaydedildi,
            initial_cari_id=initial_cari_id,
            initial_urunler=initial_urunler,
            initial_data=initial_data
        )
        dialog_layout.addWidget(self.siparis_form)

        self.siparis_form.saved_successfully.connect(self.accept)
        self.siparis_form.cancelled_successfully.connect(self.reject)
    
    def _siparis_kaydedildi(self):
        """Sipariş kaydedildiğinde çağrılan iç metot."""
        self.saved_successfully.emit()
        if self.yenile_callback:
            self.yenile_callback()

class CariHesapEkstresiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, cari_id, cari_tip, pencere_basligi, parent_list_refresh_func=None):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.cari_id = cari_id
        self.cari_tip = cari_tip
        self.cari_ad_gosterim = pencere_basligi
        self.parent_list_refresh_func = parent_list_refresh_func
        self.current_user_id = self.app.current_user.get("id") if self.app and hasattr(self.app, 'current_user') else None        
        self.hareket_detay_map = {}
        self.kasa_banka_map = {}

        self.setWindowTitle(f"Cari Hesap Ekstresi: {self.cari_ad_gosterim}")
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        # Önce tüm UI elemanlarını oluştur
        self._setup_ui()
        
        # Ardından verileri yükle ve UI'ı güncelle
        self._yukle_ozet_bilgileri()
        self.ekstreyi_yukle()
        
        self.notebook.currentChanged.connect(self._on_tab_change)

        self.finished.connect(self.on_dialog_finished)
        self.app.register_cari_ekstre_window(self)

    def _setup_ui(self):
        """Tüm UI elemanlarını tek bir metotta oluşturur."""
        main_container = QWidget(self)
        self.setLayout(QVBoxLayout(main_container))
        
        self.ozet_ve_bilgi_frame = QGroupBox("Cari Özet Bilgileri", self)
        self.layout().addWidget(self.ozet_ve_bilgi_frame)
        self._create_ozet_bilgi_alani()

        self.notebook = QTabWidget(self)
        self.layout().addWidget(self.notebook)
        
        self.hesap_hareketleri_tab = QWidget(self.notebook)
        self.notebook.addTab(self.hesap_hareketleri_tab, "Hesap Hareketleri")
        self._create_hesap_hareketleri_tab(self.hesap_hareketleri_tab)

        self.siparisler_tab = QWidget(self.notebook)
        self.notebook.addTab(self.siparisler_tab, "Siparişler")
        self._create_siparisler_tab()
        
        self.hizli_islemler_ana_frame = QFrame(self)
        self.layout().addWidget(self.hizli_islemler_ana_frame)
        self._create_hizli_islem_alanlari()
        
        today = date.today()
        start_date = today - timedelta(days=3 * 365) if self.cari_tip == "TEDARIKCI" else today - timedelta(days=6 * 30)
        self.bas_tarih_entry.setText(start_date.strftime('%Y-%m-%d'))
        self.bitis_tarih_entry.setText(today.strftime('%Y-%m-%d'))

    def on_dialog_finished(self, result):
        self.app.unregister_cari_ekstre_window(self)
        if self.parent_list_refresh_func:
            self.parent_list_refresh_func()

    def _on_tab_change(self, index):
        selected_tab_text = self.notebook.tabText(index)
        if selected_tab_text == "Siparişler":
            self._siparisleri_yukle()
        elif selected_tab_text == "Hesap Hareketleri":
            self.ekstreyi_yukle()

    def _yukle_cari_bilgileri(self):
        try:
            cari_adi = "Bilinmiyor"
            cari_telefon = ""
            
            if self.cari_tip == self.db.CARI_TIP_MUSTERI:
                # DÜZELTME: musteri_getir_by_id metoduna kullanici_id parametresi eklendi
                cari_data = self.db.musteri_getir_by_id(self.cari_id, self.current_user_id)
                if cari_data:
                    cari_adi = cari_data.get("ad", "Bilinmeyen Müşteri")
                    cari_telefon = cari_data.get("telefon", "")
                
            elif self.cari_tip == self.db.CARI_TIP_TEDARIKCI:
                # DÜZELTME: tedarikci_getir_by_id metoduna kullanici_id parametresi eklendi
                cari_data = self.db.tedarikci_getir_by_id(self.cari_id, self.current_user_id)
                if cari_data:
                    cari_adi = cari_data.get("ad", "Bilinmeyen Tedarikçi")
                    cari_telefon = cari_data.get("telefon", "")
            
            self.setWindowTitle(f"{cari_adi} - Cari Hesap Ekstresi")
        except Exception as e:
            logger.error(f"Cari bilgileri yüklenirken hata oluştu: {e}")
            QMessageBox.warning(self, "Hata", f"Cari bilgileri yüklenirken bir hata oluştu: {e}")

    def _create_hesap_hareketleri_tab(self, parent_frame):
        parent_frame.setLayout(QVBoxLayout(parent_frame))
        
        filter_frame = QFrame(parent_frame)
        parent_frame.layout().addWidget(filter_frame)
        self._create_filter_alani(filter_frame)

        tree_frame = QFrame(parent_frame)
        parent_frame.layout().addWidget(tree_frame)
        self._create_treeview_alani(tree_frame)
        
    def _create_siparisler_tab(self):
        parent_frame = self.siparisler_tab
        parent_frame.setLayout(QVBoxLayout(parent_frame))
        
        cols = ("ID", "Sipariş No", "Tarih", "Teslimat Tarihi", "Toplam Tutar", "Durum", "Fatura No")
        self.siparisler_tree = QTreeWidget(parent_frame)
        self.siparisler_tree.setHeaderLabels(cols)
        self.siparisler_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.siparisler_tree.setSortingEnabled(True)

        col_defs = [
            ("ID", 40, Qt.AlignCenter), ("Sipariş No", 150, Qt.AlignCenter), ("Tarih", 100, Qt.AlignCenter),
            ("Teslimat Tarihi", 100, Qt.AlignCenter), ("Toplam Tutar", 120, Qt.AlignCenter), ("Durum", 120, Qt.AlignCenter),
            ("Fatura No", 150, Qt.AlignCenter)
        ]
        for i, (col_id, w, a) in enumerate(col_defs):
            self.siparisler_tree.setColumnWidth(i, w)
            self.siparisler_tree.headerItem().setTextAlignment(i, a)
            self.siparisler_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))

        self.siparisler_tree.header().setStretchLastSection(False)
        self.siparisler_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        
        parent_frame.layout().addWidget(self.siparisler_tree)
        self.siparisler_tree.itemDoubleClicked.connect(self._on_siparis_double_click)

    def _siparisleri_yukle(self):
        self.siparisler_tree.clear()
        
        siparisler_data = []
        try:
            params = {
                'cari_id': self.cari_id
            }
            siparisler_data_response = self.db.siparis_listesi_al(**params) 
            siparisler_data = siparisler_data_response.get("items", []) 

            for siparis in siparisler_data:
                item_qt = QTreeWidgetItem(self.siparisler_tree)
                item_qt.setData(0, Qt.UserRole, siparis.get('id', -1))

                tarih_obj = datetime.strptime(str(siparis.get('tarih')), '%Y-%m-%d').date() if siparis.get('tarih') else None
                teslimat_tarihi_obj = datetime.strptime(str(siparis.get('teslimat_tarihi')), '%Y-%m-%d').date() if siparis.get('teslimat_tarihi') else None
                
                formatted_tarih = tarih_obj.strftime('%d.%m.%Y') if isinstance(tarih_obj, date) else '-'
                formatted_teslimat_tarihi = teslimat_tarihi_obj.strftime('%d.%m.%Y') if isinstance(teslimat_tarihi_obj, date) else '-'

                item_qt.setText(0, str(siparis.get('id', '')))
                item_qt.setText(1, siparis.get('siparis_no', ''))
                item_qt.setText(2, formatted_tarih)
                item_qt.setText(3, formatted_teslimat_tarihi)
                item_qt.setText(4, self.db._format_currency(siparis.get('toplam_tutar', 0.0)))
                item_qt.setText(5, siparis.get('durum', ''))
                
                fatura_no_text = "-"
                if siparis.get('fatura_id'):
                    try:
                        fatura_data = self.db.fatura_getir_by_id(siparis.get('fatura_id'))
                        fatura_no_text = fatura_data.get('fatura_no', '-')
                    except Exception:
                        fatura_no_text = "Hata"
                item_qt.setText(6, fatura_no_text)

                if siparis.get('durum') == "TAMAMLANDI":
                    for col_idx in range(self.siparisler_tree.columnCount()):
                        item_qt.setBackground(col_idx, QBrush(QColor("lightgreen")))
                elif siparis.get('durum') == "İPTAL_EDİLDİ":
                    for col_idx in range(self.siparisler_tree.columnCount()):
                        item_qt.setBackground(col_idx, QBrush(QColor("lightgray")))
                        item_qt.setForeground(col_idx, QBrush(QColor("gray")))
                        font = item_qt.font(col_idx)
                        font.setStrikeOut(True)
                        item_qt.setFont(col_idx, font)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Siparişler yüklenirken hata: {e}")
            logging.error(f"Cari Hesap Ekstresi - Siparişler yükleme hatası: {e}", exc_info=True)
        self.app.set_status_message(f"{self.cari_ad_gosterim} için {len(siparisler_data)} sipariş listelendi.", "blue")

    def _on_siparis_double_click(self, item, column):
        siparis_id = item.data(0, Qt.UserRole)
        if siparis_id:
            from pencereler import SiparisDetayPenceresi
            SiparisDetayPenceresi(self.app, self.db, siparis_id).exec()

    def _create_ozet_bilgi_alani(self):
        frame = self.ozet_ve_bilgi_frame
        frame.setLayout(QGridLayout(frame))

        label_font_buyuk = QFont("Segoe UI", 10, QFont.Bold)
        deger_font_buyuk = QFont("Segoe UI", 10)
        label_font_kucuk = QFont("Segoe UI", 9, QFont.Bold)
        deger_font_kucuk = QFont("Segoe UI", 9)

        finans_ozet_cerceve = QGroupBox("Finansal Özet", frame)
        finans_ozet_cerceve.setLayout(QGridLayout(finans_ozet_cerceve))
        frame.layout().addWidget(finans_ozet_cerceve, 0, 0)

        row_idx_finans = 0
        finans_ozet_cerceve.layout().addWidget(QLabel("Dönem Başı Bakiye:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_donem_basi_bakiye = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_donem_basi_bakiye, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Toplam Borç Hareketi:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_toplam_borc_hareketi = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_toplam_borc_hareketi, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Toplam Alacak Hareketi:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_toplam_alacak_hareketi = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_toplam_alacak_hareketi, row_idx_finans, 1)
        row_idx_finans += 1
        
        finans_ozet_cerceve.layout().addWidget(QLabel("Toplam Tahsilat/Ödeme:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_toplam_tahsilat_odeme = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_toplam_tahsilat_odeme, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Vadesi Gelmiş Borç/Alacak:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_vadesi_gelmis = QLabel("0,00 TL", font=deger_font_kucuk, styleSheet="color: red;")
        finans_ozet_cerceve.layout().addWidget(self.lbl_vadesi_gelmis, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Vadesi Gelecek Borç/Alacak:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_vadesi_gelecek = QLabel("0,00 TL", font=deger_font_kucuk, styleSheet="color: blue;")
        finans_ozet_cerceve.layout().addWidget(self.lbl_vadesi_gelecek, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Dönem Sonu Bakiye:", font=label_font_buyuk), row_idx_finans, 0)
        self.lbl_ozet_net_bakiye = QLabel("0,00 TL", font=deger_font_buyuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_ozet_net_bakiye, row_idx_finans, 1)

        cari_detay_cerceve = QGroupBox("Cari Detay Bilgileri", frame)
        cari_detay_cerceve.setLayout(QGridLayout(cari_detay_cerceve))
        frame.layout().addWidget(cari_detay_cerceve, 0, 1)

        row_idx_cari = 0
        cari_detay_cerceve.layout().addWidget(QLabel("Cari Adı:", font=label_font_kucuk), row_idx_cari, 0)
        self.lbl_cari_detay_ad = QLabel("-", font=deger_font_kucuk)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_ad, row_idx_cari, 1)
        row_idx_cari += 1

        cari_detay_cerceve.layout().addWidget(QLabel("Telefon:", font=label_font_kucuk), row_idx_cari, 0)
        self.lbl_cari_detay_tel = QLabel("-", font=deger_font_kucuk)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_tel, row_idx_cari, 1)
        row_idx_cari += 1

        cari_detay_cerceve.layout().addWidget(QLabel("Adres:", font=label_font_kucuk), row_idx_cari, 0, Qt.AlignTop)
        self.lbl_cari_detay_adres = QLabel("-", font=deger_font_kucuk, wordWrap=True)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_adres, row_idx_cari, 1)
        row_idx_cari += 1

        cari_detay_cerceve.layout().addWidget(QLabel("Vergi No:", font=label_font_kucuk), row_idx_cari, 0)
        self.lbl_cari_detay_vergi = QLabel("-", font=deger_font_kucuk)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_vergi, row_idx_cari, 1)
        row_idx_cari += 1

        export_buttons_frame = QFrame(frame)
        export_buttons_frame.setLayout(QVBoxLayout(export_buttons_frame))
        frame.layout().addWidget(export_buttons_frame, 0, 2, Qt.AlignTop)

        btn_pdf = QPushButton("PDF'e Aktar")
        btn_pdf.clicked.connect(self.pdf_aktar)
        export_buttons_frame.layout().addWidget(btn_pdf)

        btn_excel = QPushButton("Excel'e Aktar")
        btn_excel.clicked.connect(self.excel_aktar)
        export_buttons_frame.layout().addWidget(btn_excel)
        
        btn_update_cari = QPushButton("Cari Bilgilerini Güncelle")
        btn_update_cari.clicked.connect(self._cari_bilgileri_guncelle)
        cari_detay_cerceve.layout().addWidget(btn_update_cari, row_idx_cari, 0, 1, 2)

        btn_cari_sil = QPushButton("Cariyi Sil")
        btn_cari_sil.clicked.connect(self._cari_sil)
        btn_cari_sil.setStyleSheet("background-color: #f44336; color: white;")
        export_buttons_frame.layout().addWidget(btn_cari_sil)

    def _cari_sil(self):
        """Cariyi siler, varsayılan carilerin silinmesini engeller."""
        # Varsayılan carilerin ID'lerini kontrol et
        perakende_musteri_id = self.db.get_perakende_musteri_id()
        genel_tedarikci_id = self.db.get_genel_tedarikci_id()
        
        if (self.cari_tip == self.db.CARI_TIP_MUSTERI and self.cari_id == perakende_musteri_id) or \
        (self.cari_tip == self.db.CARI_TIP_TEDARIKCI and self.cari_id == genel_tedarikci_id):
            QMessageBox.warning(self, "Silme Engellendi", "Varsayılan cari hesaplar silinemez, sadece düzenlenebilir.")
            return

        reply = QMessageBox.question(self, 'Cariyi Sil Onayı',
                                    f"'{self.cari_ad_gosterim}' adlı cariyi silmek istediğinizden emin misiniz? Bu işlem geri alınamaz.",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            try:
                success, message = False, "Bilinmeyen hata."
                if self.cari_tip == self.db.CARI_TIP_MUSTERI:
                    # KRİTİK DÜZELTME: musteri_sil metodu 1 argüman bekler. self.current_user_id kaldırıldı.
                    success, message = self.db.musteri_sil(self.cari_id)
                elif self.cari_tip == self.db.CARI_TIP_TEDARIKCI:
                    # KRİTİK DÜZELTME: tedarikci_sil metodu 1 argüman bekler. self.current_user_id kaldırıldı.
                    success, message = self.db.tedarikci_sil(self.cari_id)
                
                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.close() # Pencereyi kapat
                    if self.parent_list_refresh_func: # Ana listeyi yenile
                        self.parent_list_refresh_func()
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                logger.error(f"Cari silinirken hata oluştu: {e}", exc_info=True)
                QMessageBox.critical(self, "Hata", f"Cari silinirken bir hata oluştu:\n{e}")

    def _create_filter_alani(self, filter_frame):
        filter_frame.setLayout(QHBoxLayout(filter_frame))
        
        filter_frame.layout().addWidget(QLabel("Başlangıç Tarihi:"))
        self.bas_tarih_entry = QLineEdit()
        filter_frame.layout().addWidget(self.bas_tarih_entry)
        
        btn_date_start = QPushButton("🗓️")
        btn_date_start.setFixedWidth(30)
        btn_date_start.clicked.connect(lambda: self._open_date_picker_for_entry(self.bas_tarih_entry))
        filter_frame.layout().addWidget(btn_date_start)

        filter_frame.layout().addWidget(QLabel("Bitiş Tarihi:"))
        self.bitis_tarih_entry = QLineEdit()
        filter_frame.layout().addWidget(self.bitis_tarih_entry)
        
        btn_date_end = QPushButton("🗓️")
        btn_date_end.setFixedWidth(30)
        btn_date_end.clicked.connect(lambda: self._open_date_picker_for_entry(self.bitis_tarih_entry))
        filter_frame.layout().addWidget(btn_date_end)

        btn_filter = QPushButton("Filtrele")
        btn_filter.clicked.connect(self.ekstreyi_yukle)
        filter_frame.layout().addWidget(btn_filter)

        today = date.today()
        start_date = today - timedelta(days=3 * 365) if self.cari_tip == "TEDARIKCI" else today - timedelta(days=6 * 30)
        self.bas_tarih_entry.setText(start_date.strftime('%Y-%m-%d'))
        self.bitis_tarih_entry.setText(today.strftime('%Y-%m-%d'))

    def _open_date_picker_for_entry(self, target_entry_qlineedit):
        from yardimcilar import DatePickerDialog
        initial_date_str = target_entry_qlineedit.text() if target_entry_qlineedit.text() else None
        
        dialog = DatePickerDialog(self.app, initial_date=initial_date_str)
        
        dialog.date_selected.connect(target_entry_qlineedit.setText)
        
        dialog.exec()

    def _create_treeview_alani(self, tree_frame):
        tree_frame.setLayout(QVBoxLayout(tree_frame))
        
        cols = ("ID", "Tarih", "Saat", "İşlem Tipi", "Referans", "Ödeme Türü", "Açıklama/Detay", "Borç", "Alacak", "Bakiye", "Vade Tarihi")
        self.ekstre_tree = QTreeWidget(tree_frame)
        self.ekstre_tree.setHeaderLabels(cols)
        self.ekstre_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ekstre_tree.setSortingEnabled(True)

        self.ekstre_tree.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.ekstre_tree.setStyleSheet("QTreeWidget::item { height: 30px; }")

        col_defs = [
            ("ID", 30, Qt.AlignCenter),
            ("Tarih", 100, Qt.AlignCenter),
            ("Saat", 80, Qt.AlignCenter),
            ("İşlem Tipi", 120, Qt.AlignCenter),
            ("Referans", 120, Qt.AlignCenter),
            ("Ödeme Türü", 120, Qt.AlignCenter),
            ("Açıklama/Detay", 180, Qt.AlignCenter),
            ("Borç", 80, Qt.AlignCenter),
            ("Alacak", 80, Qt.AlignCenter),
            ("Bakiye", 120, Qt.AlignCenter),
            ("Vade Tarihi", 90, Qt.AlignCenter)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs):
            self.ekstre_tree.setColumnWidth(i, width)
            self.ekstre_tree.headerItem().setTextAlignment(i, alignment)
            self.ekstre_tree.headerItem().setFont(i, QFont("Segoe UI", 12, QFont.Bold))
        
        self.ekstre_tree.header().setStretchLastSection(False)
        self.ekstre_tree.header().setSectionResizeMode(6, QHeaderView.Stretch)

        tree_frame.layout().addWidget(self.ekstre_tree)
        
        self.ekstre_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ekstre_tree.customContextMenuRequested.connect(self._show_context_menu)
        self.ekstre_tree.itemDoubleClicked.connect(self.on_double_click_hareket_detay)

    def _create_hizli_islem_alanlari(self):
        self.hizli_islemler_ana_frame.setLayout(QHBoxLayout(self.hizli_islemler_ana_frame))

        ot_frame_text = "Ödeme Ekle" if self.cari_tip == "TEDARIKCI" else "Tahsilat Ekle"
        odeme_tahsilat_frame = QGroupBox(ot_frame_text, self.hizli_islemler_ana_frame)
        odeme_tahsilat_frame.setLayout(QGridLayout(odeme_tahsilat_frame))
        self.hizli_islemler_ana_frame.layout().addWidget(odeme_tahsilat_frame)

        odeme_tahsilat_frame.layout().addWidget(QLabel("Ödeme Tipi:"), 0, 0)
        self.ot_odeme_tipi_combo = QComboBox()
        self.ot_odeme_tipi_combo.addItems([self.db.ODEME_TURU_NAKIT, self.db.ODEME_TURU_KART, 
                                            self.db.ODEME_TURU_EFT_HAVALE, self.db.ODEME_TURU_CEK, 
                                            self.db.ODEME_TURU_SENET])
        self.ot_odeme_tipi_combo.setCurrentText(self.db.ODEME_TURU_NAKIT)
        self.ot_odeme_tipi_combo.currentIndexChanged.connect(self._ot_odeme_tipi_degisince)
        odeme_tahsilat_frame.layout().addWidget(self.ot_odeme_tipi_combo, 0, 1)

        odeme_tahsilat_frame.layout().addWidget(QLabel("Tutar:"), 1, 0)
        self.ot_tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.ot_tutar_entry)
        odeme_tahsilat_frame.layout().addWidget(self.ot_tutar_entry, 1, 1)

        odeme_tahsilat_frame.layout().addWidget(QLabel("Kasa/Banka:"), 2, 0)
        self.ot_kasa_banka_combo = QComboBox()
        self.ot_kasa_banka_combo.setEnabled(True)  # Varsayılan olarak aktif yapıldı
        odeme_tahsilat_frame.layout().addWidget(self.ot_kasa_banka_combo, 2, 1, 1, 2) # 1 satır, 2 sütun kapla

        odeme_tahsilat_frame.layout().addWidget(QLabel("Not:"), 3, 0)
        self.ot_not_entry = QLineEdit()
        odeme_tahsilat_frame.layout().addWidget(self.ot_not_entry, 3, 1, 1, 2)

        btn_ot_save = QPushButton(ot_frame_text)
        btn_ot_save.clicked.connect(self._hizli_odeme_tahsilat_kaydet)
        odeme_tahsilat_frame.layout().addWidget(btn_ot_save, 4, 0, 1, 3)

        borc_frame = QGroupBox("Veresiye Borç Ekle", self.hizli_islemler_ana_frame)
        borc_frame.setLayout(QGridLayout(borc_frame))
        self.hizli_islemler_ana_frame.layout().addWidget(borc_frame)

        borc_frame.layout().addWidget(QLabel("Türü Seçiniz:"), 0, 0)
        self.borc_tur_combo = QComboBox()
        self.borc_tur_combo.addItems(["Diğer Borç", "Satış Faturası"])
        borc_frame.layout().addWidget(self.borc_tur_combo, 0, 1)

        borc_frame.layout().addWidget(QLabel("Tutar:"), 1, 0)
        self.borc_tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.borc_tutar_entry)
        borc_frame.layout().addWidget(self.borc_tutar_entry, 1, 1)

        borc_frame.layout().addWidget(QLabel("Not:"), 2, 0)
        self.borc_not_entry = QLineEdit()
        borc_frame.layout().addWidget(self.borc_not_entry, 2, 1)

        btn_borc_save = QPushButton("Veresiye Ekle")
        btn_borc_save.clicked.connect(self._hizli_veresiye_borc_kaydet)
        borc_frame.layout().addWidget(btn_borc_save, 3, 0, 1, 2)

        alacak_frame = QGroupBox("Alacak Ekleme", self.hizli_islemler_ana_frame)
        alacak_frame.setLayout(QGridLayout(alacak_frame))
        self.hizli_islemler_ana_frame.layout().addWidget(alacak_frame)

        alacak_frame.layout().addWidget(QLabel("Türü Seçiniz:"), 0, 0)
        self.alacak_tur_combo = QComboBox()
        self.alacak_tur_combo.addItems(["Diğer Alacak", "İade Faturası"])
        alacak_frame.layout().addWidget(self.alacak_tur_combo, 0, 1)

        alacak_frame.layout().addWidget(QLabel("Tutar:"), 1, 0)
        self.alacak_tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.alacak_tutar_entry)
        alacak_frame.layout().addWidget(self.alacak_tutar_entry, 1, 1)

        alacak_frame.layout().addWidget(QLabel("Not:"), 2, 0)
        self.alacak_not_entry = QLineEdit()
        alacak_frame.layout().addWidget(self.alacak_not_entry, 2, 1)

        btn_alacak_save = QPushButton("Alacak Kaydet")
        btn_alacak_save.clicked.connect(self._hizli_alacak_kaydet)
        alacak_frame.layout().addWidget(btn_alacak_save, 3, 0, 1, 2)
        
        self._yukle_kasa_banka_hesaplarini_hizli_islem_formu()
        # Yeni eklenen satır: Form oluşturulduktan sonra varsayılan seçimi yapmak için çağırıyoruz.
        self._ot_odeme_tipi_degisince()

    def _yukle_kasa_banka_hesaplarini_hizli_islem_formu(self):
        self.ot_kasa_banka_combo.clear()
        self.kasa_banka_map.clear()
        
        try:
            hesaplar_response = self.db.kasa_banka_listesi_al()
            if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                hesaplar = hesaplar_response["items"]
            elif isinstance(hesaplar_response, list):
                hesaplar = hesaplar_response
                self.app.set_status_message("Uyarı: Kasa/Banka listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                hesaplar = []
                self.app.set_status_message("Hata: Kasa/Banka listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Kasa/Banka listesi API'den beklenen formatta gelmedi: {type(hesaplar_response)} - {hesaplar_response}", exc_info=True)
                self.ot_kasa_banka_combo.addItem("Hesap Yok", None)
                self.ot_kasa_banka_combo.setEnabled(False)
                return

            if hesaplar:
                for h in hesaplar:
                    display_text = f"{h.get('hesap_adi')} ({h.get('tip')})"
                    if h.get('tip') == "BANKA" and h.get('banka_adi'):
                        display_text += f" - {h.get('banka_adi')}"
                    if h.get('bakiye') is not None:
                        display_text += f" (Bakiye: {self.db._format_currency(h.get('bakiye'))})"

                    self.kasa_banka_map[display_text] = h.get('id')
                    self.ot_kasa_banka_combo.addItem(display_text, h.get('id'))
                self.ot_kasa_banka_combo.setCurrentIndex(0)
                self.ot_kasa_banka_combo.setEnabled(True)
            else:
                self.ot_kasa_banka_combo.clear()
                self.ot_kasa_banka_combo.addItem("Hesap Yok", None)
                self.ot_kasa_banka_combo.setEnabled(False)

            self.app.set_status_message(f"{len(hesaplar)} kasa/banka hesabı API'den yüklendi.", "blue")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kasa/Banka hesapları yüklenirken hata: {e}")
            logging.error(f"Kasa/Banka yükleme hatası: {e}", exc_info=True)
            self.ot_kasa_banka_combo.addItem("Hesap Yok", None)
            self.ot_kasa_banka_combo.setEnabled(False)

    def _ot_odeme_tipi_degisince(self):
        """Hızlı işlem formunda ödeme tipi değiştiğinde kasa/banka seçimini ayarlar."""
        selected_odeme_sekli = self.ot_odeme_tipi_combo.currentText()
        
        self.ot_kasa_banka_combo.blockSignals(True)
        
        varsayilan_hesap_id = None
        try:
            varsayilan_hesap_data = self.db.get_kasa_banka_by_odeme_turu(selected_odeme_sekli)
            # get_kasa_banka_by_odeme_turu'nun (id, ad) tuple döndürdüğü varsayılıyor
            if varsayilan_hesap_data and isinstance(varsayilan_hesap_data, tuple):
                varsayilan_hesap_id = varsayilan_hesap_data[0]
        except Exception as e:
            logging.warning(f"Varsayılan kasa/banka ({selected_odeme_sekli}) çekilirken hata: {e}")

        # Kasa/Banka combobox'ını yeniden doldurmak yerine, var olanlar arasından seçmeye çalışıyoruz.
        # Bu, performansı artırır ve sinyal hatalarını önler.
        if varsayilan_hesap_id:
            index_to_set = -1
            for i in range(self.ot_kasa_banka_combo.count()):
                if self.ot_kasa_banka_combo.itemData(i) == varsayilan_hesap_id:
                    index_to_set = i
                    break
            
            if index_to_set != -1:
                self.ot_kasa_banka_combo.setCurrentIndex(index_to_set)
            else:
                # Varsayılan hesap listede yoksa, ilk geçerli hesabı seç
                if self.ot_kasa_banka_combo.count() > 0 and self.ot_kasa_banka_combo.itemData(0) is not None:
                    self.ot_kasa_banka_combo.setCurrentIndex(0)
                else:
                    self.ot_kasa_banka_combo.setCurrentText("")
        else:
            # Varsayılan hesap tanımlı değilse, ilk geçerli hesabı seç
            if self.ot_kasa_banka_combo.count() > 0 and self.ot_kasa_banka_combo.itemData(0) is not None:
                self.ot_kasa_banka_combo.setCurrentIndex(0)
            else:
                self.ot_kasa_banka_combo.setCurrentText("")
                self.ot_kasa_banka_combo.setEnabled(False)

        self.ot_kasa_banka_combo.blockSignals(False)
                
    def _yukle_ozet_bilgileri(self):
        try:
            cari_detail = None
            if self.cari_tip == self.db.CARI_TIP_MUSTERI:
                # DÜZELTME: musteri_getir_by_id metodu 2 parametre bekler. kullanici_id kaldırıldı.
                cari_detail = self.db.musteri_getir_by_id(self.cari_id)
            else:
                # TEDARIKCI İÇİN KULLANICI ID GEREKLİ OLDUĞU VARSAYILIYOR. (veritabani.py'ye göre)
                cari_detail = self.db.tedarikci_getir_by_id(self.cari_id, self.current_user_id)

            if not cari_detail:
                self.app.set_status_message(f"Hata: Cari bilgiler yüklenemedi. ID {self.cari_id} bulunamadı.", "red")
                return

            self.lbl_cari_detay_ad.setText(cari_detail.get('ad', '-'))
            self.lbl_cari_detay_tel.setText(cari_detail.get('telefon', '-'))
            self.lbl_cari_detay_adres.setText(cari_detail.get('adres', '-'))
            vergi_info = f"{cari_detail.get('vergi_dairesi', '-')} / {cari_detail.get('vergi_no', '-')}"
            self.lbl_cari_detay_vergi.setText(vergi_info)

            # API'den özet verilerini çekme
            # DÜZELTME: get_cari_ekstre_ozet metodu 5 parametre bekler, kullanici_id kaldırıldı.
            ekstre_ozet_data = self.db.get_cari_ekstre_ozet(
                self.cari_id, self.cari_tip,
                self.bas_tarih_entry.text(), self.bitis_tarih_entry.text()
            )
            
            self.lbl_donem_basi_bakiye.setText(self.db._format_currency(ekstre_ozet_data.get("donem_basi_bakiye", 0.0)))
            self.lbl_toplam_borc_hareketi.setText(self.db._format_currency(ekstre_ozet_data.get("toplam_borc_hareketi", 0.0)))
            self.lbl_toplam_alacak_hareketi.setText(self.db._format_currency(ekstre_ozet_data.get("toplam_alacak_hareketi", 0.0)))
            
            self.lbl_toplam_tahsilat_odeme.setText(self.db._format_currency(ekstre_ozet_data.get("toplam_alacak_hareketi", 0.0)))
            
            self.lbl_vadesi_gelmis.setText(self.db._format_currency(ekstre_ozet_data.get("vadesi_gelmis", 0.0)))
            self.lbl_vadesi_gelecek.setText(self.db._format_currency(ekstre_ozet_data.get("vadesi_gelecek", 0.0)))

            # Dönem sonu bakiyesini hesapla ve göster
            donem_sonu_bakiye = ekstre_ozet_data.get("donem_sonu_bakiye", 0.0)
            bakiye_metni = self.db._format_currency(abs(donem_sonu_bakiye))
            
            if self.cari_tip == self.db.CARI_TIP_MUSTERI:
                if donem_sonu_bakiye > 0:
                    bakiye_metni = f"<b style='color: red;'>{bakiye_metni} BORÇLU</b>"
                elif donem_sonu_bakiye < 0:
                    bakiye_metni = f"<b style='color: green;'>{bakiye_metni} ALACAKLI</b>"
                else:
                    bakiye_metni = f"<b style='color: blue;'>{bakiye_metni}</b>"
            elif self.cari_tip == self.db.CARI_TIP_TEDARIKCI:
                if donem_sonu_bakiye > 0:
                    bakiye_metni = f"<b style='color: green;'>{bakiye_metni} ALACAKLI</b>"
                elif donem_sonu_bakiye < 0:
                    bakiye_metni = f"<b style='color: red;'>{bakiye_metni} BORÇLU</b>"
                else:
                    bakiye_metni = f"<b style='color: blue;'>{bakiye_metni}</b>"
            
            self.lbl_ozet_net_bakiye.setText(bakiye_metni)

            self.app.set_status_message("Cari özet bilgileri güncellendi.", "green")

        except Exception as e:
            logging.error(f"Cari özet bilgileri yüklenirken hata oluştu: {e}", exc_info=True)
            self.app.set_status_message(f"Hata: Cari özet bilgileri yüklenemedi. Detay: {e}", "red")

    def _cari_bilgileri_guncelle(self):
        try:
            cari_data = None
            if self.cari_tip == "MUSTERI":
                cari_data = self.db.musteri_getir_by_id(self.cari_id)
                if cari_data:
                    dialog = YeniMusteriEklePenceresi(self, self.db, self._ozet_ve_liste_yenile, musteri_duzenle=cari_data, app_ref=self.app)
                    dialog.exec()
                else:
                    self.app.set_status_message(f"Hata: Müşteri bilgileri yüklenemedi. ID {self.cari_id} bulunamadı.", "red")
                    return
            elif self.cari_tip == "TEDARIKCI":
                cari_data = self.db.tedarikci_getir_by_id(self.cari_id, self.current_user_id)
                if cari_data:
                    from pencereler import YeniTedarikciEklePenceresi
                    dialog = YeniTedarikciEklePenceresi(self, self.db, self._ozet_ve_liste_yenile, tedarikci_duzenle=cari_data, app_ref=self.app)
                    dialog.exec()
                else:
                    self.app.set_status_message(f"Hata: Tedarikçi bilgileri yüklenemedi. ID {self.cari_id} bulunamadı.", "red")
                    return

            self.app.set_status_message(f"{self.cari_tip} kartı açıldı.", "blue")

        except Exception as e:
            logger.error(f"Cari bilgiler güncellenmek üzere yüklenirken hata oluştu: {e}", exc_info=True)
            self.app.set_status_message(f"Hata: Cari bilgiler yüklenemedi. Detay: {e}", "red")

    def _ozet_ve_liste_yenile(self):
        self._yukle_ozet_bilgileri()
        self.ekstreyi_yukle()

    def _hizli_odeme_tahsilat_kaydet(self):
        islem_turu = self.sender().text()
        islem_turu_enum = "GİDER" if islem_turu == "Ödeme Ekle" else "GELİR"

        tutar_str = self.ot_tutar_entry.text().replace(".", "").replace(",", ".")
        try:
            tutar = float(tutar_str)
            if tutar <= 0:
                self.app.set_status_message("Tutar sıfırdan büyük olmalıdır.", "orange")
                return
        except ValueError:
            self.app.set_status_message("Geçerli bir tutar girin.", "orange")
            return

        aciklama = self.ot_not_entry.text().strip()
        if not aciklama:
            self.app.set_status_message("Açıklama alanı boş bırakılamaz.", "orange")
            return

        selected_hesap_idx = self.ot_kasa_banka_combo.currentIndex()
        if selected_hesap_idx < 0:
            self.app.set_status_message("Lütfen bir Kasa/Banka hesabı seçin.", "orange")
            return
        
        kasa_banka_id = self.ot_kasa_banka_combo.currentData()
        odeme_turu = self.ot_odeme_tipi_combo.currentText()

        gelir_gider_data = {
            "tarih": datetime.now().strftime('%Y-%m-%d'),
            "islem_saati": datetime.now().strftime('%H:%M:%S'),
            "tip": islem_turu_enum,
            "tutar": tutar,
            "aciklama": aciklama,
            "kaynak": "MANUEL",
            "kasa_banka_id": kasa_banka_id,
            "cari_id": self.cari_id,
            "cari_tip": self.cari_tip,
            "odeme_turu": odeme_turu,
            "kullanici_id": self.current_user_id # DÜZELTME: kullanici_id eklendi
        }

        try:
            success = self.db.gelir_gider_ekle(gelir_gider_data, self.current_user_id) # DÜZELTME: kullanici_id eklendi
            if success:
                self.app.set_status_message(f"Hızlı {islem_turu.lower()} kaydı başarıyla oluşturuldu.", "green")
                self.ot_tutar_entry.clear()
                self.ot_not_entry.clear()
                self.ot_odeme_tipi_combo.setCurrentText(self.db.ODEME_TURU_NAKIT)
                self._ot_odeme_tipi_degisince()
                self._ozet_ve_liste_yenile()
            else:
                self.app.set_status_message(f"Hızlı {islem_turu.lower()} kaydı oluşturulamadı.", "red")
        except Exception as e:
            logger.error(f"Hızlı {islem_turu.lower()} kaydı oluşturulurken hata oluştu: {e}", exc_info=True)
            self.app.set_status_message(f"Hata: Hızlı {islem_turu.lower()} kaydı oluşturulamadı. Detay: {e}", "red")

    def _hizli_veresiye_borc_kaydet(self):
        borc_tur = self.borc_tur_combo.currentText()
        tutar_str = self.borc_tutar_entry.text().replace(',', '.')
        not_str = self.borc_not_entry.text()

        if not tutar_str or float(tutar_str) <= 0:
            QMessageBox.warning(self, "Eksik Bilgi", "Lütfen geçerli bir tutar giriniz.")
            return

        if borc_tur == "Satış Faturası":
            QMessageBox.information(self, "Yönlendirme", "Fatura oluşturmak için lütfen ana menüden 'Yeni Satış Faturası' ekranını kullanın.")
        else:
            try:
                tutar_f = float(tutar_str)
                data = {
                    "cari_id": self.cari_id,
                    "cari_turu": self.cari_tip,
                    "tarih": date.today().strftime('%Y-%m-%d'),
                    "tutar": tutar_f,
                    "aciklama": not_str,
                    "islem_turu": "VERESİYE_BORÇ",
                    "islem_yone": "BORC",
                    "kaynak": self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL,
                    "kullanici_id": self.current_user_id # DÜZELTME: kullanici_id eklendi
                }
                # DÜZELTME: cari_hareket_ekle_manuel metoduna kullanici_id eklendi
                success = self.db.cari_hareket_ekle_manuel(data, self.current_user_id)

                if success:
                    QMessageBox.information(self, "Başarılı", "Veresiye borç başarıyla eklendi.")
                    self.borc_tutar_entry.clear()
                    self.borc_not_entry.clear()
                    self._ozet_ve_liste_yenile()
                else:
                    QMessageBox.critical(self, "Hata", "Veresiye borç eklenirken hata.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Veresiye borç eklenirken hata: {e}")
                logging.error(f"Hızlı veresiye borç kaydetme hatası: {e}", exc_info=True)

    def _hizli_alacak_kaydet(self):
        QMessageBox.information(self, "Geliştirme Aşamasında", "Alacak ekleme özelliği henüz tamamlanmamıştır.")

    def excel_aktar(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Cari Hesap Ekstresini Excel'e Kaydet", 
                                                 f"Cari_Ekstresi_{self.cari_ad_gosterim.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx", 
                                                 "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)")
        if file_path:
            bekleme_penceresi = BeklemePenceresi(self, message="Ekstre Excel'e aktarılıyor, lütfen bekleyiniz...")
            threading.Thread(target=lambda: self._generate_ekstre_excel_threaded(
                self.cari_tip, self.cari_id, self.bas_tarih_entry.text(), self.bitis_tarih_entry.text(),
                file_path, bekleme_penceresi
            )).start()
        else:
            self.app.set_status_message("Excel'e aktarma iptal edildi.", "blue")

    def pdf_aktar(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Cari Hesap Ekstresini PDF'e Kaydet", 
                                                 f"Cari_Ekstresi_{self.cari_ad_gosterim.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf", 
                                                 "PDF Dosyaları (*.pdf);;Tüm Dosyalar (*)")
        if file_path:
            bekleme_penceresi = BeklemePenceresi(self, message="Ekstre PDF'e aktarılıyor, lütfen bekleyiniz...")
            
            result_queue = multiprocessing.Queue()
            pdf_process = multiprocessing.Process(target=self.db.cari_ekstresi_pdf_olustur, args=(
                self.db.data_dir,
                self.cari_tip,
                self.cari_id,
                self.bas_tarih_entry.text(),
                self.bitis_tarih_entry.text(),
                file_path,
                result_queue
            ))
            pdf_process.start()

            self.app.process_queue_timer = QTimer(self.app)
            self.app.process_queue_timer.timeout.connect(lambda: self._check_pdf_process_completion(result_queue, pdf_process, bekleme_penceresi))
            self.app.process_queue_timer.start(100)
        else:
            self.app.set_status_message("PDF'e aktarma iptal edildi.", "blue")

    def _check_pdf_process_completion(self, result_queue, pdf_process, bekleme_penceresi):
        if not result_queue.empty():
            success, message = result_queue.get()
            bekleme_penceresi.close()
            self.app.process_queue_timer.stop()

            if success:
                QMessageBox.information(self, "Başarılı", message)
                self.app.set_status_message(message, "green")
            else:
                QMessageBox.critical(self, "Hata", message)
                self.app.set_status_message(f"Ekstre PDF'e aktarılırken hata: {message}", "red")
            pdf_process.join()
            
        elif not pdf_process.is_alive():
            bekleme_penceresi.close()
            self.app.process_queue_timer.stop()
            QMessageBox.critical(self, "Hata", "PDF işlemi beklenmedik şekilde sonlandı.")
            self.app.set_status_message("PDF işlemi beklenmedik şekilde sonlandı.", "red")
            pdf_process.join()

    def _generate_ekstre_excel_threaded(self, cari_tip, cari_id, bas_t, bit_t, dosya_yolu, bekleme_penceresi):
        local_db_manager = self.db.__class__(api_base_url=self.db.api_base_url, app_ref=self.app)
        
        success = False
        message = ""
        try:
            hareketler_listesi, devreden_bakiye, success_db, message_db = local_db_manager.cari_hesap_ekstresi_al(
                cari_id, cari_tip, bas_t, bit_t
            )
            
            if not success_db:
                message = f"Ekstre verisi alınırken hata: {message_db}"
            elif not hareketler_listesi and devreden_bakiye == 0:
                message = "Excel'e aktarılacak cari ekstre verisi bulunamadı."
            else:
                success, message = local_db_manager.tarihsel_satis_raporu_excel_olustur(
                    rapor_verileri=hareketler_listesi,
                    dosya_yolu=dosya_yolu,
                    bas_t=bas_t,
                    bit_t=bit_t
                )
                if not success: message = f"Excel oluşturulurken hata: {message}"

        except Exception as e:
            message = f"Rapor Excel'e aktarılırken bir hata oluştu:\n{e}"
            logging.error(f"Excel export thread error: {e}", exc_info=True)
        finally:
            self.app.set_status_message(message, "blue" if success else "red")
            self.app.after(0, bekleme_penceresi.kapat)
            self.app.after(0, lambda: QMessageBox.information(self, "Excel Aktarım", message) if success else QMessageBox.critical(self, "Excel Aktarım Hatası", message))

    def ekstreyi_yukle(self):
        self.ekstre_tree.clear()
        self.hareket_detay_map.clear()

        bas_tarih_str = self.bas_tarih_entry.text()
        bitis_tarih_str = self.bitis_tarih_entry.text()

        try:
            datetime.strptime(bas_tarih_str, '%Y-%m-%d')
            datetime.strptime(bitis_tarih_str, '%Y-%m-%d')
        except ValueError:
            QMessageBox.critical(self, "Hata", "Tarih formatı 'YYYY-AA-GG' şeklinde olmalıdır.")
            return
        
        # API'den gelen veriyi çek
        # DÜZELTME: cari_hesap_ekstresi_al metodu 5 parametre bekler, self.current_user_id kaldırıldı.
        hareketler_listesi, devreden_bakiye, success_db, message_db = self.db.cari_hesap_ekstresi_al(
            self.cari_id, self.cari_tip, bas_tarih_str, bitis_tarih_str
        )

        if not success_db:
            QMessageBox.critical(self, "Hata", f"Ekstre verisi alınırken hata: {message_db}")
            self.app.set_status_message(f"{self.cari_ad_gosterim} için ekstre yüklenemedi: {message_db}", "red")
            return
        
        # Devir bakiyesi satırı
        devir_item = QTreeWidgetItem(self.ekstre_tree)
        devir_item.setText(0, "")
        devir_item.setText(1, bas_tarih_str)
        devir_item.setText(2, "")
        devir_item.setText(3, "DEVİR")
        devir_item.setText(4, "")
        devir_item.setText(5, "Devreden Bakiye")
        devir_item.setText(6, "")
        devir_item.setText(7, self.db._format_currency(devreden_bakiye) if devreden_bakiye > 0 else "")
        devir_item.setText(8, self.db._format_currency(abs(devreden_bakiye)) if devreden_bakiye < 0 else "")
        devir_item.setText(9, self.db._format_currency(devreden_bakiye))
        devir_item.setText(10, "")
        
        for col_idx in range(self.ekstre_tree.columnCount()):
            devir_item.setBackground(col_idx, QBrush(QColor("#EFEFEF")))
            devir_item.setFont(col_idx, QFont("Segoe UI", 9, QFont.Bold))

        current_bakiye = devreden_bakiye
        
        for hareket in hareketler_listesi:
            # Borç ve alacak yönüne göre çalışan bakiyeyi güncelle
            if self.cari_tip == 'MUSTERI':
                if hareket.get('islem_yone') == 'BORC':
                    current_bakiye += hareket.get('tutar', 0)
                elif hareket.get('islem_yone') == 'ALACAK':
                    current_bakiye -= hareket.get('tutar', 0)
            elif self.cari_tip == 'TEDARIKCI':
                if hareket.get('islem_yone') == 'BORC':
                    current_bakiye -= hareket.get('tutar', 0)
                elif hareket.get('islem_yone') == 'ALACAK':
                    current_bakiye += hareket.get('tutar', 0)

            self._ekstreye_satir_ekle(hareket, current_bakiye)

        self.app.set_status_message(f"{self.cari_ad_gosterim} için {len(hareketler_listesi)} hareket yüklendi.", "blue")

    def _ekstreye_satir_ekle(self, hareket, current_bakiye):
        item_qt = QTreeWidgetItem(self.ekstre_tree)
            
        tarih_formatted = hareket['tarih'].strftime('%d.%m.%Y') if isinstance(hareket['tarih'], date) else str(hareket['tarih'])
        vade_tarihi_formatted = hareket['vade_tarihi'].strftime('%d.%m.%Y') if isinstance(hareket['vade_tarihi'], date) else (str(hareket['vade_tarihi']) if hareket['vade_tarihi'] else '-')
        
        borc_val = ""
        alacak_val = ""
        
        if hareket['islem_yone'] == 'BORC':
            borc_val = self.db._format_currency(hareket['tutar'])
        elif hareket['islem_yone'] == 'ALACAK':
            alacak_val = self.db._format_currency(hareket['tutar'])
        
        display_islem_tipi = hareket['islem_turu']
        display_ref_gosterim = hareket['fatura_no'] if hareket.get('fatura_no') else (hareket.get('kaynak') or '-')

        if hareket.get('kaynak') in (self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA):
            if hareket.get('fatura_turu') == self.db.FATURA_TIP_SATIS:
                display_islem_tipi = "Satış Faturası"
            elif hareket.get('fatura_turu') == self.db.FATURA_TIP_ALIS:
                display_islem_tipi = "Alış Faturası"
            elif hareket.get('fatura_turu') == self.db.FATURA_TIP_SATIS_IADE:
                display_islem_tipi = "Satış İade Faturası"
            elif hareket.get('fatura_turu') == self.db.FATURA_TIP_ALIS_IADE:
                display_islem_tipi = "Alış İade Faturası"
            display_ref_gosterim = hareket['fatura_no']
        elif hareket.get('kaynak') in (self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME):
            display_islem_tipi = "Tahsilat" if hareket.get('islem_turu') == "GELIR" else "Ödeme"
            display_ref_gosterim = hareket.get('kaynak')
        
        islem_saati_str = hareket.get('islem_saati') or ''

        item_qt.setText(0, str(hareket['id']))
        item_qt.setText(1, tarih_formatted)
        item_qt.setText(2, islem_saati_str)
        item_qt.setText(3, display_islem_tipi)
        item_qt.setText(4, display_ref_gosterim)
        item_qt.setText(5, hareket.get('odeme_turu') or '-')
        item_qt.setText(6, hareket.get('aciklama') or '-')
        item_qt.setText(7, borc_val)
        item_qt.setText(8, alacak_val)
        item_qt.setText(9, self.db._format_currency(current_bakiye))
        item_qt.setText(10, vade_tarihi_formatted)

        self.hareket_detay_map[hareket['id']] = hareket

        if hareket.get('kaynak') in (self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA):
            if hareket.get('odeme_turu') in self.db.pesin_odeme_turleri:
                for col_idx in range(self.ekstre_tree.columnCount()):
                    item_qt.setBackground(col_idx, QBrush(QColor("lightgray")))
                    item_qt.setForeground(col_idx, QBrush(QColor("darkgray")))
            else:
                for col_idx in range(self.ekstre_tree.columnCount()):
                    item_qt.setForeground(col_idx, QBrush(QColor("red")))
            if "İADE" in hareket.get('fatura_turu', ''):
                for col_idx in range(self.ekstre_tree.columnCount()):
                    item_qt.setBackground(col_idx, QBrush(QColor("#FFF2CC")))
                    item_qt.setForeground(col_idx, QBrush(QColor("#A67400")))
        elif hareket.get('kaynak') in (self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME, self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL) or hareket.get('kaynak') == self.db.KAYNAK_TIP_MANUEL:
            for col_idx in range(self.ekstre_tree.columnCount()):
                item_qt.setForeground(col_idx, QBrush(QColor("green")))

    def _show_context_menu(self, pos):
        item = self.ekstre_tree.itemAt(pos)
        if not item: return

        item_id_str = item.text(0)
        if not item_id_str: return
        try:
            item_id = int(item_id_str)
        except ValueError:
            return

        hareket_detayi = self.hareket_detay_map.get(item_id)
        if not hareket_detayi: return

        ref_tip = hareket_detayi.get('kaynak')

        context_menu = QMenu(self)
        
        # DÜZELTME: Sadece manuel eklenen veya fatura kaynaklı işlemler için silme seçeneği
        if ref_tip in [self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME, self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL] or \
           ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            context_menu.addAction("İşlemi Sil").triggered.connect(self.secili_islemi_sil)
        
        if ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            context_menu.addAction("Faturayı Güncelle").triggered.connect(self.secili_islemi_guncelle)
        
        if context_menu.actions():
            context_menu.exec(self.ekstre_tree.mapToGlobal(pos))

    def secili_islemi_sil(self):
        selected_items = self.ekstre_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir işlem seçin.")
            return

        item_qt = selected_items[0]
        hareket_id = int(item_qt.text(0))

        hareket_detayi = self.hareket_detay_map.get(hareket_id)
        if not hareket_detayi:
            QMessageBox.critical(self, "Hata", "İşlem detayları bulunamadı.")
            return
        
        ref_id = hareket_detayi.get('kaynak_id')
        ref_tip = hareket_detayi.get('kaynak')
        aciklama_text = hareket_detayi.get('aciklama')
        fatura_no = hareket_detayi.get('fatura_no')
        
        confirm_msg = f"'{aciklama_text}' açıklamalı işlemi silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz."
        
        if ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            confirm_msg = f"'{fatura_no}' numaralı FATURA ve ilişkili tüm hareketlerini silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz."
            reply = QMessageBox.question(self, "Silme Onayı", confirm_msg, QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                try:
                    # DÜZELTME: fatura_sil metoduna kullanici_id parametresi eklendi
                    success, message = self.db.fatura_sil(ref_id, self.current_user_id)
                    if success:
                        QMessageBox.information(self, "Başarılı", message)
                        self._ozet_ve_liste_yenile()
                    else:
                        QMessageBox.critical(self, "Hata", message)
                except Exception as e:
                    QMessageBox.critical(self, "Hata", f"Fatura silinirken hata: {e}")
            else:
                self.app.set_status_message("Silme işlemi iptal edildi.", "blue")
                return
        
        elif ref_tip in [self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME, self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL, self.db.KAYNAK_TIP_MANUEL]:
            confirm_msg = f"'{aciklama_text}' açıklamalı işlemi silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz."
            reply = QMessageBox.question(self, "Silme Onayı", confirm_msg, QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                try:
                    # DÜZELTME: gelir_gider_sil metoduna kullanici_id parametresi eklendi
                    success, message = self.db.gelir_gider_sil(hareket_id, self.current_user_id)
                    if success:
                        QMessageBox.information(self, "Başarılı", message)
                        self._ozet_ve_liste_yenile()
                    else:
                        QMessageBox.critical(self, "Hata", message)
                except Exception as e:
                    QMessageBox.critical(self, "Hata", f"Manuel işlem silinirken hata: {e}")
            else:
                self.app.set_status_message("Silme işlemi iptal edildi.", "blue")
                return
        
        else: # Diğer otomatik veya bilinmeyen kaynaklar için uyarı
            QMessageBox.warning(self, "Silme Engellendi", f"Sadece manuel olarak eklenen kayıtlar veya fatura kaynaklı işlemler buradan silinebilir. Bu kayıt türü ({ref_tip}) buradan silinemez.")
            return

        # Ortak yenileme işlemleri
        if hasattr(self.app, 'fatura_listesi_sayfasi'):
            self.app.fatura_listesi_sayfasi.fatura_listesini_yukle()
        if hasattr(self.app, 'gelir_gider_sayfasi'):
            if hasattr(self.app.gelir_gider_sayfasi, 'gelir_listesi_frame'):
                self.app.gelir_gider_sayfasi.gelir_listesi_frame.gg_listesini_yukle()
            if hasattr(self.app.gelir_gider_sayfasi, 'gider_listesi_frame'):
                self.app.gelir_gider_sayfasi.gider_listesi_frame.gg_listesini_yukle()
        if hasattr(self.app, 'kasa_banka_yonetimi_sayfasi'):
            self.app.kasa_banka_yonetimi_sayfasi.hesap_listesini_yenile()

    def secili_islemi_guncelle(self):
        selected_items = self.ekstre_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen güncellemek için bir fatura işlemi seçin.")
            return

        item_qt = selected_items[0]
        hareket_id = int(item_qt.text(0))

        hareket_detayi = self.hareket_detay_map.get(hareket_id)
        if not hareket_detayi:
            QMessageBox.critical(self, "Hata", "İşlem detayları bulunamadı.")
            return
        
        ref_id = hareket_detayi.get('kaynak_id')
        ref_tip = hareket_detayi.get('kaynak')

        if ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            if ref_id:
                from pencereler import FaturaGuncellemePenceresi
                # DÜZELTME: FaturaGuncellemePenceresi'ne ana uygulama referansı (self.app) gönderildi
                FaturaGuncellemePenceresi(self.app, self.db, ref_id, self._ozet_ve_liste_yenile).exec()
            else:
                QMessageBox.information(self, "Detay", "Fatura referansı bulunamadı.")
        else:
            QMessageBox.information(self, "Bilgi", "Sadece fatura işlemleri güncellenebilir.")

    def on_double_click_hareket_detay(self, item, column):
        if item.text(3) == "DEVİR":
            QMessageBox.warning(self, "Uyarı", "Devir satırı için detay görüntülenemez.")
            return

        hareket_id = int(item.text(0))
        hareket_detay = self.hareket_detay_map.get(hareket_id)

        if not hareket_detay:
            QMessageBox.critical(self, "Hata", "Seçilen işlemin detayları bulunamadı.")
            return

        ref_id = hareket_detay.get('kaynak_id')
        ref_tip_str = hareket_detay.get('kaynak')

        if ref_tip_str in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            if ref_id:
                from pencereler import FaturaDetayPenceresi
                # DÜZELTME: FaturaDetayPenceresi'ne ana uygulama referansı (self.app) gönderildi
                FaturaDetayPenceresi(self.app, self.db, ref_id).exec()
            else:
                QMessageBox.information(self, "Detay", "Fatura referansı bulunamadı.")
        elif ref_tip_str in [self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME, self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL]:
            tarih_gosterim = hareket_detay.get('tarih').strftime('%d.%m.%Y') if isinstance(hareket_detay.get('tarih'), date) else str(hareket_detay.get('tarih'))
            tutar_gosterim = self.db._format_currency(hareket_detay.get('tutar'))
            aciklama_gosterim = hareket_detay.get('aciklama') or "Açıklama yok."
            
            QMessageBox.information(self, "İşlem Detayı",
                                 f"Bu bir {ref_tip_str} işlemidir.\n"
                                 f"Tarih: {tarih_gosterim}\n"
                                 f"Tutar: {tutar_gosterim}\n" 
                                 f"Açıklama: {aciklama_gosterim}\n"
                                 f"Referans ID: {hareket_id}")
        else:
            QMessageBox.information(self, "Detay", "Bu işlem tipi için detay görüntüleme mevcut değil.")
            
    def clear_layout(self, layout):
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
            else:
                self.clear_layout(item.layout())
            
# pencereler.py dosyasında FaturaGuncellemePenceresi'nin yeni içeriği
class FaturaGuncellemePenceresi(QDialog):
    saved_successfully = Signal()

    def __init__(self, parent, db_manager, fatura_id_duzenle, yenile_callback_liste=None):
        super().__init__(parent)
        self.app = parent.app if hasattr(parent, 'app') else parent
        self.db = db_manager
        self.yenile_callback_liste = yenile_callback_liste
        self.fatura_id_duzenle = fatura_id_duzenle
        
        try:
            fatura_ana_bilgileri = self.db.fatura_getir_by_id(self.fatura_id_duzenle)
            if not fatura_ana_bilgileri:
                QMessageBox.critical(self, "Hata", f"ID {self.fatura_id_duzenle} olan fatura bulunamadı.")
                self.reject()
                return
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Fatura bilgileri çekilirken bir hata oluştu: {e}")
            logger.error(f"Fatura bilgileri yüklenirken hata oluştu: {e}", exc_info=True)
            self.reject()
            return
        
        fatura_turu = fatura_ana_bilgileri.get('fatura_turu')
        fatura_no = fatura_ana_bilgileri.get('fatura_no', 'Bilinmiyor')

        self.setWindowTitle(f"Fatura Güncelleme: {fatura_no}")
        # KRİTİK DÜZELTME: Pencereyi tam ekran aç
        self.setWindowState(Qt.WindowMaximized) 
        self.setModal(True)
        
        dialog_layout = QVBoxLayout(self)

        from arayuz import FaturaOlusturmaSayfasi
        self.fatura_olusturma_form = FaturaOlusturmaSayfasi(
            self,
            self.db,
            self.app,
            fatura_tipi=fatura_turu,
            duzenleme_id=self.fatura_id_duzenle,
            yenile_callback=self._fatura_kaydedildi
        )
        dialog_layout.addWidget(self.fatura_olusturma_form)

        # FaturaOlusturmaSayfasi'ndaki sinyallere bağlanıyoruz.
        self.fatura_olusturma_form.saved_successfully.connect(self.accept)
        self.fatura_olusturma_form.cancelled_successfully.connect(self.reject)

        self.finished.connect(self.on_dialog_finished)

    def on_dialog_finished(self, result):
        if self.yenile_callback_liste:
            self.yenile_callback_liste()
    
    def _fatura_kaydedildi(self):
        """Kaydetme işlemi tamamlandığında çağrılan iç metot."""
        self.saved_successfully.emit()

class FaturaDetayPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, fatura_id):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.fatura_id = fatura_id
        
        self.fatura_ana = None
        self.fatura_kalemleri_db = None

        # Fatura tip sabitlerini db_manager'dan al
        self.FATURA_TIP_ALIS = self.db.FATURA_TIP_ALIS
        self.FATURA_TIP_SATIS = self.db.FATURA_TIP_SATIS
        self.FATURA_TIP_DEVIR_GIRIS = self.db.FATURA_TIP_DEVIR_GIRIS
        self.FATURA_TIP_SATIS_IADE = self.db.FATURA_TIP_SATIS_IADE
        self.FATURA_TIP_ALIS_IADE = self.db.FATURA_TIP_ALIS_IADE

        self.ODEME_TURU_NAKIT = self.db.ODEME_TURU_NAKIT
        self.ODEME_TURU_KART = self.db.ODEME_TURU_KART
        self.ODEME_TURU_EFT_HAVALE = self.db.ODEME_TURU_EFT_HAVALE
        self.ODEME_TURU_CEK = self.db.ODEME_TURU_CEK
        self.ODEME_TURU_SENET = self.db.ODEME_TURU_SENET
        self.ODEME_TURU_ACIK_HESAP = self.db.ODEME_TURU_ACIK_HESAP
        self.ODEME_TURU_ETKISIZ_FATURA = self.db.ODEME_TURU_ETKISIZ_FATURA

        self.f_no = "Yükleniyor..."
        self.tip = ""

        self.setWindowTitle(f"Fatura Detayları: {self.f_no}")
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        self.main_layout = QVBoxLayout(self)
        
        self._create_ui_and_populate_data()

    def _create_ui_and_populate_data(self):
        """Arayüzü oluşturur, API'den verileri çeker ve arayüzü doldurur."""
        if self.main_layout.layout():
            self.clear_layout(self.main_layout)
        try:
            # DÜZELTME: fatura_getir_by_id metodu 2 parametre bekler, kullanici_id kaldırıldı.
            self.fatura_ana = self.db.fatura_getir_by_id(self.fatura_id)
            if not self.fatura_ana:
                raise Exception("Fatura ana bilgileri API'den alınamadı.")
            
            # DÜZELTME: fatura_kalemleri_al metodu 2 parametre bekler, kullanici_id kaldırıldı.
            self.fatura_kalemleri_db = self.db.fatura_kalemleri_al(self.fatura_id)
            if not self.fatura_kalemleri_db:
                logging.warning(f"Fatura ID {self.fatura_id} için fatura kalemi bulunamadı.")
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Fatura bilgileri çekilemedi: {e}")
            self.close()
            return
        
        # --- Veri Çekme ve Hazırlama ---
        self.f_no = self.fatura_ana.get('fatura_no', '-')
        self.tip = self.fatura_ana.get('fatura_turu', '-')
        self.setWindowTitle(f"Fatura Detayları: {self.f_no} ({self.tip})")

        tarih_db = self.fatura_ana.get('tarih')
        c_id = self.fatura_ana.get('cari_id')
        toplam_kdv_haric_fatura_ana_db = self.db.safe_float(self.fatura_ana.get('toplam_kdv_haric'))
        toplam_kdv_dahil_fatura_ana_db = self.db.safe_float(self.fatura_ana.get('toplam_kdv_dahil'))
        odeme_turu_db = self.fatura_ana.get('odeme_turu')
        misafir_adi_db = self.fatura_ana.get('misafir_adi')
        kasa_banka_id_db = self.fatura_ana.get('kasa_banka_id')
        olusturma_tarihi_saat = self.fatura_ana.get('olusturma_tarihi_saat')
        olusturan_kullanici_id = self.fatura_ana.get('olusturan_kullanici_id')
        son_guncelleme_tarihi_saat = self.fatura_ana.get('son_guncelleme_tarihi_saat')
        son_guncelleyen_kullanici_id = self.fatura_ana.get('son_guncelleyen_kullanici_id')
        fatura_notlari_db = self.fatura_ana.get('fatura_notlari')
        vade_tarihi_db = self.fatura_ana.get('vade_tarihi')
        genel_iskonto_tipi_db = self.fatura_ana.get('genel_iskonto_tipi')
        genel_iskonto_degeri_db = self.db.safe_float(self.fatura_ana.get('genel_iskonto_degeri'))
        
        try:
            # DÜZELTME: kullanici_listele metoduna kullanici_id parametresi eklendi
            kullanicilar_list = self.db.kullanici_listele(kullanici_id=self.app.current_user[0])
            kullanicilar_map = {k.get('id'): k.get('kullanici_adi') for k in kullanicilar_list.get('items', [])}
        except Exception as e:
            logger.error(f"Kullanıcı listesi API'den alınamadı: {e}")
            kullanicilar_map = {}
        
        olusturan_adi = kullanicilar_map.get(olusturan_kullanici_id, "Bilinmiyor")
        son_guncelleyen_adi = kullanicilar_map.get(son_guncelleyen_kullanici_id, "Bilinmiyor")
        
        cari_adi_text = "Bilinmiyor"
        if str(c_id) == str(self.db.get_perakende_musteri_id()) and self.fatura_ana.get('fatura_turu') == self.db.FATURA_TIP_SATIS:
            cari_adi_text = "Perakende Satış Müşterisi"
            if misafir_adi_db: cari_adi_text += f" (Misafir: {misafir_adi_db})"
        else:
            cari_bilgi_db = None
            if self.fatura_ana.get('fatura_turu') in [self.db.FATURA_TIP_SATIS, self.db.FATURA_TIP_SATIS_IADE]:
                # DÜZELTME: musteri_getir_by_id metoduna kullanici_id parametresi eklendi
                cari_bilgi_db = self.db.musteri_getir_by_id(c_id, self.app.current_user[0])
                if cari_bilgi_db and cari_bilgi_db.get('kod'):
                    cari_adi_text = f"{cari_bilgi_db.get('ad')} (Kod: {cari_bilgi_db.get('kod')})"
            elif self.fatura_ana.get('fatura_turu') in [self.db.FATURA_TIP_ALIS, self.db.FATURA_TIP_ALIS_IADE]:
                # DÜZELTME: tedarikci_getir_by_id metoduna kullanici_id parametresi eklendi
                cari_bilgi_db = self.db.tedarikci_getir_by_id(c_id, self.app.current_user[0])
                if cari_bilgi_db and cari_bilgi_db.get('kod'):
                    cari_adi_text = f"{cari_bilgi_db.get('ad')} (Kod: {cari_bilgi_db.get('kod')})"

        # --- Arayüz Oluşturma ---
        # Tutarlı yazı tipleri tanımlandı.
        font_label = QFont("Segoe UI", 14, QFont.Bold)
        font_value = QFont("Segoe UI", 13)
        font_header = QFont("Segoe UI", 14, QFont.Bold)
        font_groupbox = QFont("Segoe UI", 13, QFont.Bold)
        
        # Ana yatay layout: Üst bilgiler ve toplamlar bir arada
        self.ust_bilgiler_frame = QFrame(self)
        self.ust_bilgiler_layout = QHBoxLayout(self.ust_bilgiler_frame)
        self.ust_bilgiler_layout.setContentsMargins(10, 10, 10, 10)
        self.ust_bilgiler_layout.setSpacing(15)
        self.main_layout.addWidget(self.ust_bilgiler_frame)

        # Üç ana panel için çerçeveler oluşturuldu
        self.sol_panel_frame = QGroupBox("Fatura Bilgileri", self.ust_bilgiler_frame)
        self.orta_panel_frame = QGroupBox("Finansal Bilgiler", self.ust_bilgiler_frame)
        self.sag_panel_frame = QGroupBox("Finansal Özet", self.ust_bilgiler_frame)
        
        self.ust_bilgiler_layout.addWidget(self.sol_panel_frame)
        self.ust_bilgiler_layout.addWidget(self.orta_panel_frame)
        self.ust_bilgiler_layout.addWidget(self.sag_panel_frame)

        # --- Sol Panel (Fatura Bilgileri) ---
        sol_layout = QGridLayout(self.sol_panel_frame)
        sol_layout.setContentsMargins(10, 15, 10, 10)
        sol_layout.setSpacing(5)

        try: fatura_tarihi_formatted = datetime.strptime(str(tarih_db), '%Y-%m-%d').strftime('%d.%m.%Y')
        except: fatura_tarihi_formatted = str(tarih_db)
        
        sol_layout.addWidget(QLabel("Fatura No:", font=font_label), 0, 0)
        sol_layout.addWidget(QLabel(self.f_no, font=font_value), 0, 1)

        sol_layout.addWidget(QLabel("Tarih:", font=font_label), 1, 0)
        sol_layout.addWidget(QLabel(fatura_tarihi_formatted, font=font_value), 1, 1)

        cari_label_tipi = "Müşteri/Misafir:" if self.fatura_ana.get('fatura_turu') == self.db.FATURA_TIP_SATIS else "Tedarikçi:"
        sol_layout.addWidget(QLabel(cari_label_tipi, font=font_label), 2, 0)
        sol_layout.addWidget(QLabel(cari_adi_text, font=font_value), 2, 1)

        sol_layout.addWidget(QLabel("Oluşturan:", font=font_label), 3, 0)
        sol_layout.addWidget(QLabel(olusturan_adi, font=font_value), 3, 1)
        
        sol_layout.addWidget(QLabel("Fatura Notları:", font=font_label), 4, 0, Qt.AlignTop)
        fatura_notlari_display = QTextEdit(); fatura_notlari_display.setPlainText(fatura_notlari_db if fatura_notlari_db else ""); fatura_notlari_display.setReadOnly(True); fatura_notlari_display.setFixedHeight(60); fatura_notlari_display.setFont(font_value)
        sol_layout.addWidget(fatura_notlari_display, 4, 1)

        sol_layout.setColumnStretch(1, 1)
        
        # --- Orta Panel (Finansal Bilgiler) ---
        orta_layout = QGridLayout(self.orta_panel_frame)
        orta_layout.setContentsMargins(10, 15, 10, 10)
        orta_layout.setSpacing(5)

        orta_layout.addWidget(QLabel("Ödeme Türü:", font=font_label), 0, 0)
        orta_layout.addWidget(QLabel(odeme_turu_db, font=font_value), 0, 1)

        if kasa_banka_id_db:
            try:
                kb_bilgi = self.db.kasa_banka_getir_by_id(kasa_banka_id_db)
                if kb_bilgi:
                    orta_layout.addWidget(QLabel("Kasa/Banka:", font=font_label), 1, 0)
                    orta_layout.addWidget(QLabel(kb_bilgi.get('hesap_adi', '-'), font=font_value), 1, 1)
            except Exception as e: logging.error(f"Kasa/Banka bilgisi çekilirken hata: {e}")

        genel_iskonto_gosterim_text = "Uygulanmadı"
        if genel_iskonto_tipi_db == 'YUZDE' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Yüzde %{genel_iskonto_degeri_db:.2f}".replace('.', ',').rstrip('0').rstrip(',')
        elif genel_iskonto_tipi_db == 'TUTAR' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = self.db._format_currency(genel_iskonto_degeri_db)
        
        orta_layout.addWidget(QLabel("Genel İskonto:", font=font_label), 2, 0)
        orta_layout.addWidget(QLabel(genel_iskonto_gosterim_text, font=font_value), 2, 1)
        
        olusturma_tarihi_formatted = datetime.fromisoformat(str(olusturma_tarihi_saat)).strftime('%d.%m.%Y %H:%M:%S') if olusturma_tarihi_saat else "-"
        orta_layout.addWidget(QLabel("Oluşturma Tarih/Saat:", font=font_label), 3, 0)
        orta_layout.addWidget(QLabel(olusturma_tarihi_formatted, font=font_value), 3, 1)
        
        if odeme_turu_db == self.db.ODEME_TURU_ACIK_HESAP and vade_tarihi_db:
            orta_layout.addWidget(QLabel("Vade Tarihi:", font=font_label), 4, 0)
            orta_layout.addWidget(QLabel(str(vade_tarihi_db), font=font_value), 4, 1)

        orta_layout.setColumnStretch(1, 1)

        # --- Sağ Panel (Finansal Özet) ---
        sag_layout = QGridLayout(self.sag_panel_frame)
        sag_layout.setContentsMargins(10, 15, 10, 10)
        sag_layout.setSpacing(5)
        
        toplam_kdv_hesaplanan_detay = toplam_kdv_dahil_fatura_ana_db - toplam_kdv_haric_fatura_ana_db
        toplam_kdv_dahil_kalemler_genel_iskonto_oncesi = sum(self.db.safe_float(k.get('kalem_toplam_kdv_dahil')) for k in self.fatura_kalemleri_db if isinstance(k, dict))
        gercek_uygulanan_genel_iskonto = self.db.safe_float(toplam_kdv_dahil_kalemler_genel_iskonto_oncesi) - self.db.safe_float(toplam_kdv_dahil_fatura_ana_db)
        if gercek_uygulanan_genel_iskonto < 0: gercek_uygulanan_genel_iskonto = 0.0

        sag_layout.addWidget(QLabel("Toplam KDV Hariç:", font=font_label), 0, 0)
        self.tkh_l = QLabel(self.db._format_currency(toplam_kdv_haric_fatura_ana_db), font=font_value)
        sag_layout.addWidget(self.tkh_l, 0, 1)
        
        sag_layout.addWidget(QLabel("Toplam KDV:", font=font_label), 1, 0)
        self.tkdv_l = QLabel(self.db._format_currency(toplam_kdv_hesaplanan_detay), font=font_value)
        sag_layout.addWidget(self.tkdv_l, 1, 1)
        
        sag_layout.addWidget(QLabel("Genel Toplam:", font=font_header), 2, 0)
        self.gt_l = QLabel(self.db._format_currency(toplam_kdv_dahil_fatura_ana_db), font=QFont("Segoe UI", 16, QFont.Bold))
        sag_layout.addWidget(self.gt_l, 2, 1)
        
        sag_layout.addWidget(QLabel("Uygulanan Genel İsk:", font=font_label), 3, 0)
        self.lbl_uygulanan_genel_iskonto = QLabel(self.db._format_currency(gercek_uygulanan_genel_iskonto), font=font_value)
        sag_layout.addWidget(self.lbl_uygulanan_genel_iskonto, 3, 1)
        
        sag_layout.setColumnStretch(1, 1)

        # Fatura Kalemleri GroupBox
        kalemler_frame = QGroupBox("Fatura Kalemleri", self)
        kalemler_frame.setFont(font_groupbox)
        kalemler_frame_layout = QVBoxLayout(kalemler_frame)
        self.main_layout.addWidget(kalemler_frame, stretch=1)
        
        cols_kalem = ("Sıra", "Ürün Kodu", "Ürün Adı", "Miktar", "Birim Fiyat", "KDV %", "İsk 1 (%)", "İsk 2 (%)", "Uyg. İsk. Tutarı", "Tutar (Dah.)", "Alış.F (Fatura Anı)")
        self.kalem_tree = QTreeWidget(kalemler_frame)
        self.kalem_tree.setHeaderLabels(cols_kalem)
        self.kalem_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kalem_tree.setSortingEnabled(True)

        font_header = QFont("Segoe UI", 13, QFont.Bold)
        col_defs_kalem = [
            ("Sıra", 50, Qt.AlignCenter), ("Ürün Kodu", 120, Qt.AlignCenter), ("Ürün Adı", 190, Qt.AlignCenter),
            ("Miktar", 80, Qt.AlignCenter), ("Birim Fiyat", 100, Qt.AlignCenter), ("KDV %", 60, Qt.AlignCenter),
            ("İsk 1 (%)", 90, Qt.AlignCenter), ("İsk 2 (%)", 90, Qt.AlignCenter),
            ("Uyg. İsk. Tutarı", 110, Qt.AlignCenter), ("Tutar (Dah.)", 120, Qt.AlignCenter),
            ("Alış.F (Fatura Anı)", 150, Qt.AlignCenter)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs_kalem):
            self.kalem_tree.setColumnWidth(i, width)
            self.kalem_tree.headerItem().setTextAlignment(i, alignment)
            self.kalem_tree.headerItem().setFont(i, font_header)
        self.kalem_tree.header().setStretchLastSection(False)
        self.kalem_tree.header().setSectionResizeMode(2, QHeaderView.Stretch)
        
        kalemler_frame_layout.addWidget(self.kalem_tree)
        self._load_fatura_kalemleri_to_treeview(self.fatura_kalemleri_db)

        # Butonlar
        self._butonlari_olustur()

    def _butonlari_olustur(self):
        button_frame_alt = QFrame(self)
        button_layout_alt = QHBoxLayout(button_frame_alt)
        self.main_layout.addWidget(button_frame_alt)

        btn_guncelle = QPushButton("Güncelle")
        btn_guncelle.clicked.connect(self._open_fatura_guncelleme_penceresi)
        button_layout_alt.addWidget(btn_guncelle)
        
        btn_pdf_yazdir = QPushButton("PDF Yazdır")
        btn_pdf_yazdir.clicked.connect(self._handle_pdf_print)
        button_layout_alt.addWidget(btn_pdf_yazdir)

        button_layout_alt.addStretch()
        
        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        button_layout_alt.addWidget(btn_kapat)

    def _handle_pdf_print(self):
        """Fatura detay penceresinden PDF yazdırma işlemini başlatır."""
        dosya_adi_onek = f"{self.tip.capitalize()}Faturasi"
        file_path, _ = QFileDialog.getSaveFileName(self, f"{self.tip.capitalize()} Faturasını PDF Kaydet", 
                                                 f"{dosya_adi_onek}_{self.f_no.replace('/','_')}.pdf", 
                                                 "PDF Dosyaları (*.pdf);;Tüm Dosyalar (*)")
        if file_path:
            from pencereler import BeklemePenceresi
            bekleme_penceresi = BeklemePenceresi(self, message="Fatura PDF'e aktarılıyor, lütfen bekleyiniz...")
            QTimer.singleShot(0, bekleme_penceresi.exec)

            result_queue = multiprocessing.Queue()
            pdf_process = multiprocessing.Process(target=self.db.fatura_pdf_olustur, args=(self.fatura_id, file_path, result_queue))
            pdf_process.start()

            self.pdf_check_timer = QTimer(self)
            self.pdf_check_timer.timeout.connect(lambda: self._check_pdf_process_completion(result_queue, pdf_process, bekleme_penceresi))
            self.pdf_check_timer.start(100)
        else:
            self.app.set_status_message("PDF kaydetme iptal edildi.")

    def _check_pdf_process_completion(self, result_queue, pdf_process, bekleme_penceresi):
        if not result_queue.empty():
            success, message = result_queue.get()
            bekleme_penceresi.close()
            self.pdf_check_timer.stop()

            if success:
                QMessageBox.information(self, "Başarılı", message)
                self.app.set_status_message(message)
            else:
                QMessageBox.critical(self, "Hata", message)
                self.app.set_status_message(f"PDF kaydetme başarısız: {message}")
            pdf_process.join()
            
        elif not pdf_process.is_alive():
            bekleme_penceresi.close()
            self.pdf_check_timer.stop()
            QMessageBox.critical(self, "Hata", "PDF işlemi beklenmedik şekilde sonlandı.")
            pdf_process.join()

    def _open_fatura_guncelleme_penceresi(self):
        from pencereler import FaturaGuncellemePenceresi
        dialog = FaturaGuncellemePenceresi(
            self.app,
            self.db,
            self.fatura_id,
            yenile_callback_liste=self._fatura_guncellendi_callback_detay
        )
        dialog.exec()
        
    def _fatura_guncellendi_callback_detay(self):
        try:
            self.fatura_ana = self.db.fatura_getir_by_id(self.fatura_id)
            if not self.fatura_ana:
                raise Exception("Fatura ana bilgileri API'den alınamadı.")
            
            self.fatura_kalemleri_db = self.db.fatura_kalemleri_al(self.fatura_id)
            if not self.fatura_kalemleri_db:
                raise Exception("Fatura kalemleri API'den alınamadı.")
            
            self._create_ui_and_populate_data()
            self.app.set_status_message(f"Fatura '{self.f_no}' detayları güncellendi.")

        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Fatura detayları yenilenirken hata: {e}")
            logging.error(f"Fatura detay yenileme hatası: {e}", exc_info=True)
            self.close()
            return
                
        if hasattr(self.app, 'fatura_listesi_sayfasi'):
            if hasattr(self.app.fatura_listesi_sayfasi, 'satis_fatura_frame') and hasattr(self.app.fatura_listesi_sayfasi.satis_fatura_frame, 'fatura_listesini_yukle'):
                self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
            if hasattr(self.app.fatura_listesi_sayfasi, 'alis_fatura_frame') and hasattr(self.app.fatura_listesi_sayfasi.alis_fatura_frame, 'fatura_listesini_yukle'):
                self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()
                
    def _load_fatura_kalemleri_to_treeview(self, kalemler_list):
        """API'den gelen fatura kalemlerini QTreeWidget'a yükler."""
        self.kalem_tree.clear()
        
        # DEĞİŞİKLİK: Liste içeriğinin fontu büyütüldü
        font_item = QFont("Segoe UI", 12)

        sira_idx = 1
        for kalem_item in kalemler_list:
            miktar_db = self.db.safe_float(kalem_item.get('miktar', 0.0))
            toplam_dahil_db = self.db.safe_float(kalem_item.get('kalem_toplam_kdv_dahil', 0.0))
            original_birim_fiyat_kdv_haric_item = self.db.safe_float(kalem_item.get('birim_fiyat', 0.0))
            original_kdv_orani_item = self.db.safe_float(kalem_item.get('kdv_orani', 0.0))
            
            iskontolu_birim_fiyat_kdv_dahil = 0.0
            uygulanan_toplam_iskonto_tutari_detay = 0.0

            if miktar_db != 0:
                iskontolu_birim_fiyat_kdv_dahil = toplam_dahil_db / miktar_db
                original_birim_fiyat_kdv_dahil_kalem = original_birim_fiyat_kdv_haric_item * (1 + original_kdv_orani_item / 100)
                iskonto_farki_per_birim_detay = original_birim_fiyat_kdv_dahil_kalem - iskontolu_birim_fiyat_kdv_dahil
                uygulanan_toplam_iskonto_tutari_detay = iskonto_farki_per_birim_detay * miktar_db

            item_qt = QTreeWidgetItem(self.kalem_tree)
            item_qt.setText(0, str(sira_idx))
            item_qt.setText(1, kalem_item.get('urun_kodu', ''))
            item_qt.setText(2, kalem_item.get('urun_adi', ''))
            item_qt.setText(3, f"{miktar_db:.2f}".rstrip('0').rstrip('.'))
            item_qt.setText(4, self.db._format_currency(iskontolu_birim_fiyat_kdv_dahil))
            item_qt.setText(5, f"%{kalem_item.get('kdv_orani', 0):.0f}")
            item_qt.setText(6, f"{kalem_item.get('iskonto_yuzde_1', 0):.2f}".replace('.', ',').rstrip('0').rstrip('.'))
            item_qt.setText(7, f"{kalem_item.get('iskonto_yuzde_2', 0):.2f}".replace('.', ',').rstrip('0').rstrip('.'))
            item_qt.setText(8, self.db._format_currency(uygulanan_toplam_iskonto_tutari_detay))
            item_qt.setText(9, self.db._format_currency(toplam_dahil_db))
            item_qt.setText(10, self.db._format_currency(kalem_item.get('alis_fiyati_fatura_aninda', 0.0)))
            
            # DEĞİŞİKLİK: Tüm hücrelerin içeriğini merkeze hizala ve fontunu ayarla
            for i in range(self.kalem_tree.columnCount()):
                item_qt.setTextAlignment(i, Qt.AlignCenter)
                item_qt.setFont(i, font_item)
            
            sira_idx += 1


    def clear_layout(self, layout):
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
            else:
                self.clear_layout(item.layout())

class YeniMusteriEklePenceresi(QDialog):
    def __init__(self, parent, db_manager, yenile_callback, musteri_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.app = app_ref
        self.yenile_callback = yenile_callback
        self.musteri_duzenle_data = musteri_duzenle

        self.musteri_duzenle_id = self.musteri_duzenle_data.get('id') if self.musteri_duzenle_data else None

        from hizmetler import CariService
        self.cari_service = CariService(self.db)

        title = "Yeni Müşteri Ekle" if not self.musteri_duzenle_id else "Müşteri Düzenle"
        self.setWindowTitle(title)
        self.setFixedSize(500, 330)  # Değişiklik: Pencere boyutu küçültüldü
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10) # Değişiklik: Marjinleri ayarladık
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter) # Değişiklik: Başlık ortaya hizalandı
        main_layout.addWidget(title_label)

        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)
        
        self.entries = {}
        labels_entries = {
            "Müşteri Kodu:": "entry_kod",
            "Ad Soyad (*):": "entry_ad",
            "Telefon:": "entry_tel",
            "Adres:": "entry_adres",
            "Vergi Dairesi:": "entry_vd",
            "Vergi No:": "entry_vn"
        }

        for i, (label_text, entry_name) in enumerate(labels_entries.items()):
            form_layout.addWidget(QLabel(label_text), i, 0, alignment=Qt.AlignCenter)
            if entry_name == "entry_adres":
                widget = QTextEdit()
                widget.setFixedHeight(80)
                widget.textChanged.connect(lambda w=widget: self._limit_text_length(w, 300)) 
            else:
                widget = QLineEdit()
            
            self.entries[entry_name] = widget
            form_layout.addWidget(widget, i, 1)

        # KRİTİK GÜNCELLEME: QLineEdit'lar için karakter kısıtlamaları
        self.entries["entry_kod"].setMaxLength(50)  # Kod
        self.entries["entry_ad"].setMaxLength(100) # Ad Soyad
        self.entries["entry_tel"].setMaxLength(11)  # Telefon (İstenen kısıtlama)
        self.entries["entry_vd"].setMaxLength(100) # Vergi Dairesi
        self.entries["entry_vn"].setMaxLength(11)  # Vergi No/TCKN
        
        main_layout.addStretch() # Değişiklik: Butonları alta itmek için boşluk ekledik

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()

        self.kaydet_button = QPushButton("Kaydet")
        self.kaydet_button.clicked.connect(self.kaydet)
        button_layout.addWidget(self.kaydet_button)
        
        self.iptal_button = QPushButton("İptal")
        self.iptal_button.clicked.connect(self.reject)
        button_layout.addWidget(self.iptal_button)
        
        if self.musteri_duzenle_id:
            self._mevcut_musteriyi_yukle()
        else:
            # Yeni müşteri: Formu sıfırla ve kodu üret (Hata veren akış düzeltildi)
            self._formu_sifirla()
            kullanici_id = self.app.current_user.get("id") # ID'yi al
            
            # Hata veren doğrudan çağrıyı ID parametresi ile düzelt
            yeni_kod = self.db.get_next_musteri_kodu(kullanici_id=kullanici_id)
            self._oto_kod_uret(yeni_kod) 

    def _limit_text_length(self, text_edit_widget, max_length):
        """QTextEdit için metin uzunluğunu sınırlayan yardımcı metot."""
        if len(text_edit_widget.toPlainText()) > max_length:
            cursor = text_edit_widget.textCursor()
            cursor.movePosition(cursor.End)
            text_edit_widget.setText(text_edit_widget.toPlainText()[:max_length])
            text_edit_widget.setTextCursor(cursor)

    def _oto_kod_uret(self, yeni_kod):
        """Yeni cari kodu oluşturur ve forma yazar."""
        self.entries["entry_kod"].setText(yeni_kod)
        self.entries["entry_kod"].setReadOnly(True)

    def _mevcut_musteriyi_yukle(self):
        """Mevcut müşteri verilerini düzenleme modunda forma yükler."""
        if self.musteri_duzenle_data:
            self.entries["entry_kod"].setText(self.musteri_duzenle_data.get('kod', ''))
            self.entries["entry_ad"].setText(self.musteri_duzenle_data.get('ad', ''))
            self.entries["entry_tel"].setText(self.musteri_duzenle_data.get('telefon', ''))
            self.entries["entry_adres"].setPlainText(self.musteri_duzenle_data.get('adres', ''))
            self.entries["entry_vd"].setText(self.musteri_duzenle_data.get('vergi_dairesi', ''))
            self.entries["entry_vn"].setText(self.musteri_duzenle_data.get('vergi_no', ''))
            self.entries["entry_kod"].setReadOnly(True)

    def _verileri_yukle(self):
        """
        Müşteri verilerini forma yükler (düzenleme modu) VEYA yeni kod üretir (yeni ekleme modu).
        [KRİTİK DÜZELTME 1: Yeni müşteriler için _mevcut_musteriyi_yukle çağrısı _verileri_yukle ile değiştirildi.]
        """
        kullanici_id = self.app.current_user.get("id")
        
        if self.musteri_duzenle_data:
            # Düzenleme modu: Mevcut veriyi yükle (bu zaten doğru çalışıyordu)
            self._mevcut_musteriyi_yukle() 
        else:
            # Yeni ekleme modu: Kodu üret ve ata
            self._formu_sifirla()
            generated_code = self.db.get_next_musteri_kodu(kullanici_id=kullanici_id)
            self._oto_kod_uret(generated_code) 
            self.entries["entry_kod"].setReadOnly(True)

    def _formu_sifirla(self):
        """Formu boşaltır ve varsayılan değerleri atar."""
        self.entries["entry_kod"].clear()
        self.entries["entry_ad"].clear()
        self.entries["entry_tel"].clear()
        self.entries["entry_adres"].clear()
        self.entries["entry_vd"].clear()
        self.entries["entry_vn"].clear()

    def kaydet(self):
        ad = self.entries["entry_ad"].text().strip()
        if not ad:
            QMessageBox.warning(self, "Eksik Bilgi", "Müşteri Adı alanı boş bırakılamaz.")
            return

        data = {
            "ad": ad,
            "kod": self.entries["entry_kod"].text().strip(),
            "telefon": self.entries["entry_tel"].text().strip(),
            "adres": self.entries["entry_adres"].toPlainText().strip(),
            "vergi_dairesi": self.entries["entry_vd"].text().strip(),
            "vergi_no": self.entries["entry_vn"].text().strip()
        }

        try:
            if self.musteri_duzenle_id:
                success, message = self.cari_service.cari_guncelle(self.musteri_duzenle_id, data)
            else:
                success, message = self.cari_service.cari_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Müşteri bilgileri başarıyla kaydedildi.")
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", "Müşteri kaydedilirken bir hata oluştu.")

        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "Hata", f"Müşteri kaydedilirken bir hata oluştu:\n{error_detail}")
            logging.error(f"Müşteri kaydetme hatası: {error_detail}", exc_info=True)

class SiparisDetayPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, siparis_id, yenile_callback=None):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.siparis_id = siparis_id
        self.yenile_callback = yenile_callback
        
        self.siparis_ana = None
        self.siparis_kalemleri_db = None
        
        self.s_no = "Yükleniyor..."
        self.tip = ""

        self.setWindowTitle(f"Sipariş Detayları: {self.s_no}")
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        self.main_layout = QVBoxLayout(self)
        
        self._create_ui_and_populate_data()

        self.finished.connect(self.on_dialog_finished)

    def _create_ui_and_populate_data(self):
        """Arayüzü oluşturur, API'den verileri çeker ve arayüzü doldurur."""
        if self.main_layout.layout():
            self.clear_layout(self.main_layout)

        try:
            # DÜZELTME: siparis_getir_by_id metoduna kullanici_id parametresi eklendi
            self.siparis_ana = self.db.siparis_getir_by_id(self.siparis_id, self.app.current_user[0])
            if not self.siparis_ana:
                raise Exception("Sipariş ana bilgileri API'den alınamadı.")
            
            # DÜZELTME: siparis_kalemleri_al metoduna kullanici_id parametresi eklendi
            self.siparis_kalemleri_db = self.db.siparis_kalemleri_al(self.siparis_id, self.app.current_user[0])
            if not self.siparis_kalemleri_db:
                logging.warning(f"Sipariş ID {self.siparis_id} için sipariş kalemi bulunamadı.")
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Sipariş bilgileri çekilemedi: {e}")
            self.close()
            return
        
        # --- Veri Çekme ve Hazırlama ---
        self.s_no = self.siparis_ana.get('siparis_no', '-')
        durum_db = self.siparis_ana.get('durum', '-')
        self.setWindowTitle(f"Sipariş Detayları: {self.s_no} ({durum_db})")
        
        # Kullanıcı bilgileri
        try:
            # DÜZELTME: kullanici_listele metoduna kullanici_id parametresi eklendi
            kullanicilar_list = self.db.kullanici_listele(self.app.current_user[0])
            kullanicilar_map = {k.get('id'): k.get('kullanici_adi') for k in kullanicilar_list}
        except Exception as e:
            logger.error(f"Kullanıcı listesi API'den alınamadı: {e}")
            kullanicilar_map = {}
        
        olusturan_adi = kullanicilar_map.get(self.siparis_ana.get('olusturan_kullanici_id'), "Bilinmiyor")
        son_guncelleyen_adi = kullanicilar_map.get(self.siparis_ana.get('son_guncelleyen_kullanici_id'), "Bilinmiyor")
        
        # Cari Bilgisi
        cari_adi_text = "Bilinmiyor"
        if self.siparis_ana.get('cari_tip') == 'MUSTERI':
            # DÜZELTME: musteri_getir_by_id metoduna kullanici_id parametresi eklendi
            cari_bilgi = self.db.musteri_getir_by_id(self.siparis_ana.get('cari_id'), self.app.current_user[0])
            cari_adi_text = f"{cari_bilgi.get('ad')}" if cari_bilgi else "Bilinmiyor"
        elif self.siparis_ana.get('cari_tip') == 'TEDARIKCI':
            # DÜZELTME: tedarikci_getir_by_id metoduna kullanici_id parametresi eklendi
            cari_bilgi = self.db.tedarikci_getir_by_id(self.siparis_ana.get('cari_id'), self.app.current_user[0])
            cari_adi_text = f"{cari_bilgi.get('ad')}" if cari_bilgi else "Bilinmiyor"

        # --- Arayüz Oluşturma ---
        font_label = QFont("Segoe UI", 9, QFont.Bold)
        font_value = QFont("Segoe UI", 10)
        font_header = QFont("Segoe UI", 11, QFont.Bold)
        font_groupbox = QFont("Segoe UI", 10, QFont.Bold)

        self.ust_bilgiler_frame = QFrame(self)
        self.ust_bilgiler_layout = QHBoxLayout(self.ust_bilgiler_frame)
        self.ust_bilgiler_layout.setContentsMargins(0, 0, 0, 0)
        self.ust_bilgiler_layout.setSpacing(15)
        self.main_layout.addWidget(self.ust_bilgiler_frame)
        self.ust_bilgiler_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.sol_panel_frame = QFrame(self.ust_bilgiler_frame)
        self.sol_panel_layout = QGridLayout(self.sol_panel_frame)
        self.sol_panel_layout.setContentsMargins(0, 0, 0, 0)
        self.sol_panel_layout.setSpacing(5)
        self.ust_bilgiler_layout.addWidget(self.sol_panel_frame)

        try: siparis_tarihi_formatted = datetime.strptime(str(self.siparis_ana.get('tarih')), '%Y-%m-%d').strftime('%d.%m.%Y')
        except: siparis_tarihi_formatted = str(self.siparis_ana.get('tarih'))
        
        self.sol_panel_layout.addWidget(QLabel("Sipariş No:", font=font_label), 0, 0)
        self.sol_panel_layout.addWidget(QLabel(self.s_no, font=font_value), 0, 1)
        self.sol_panel_layout.addWidget(QLabel("Tarih:", font=font_label), 0, 2)
        self.sol_panel_layout.addWidget(QLabel(siparis_tarihi_formatted, font=font_value), 0, 3)

        cari_label_tipi = "Müşteri:" if self.siparis_ana.get('cari_tip') == self.db.CARI_TIP_MUSTERI else "Tedarikçi:"
        self.sol_panel_layout.addWidget(QLabel(cari_label_tipi, font=font_label), 1, 0)
        self.sol_panel_layout.addWidget(QLabel(cari_adi_text, font=font_value), 1, 1, 1, 3)
        
        try: teslimat_tarihi_formatted = datetime.strptime(str(self.siparis_ana.get('teslimat_tarihi')), '%Y-%m-%d').strftime('%d.%m.%Y')
        except: teslimat_tarihi_formatted = str(self.siparis_ana.get('teslimat_tarihi'))
        self.sol_panel_layout.addWidget(QLabel("Teslimat Tarihi:", font=font_label), 2, 0)
        self.sol_panel_layout.addWidget(QLabel(teslimat_tarihi_formatted, font=font_value), 2, 1)

        self.sol_panel_layout.addWidget(QLabel("Durum:", font=font_label), 2, 2)
        self.sol_panel_layout.addWidget(QLabel(self.siparis_ana.get('durum'), font=font_value), 2, 3)

        genel_iskonto_gosterim_text = "Uygulanmadı"
        genel_iskonto_tipi_db = self.siparis_ana.get('genel_iskonto_tipi')
        genel_iskonto_degeri_db = self.db.safe_float(self.siparis_ana.get('genel_iskonto_degeri'))
        if genel_iskonto_tipi_db == 'YUZDE' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Yüzde %{genel_iskonto_degeri_db:.2f}".replace('.', ',').rstrip('0').rstrip(',')
        elif genel_iskonto_tipi_db == 'TUTAR' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = self.db._format_currency(genel_iskonto_degeri_db)
        
        self.sol_panel_layout.addWidget(QLabel("Genel İsk:", font=font_label), 3, 0)
        self.sol_panel_layout.addWidget(QLabel(genel_iskonto_gosterim_text, font=font_value), 3, 1)

        siparis_notlari_display = QTextEdit()
        siparis_notlari_display.setPlainText(self.siparis_ana.get('siparis_notlari', '-') if self.siparis_ana.get('siparis_notlari') else "")
        siparis_notlari_display.setReadOnly(True)
        siparis_notlari_display.setFixedHeight(60)
        self.sol_panel_layout.addWidget(QLabel("Sipariş Notları:", font=font_label), 4, 0, Qt.AlignTop)
        self.sol_panel_layout.addWidget(siparis_notlari_display, 4, 1, 1, 4)
        
        self.sag_panel_frame = QFrame(self.ust_bilgiler_frame)
        self.sag_panel_layout = QGridLayout(self.sag_panel_frame)
        self.sag_panel_layout.setContentsMargins(0, 0, 0, 0)
        self.sag_panel_layout.setSpacing(5)
        self.ust_bilgiler_layout.addWidget(self.sag_panel_frame)
        self.ust_bilgiler_layout.setStretch(1, 1)
        
        toplam_tutar = self.db.safe_float(self.siparis_ana.get('toplam_tutar'))
        
        self.sag_panel_layout.addWidget(QLabel("Genel Toplam:", font=font_header), 0, 0, Qt.AlignCenter)
        self.lbl_genel_toplam = QLabel(self.db._format_currency(toplam_tutar), font=QFont("Segoe UI", 20, QFont.Bold))
        self.sag_panel_layout.addWidget(self.lbl_genel_toplam, 0, 1, Qt.AlignCenter)
        
        self.sag_panel_layout.addWidget(QLabel("Oluşturan:", font=font_label), 1, 0, Qt.AlignCenter)
        self.sag_panel_layout.addWidget(QLabel(olusturan_adi, font=font_value), 1, 1, Qt.AlignCenter)
        
        self.sag_panel_layout.addWidget(QLabel("Son Güncelleyen:", font=font_label), 2, 0, Qt.AlignCenter)
        self.sag_panel_layout.addWidget(QLabel(son_guncelleyen_adi, font=font_value), 2, 1, Qt.AlignCenter)
        
        self.sag_panel_layout.addWidget(QLabel("Oluşturulma Tarihi:", font=font_label), 3, 0, Qt.AlignCenter)
        self.sag_panel_layout.addWidget(QLabel(str(self.siparis_ana.get('olusturma_tarihi_saat', '-')), font=font_value), 3, 1, Qt.AlignCenter)

        # Sipariş Kalemleri GroupBox
        kalemler_frame = QGroupBox("Sipariş Kalemleri", self)
        kalemler_frame.setFont(font_groupbox)
        kalemler_frame_layout = QVBoxLayout(kalemler_frame)
        self.main_layout.addWidget(kalemler_frame)
        
        cols_kalem = ("Sıra", "Ürün Kodu", "Ürün Adı", "Miktar", "Birim Fiyat", "KDV %", "İsk 1 (%)", "İsk 2 (%)", "Uyg. İsk. Tutarı", "Tutar (Dah.)", "Alış.F (Sipariş Anı)")
        self.kalem_tree = QTreeWidget(kalemler_frame)
        self.kalem_tree.setHeaderLabels(cols_kalem)
        self.kalem_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kalem_tree.setSortingEnabled(True)

        font_header = QFont("Segoe UI", 11, QFont.Bold)
        col_defs_kalem = [
            ("Sıra", 40, Qt.AlignCenter), ("Ürün Kodu", 120, Qt.AlignCenter), ("Ürün Adı", 190, Qt.AlignCenter),
            ("Miktar", 80, Qt.AlignCenter), ("Birim Fiyat", 100, Qt.AlignCenter), ("KDV %", 60, Qt.AlignCenter),
            ("İsk 1 (%)", 90, Qt.AlignCenter), ("İsk 2 (%)", 90, Qt.AlignCenter),
            ("Uyg. İsk. Tutarı", 110, Qt.AlignCenter), ("Tutar (Dah.)", 120, Qt.AlignCenter),
            ("Alış.F (Sipariş Anı)", 130, Qt.AlignCenter)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs_kalem):
            self.kalem_tree.setColumnWidth(i, width)
            self.kalem_tree.headerItem().setTextAlignment(i, alignment)
            self.kalem_tree.headerItem().setFont(i, font_header)
        self.kalem_tree.header().setStretchLastSection(False)
        self.kalem_tree.header().setSectionResizeMode(2, QHeaderView.Stretch)
        
        kalemler_frame_layout.addWidget(self.kalem_tree)
        self._load_siparis_kalemleri_to_treeview(self.siparis_kalemleri_db)

        # Butonlar
        self._butonlari_olustur()
            
    def _load_siparis_kalemleri_to_treeview(self, kalemler_list):
        """API'den gelen sipariş kalemlerini QTreeWidget'a yükler."""
        self.kalem_tree.clear()
        
        font_item = QFont("Segoe UI", 12)

        if not kalemler_list:
            item_qt = QTreeWidgetItem(self.kalem_tree)
            item_qt.setText(2, "Sipariş kalemi bulunamadı.")
            for i in range(self.kalem_tree.columnCount()):
                item_qt.setTextAlignment(i, Qt.AlignCenter)
                item_qt.setFont(i, font_item)
            return

        sira_idx = 1
        for k_db in kalemler_list:
            urun_info = self.db.stok_getir_by_id(k_db.get('urun_id'))
            if not urun_info:
                urun_kodu_db = "Bilinmiyor"
                urun_adi_db = "Bilinmiyor"
            else:
                urun_kodu_db = urun_info.get('kod', 'Bilinmiyor')
                urun_adi_db = urun_info.get('ad', 'Bilinmiyor')

            miktar_gosterim = f"{k_db.get('miktar'):.2f}".rstrip('0').rstrip('.')
            iskontolu_birim_fiyat_kdv_dahil_display = (k_db.get('kalem_toplam_kdv_dahil') / k_db.get('miktar')) if k_db.get('miktar') != 0 else 0.0
            iskonto_yuzde_1_display = f"{k_db.get('iskonto_yuzde_1'):.2f}".replace('.', ',').rstrip('0').rstrip(',')
            iskonto_yuzde_2_display = f"{k_db.get('iskonto_yuzde_2'):.2f}".replace('.', ',').rstrip('0').rstrip(',')

            original_birim_fiyat_kdv_dahil_kalem = k_db.get('birim_fiyat') * (1 + k_db.get('kdv_orani') / 100)
            iskonto_farki_per_birim_detay = original_birim_fiyat_kdv_dahil_kalem - iskontolu_birim_fiyat_kdv_dahil_display
            uygulanan_toplam_iskonto_tutari_detay = iskonto_farki_per_birim_detay * k_db.get('miktar')

            item_qt = QTreeWidgetItem(self.kalem_tree)
            item_qt.setText(0, str(sira_idx))
            item_qt.setText(1, urun_kodu_db)
            item_qt.setText(2, urun_adi_db)
            item_qt.setText(3, miktar_gosterim)
            item_qt.setText(4, self.db._format_currency(iskontolu_birim_fiyat_kdv_dahil_display))
            item_qt.setText(5, f"%{k_db.get('kdv_orani'):.0f}")
            item_qt.setText(6, iskonto_yuzde_1_display)
            item_qt.setText(7, iskonto_yuzde_2_display)
            item_qt.setText(8, self.db._format_currency(uygulanan_toplam_iskonto_tutari_detay))
            item_qt.setText(9, self.db._format_currency(k_db.get('kalem_toplam_kdv_dahil')))
            item_qt.setText(10, self.db._format_currency(k_db.get('alis_fiyati_fatura_aninda', 0.0)))
            
            for i in range(self.kalem_tree.columnCount()):
                item_qt.setTextAlignment(i, Qt.AlignCenter)
                item_qt.setFont(i, font_item)
            
            sira_idx += 1
            
    def _butonlari_olustur(self):
        button_frame_alt = QFrame(self)
        button_frame_alt_layout = QHBoxLayout(button_frame_alt)
        self.main_layout.addWidget(button_frame_alt)

        self.faturaya_donustur_button_detail = QPushButton("Faturaya Dönüştür")
        self.faturaya_donustur_button_detail.clicked.connect(self._faturaya_donustur)
        button_frame_alt_layout.addWidget(self.faturaya_donustur_button_detail)

        btn_siparisi_duzenle = QPushButton("Siparişi Düzenle")
        btn_siparisi_duzenle.clicked.connect(self._siparisi_duzenle)
        button_frame_alt_layout.addWidget(btn_siparisi_duzenle)

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        button_frame_alt_layout.addWidget(btn_kapat)

        if self.siparis_ana.get('fatura_id'):
            self.faturaya_donustur_button_detail.setEnabled(False)
            fatura_no_text = ""
            try:
                fatura_data = self.db.fatura_getir_by_id(self.siparis_ana.get('fatura_id'))
                fatura_no_text = fatura_data.get('fatura_no', '-')
            except Exception:
                fatura_no_text = "Hata"

            lbl_fatura_iliskisi = QLabel(f"Bu sipariş Fatura No: '{fatura_no_text}' ile ilişkilendirilmiştir.")
            lbl_fatura_iliskisi.setStyleSheet("color: blue; font-style: italic;")
            button_frame_alt_layout.addWidget(lbl_fatura_iliskisi)

    def _faturaya_donustur(self):
        """Bu siparişi satış veya alış faturasına dönüştürür."""
        from pencereler import OdemeTuruSecimDialog
        fatura_tipi_for_dialog = 'SATIŞ' if self.siparis_ana.get('cari_tip') == 'MUSTERI' else 'ALIŞ'
        dialog = OdemeTuruSecimDialog(
            self.app,
            self.db,
            fatura_tipi_for_dialog,
            self.siparis_ana.get('cari_id'),
            self.app.current_user[0], # DÜZELTME: kullanici_id parametresi eklendi
            self._faturaya_donustur_on_dialog_confirm
        )
        dialog.exec()

    def _faturaya_donustur_on_dialog_confirm(self, selected_odeme_turu, selected_kasa_banka_id, selected_vade_tarihi):
        if selected_odeme_turu is None:
            self.app.set_status_message("Faturaya dönüştürme iptal edildi (ödeme türü seçilmedi).")
            return
        confirm_msg = (f"'{self.s_no}' numaralı siparişi '{selected_odeme_turu}' ödeme türü ile faturaya dönüştürmek istediğinizden emin misiniz?\n"
                    f"Bu işlem sonucunda yeni bir fatura oluşturulacak ve sipariş durumu güncellenecektir.")
        if selected_odeme_turu == "AÇIK HESAP" and selected_vade_tarihi:
            confirm_msg += f"\nVade Tarihi: {selected_vade_tarihi}"
        if selected_kasa_banka_id:
            try:
                kb_bilgi = self.db.kasa_banka_getir_by_id(selected_kasa_banka_id)
                if kb_bilgi:
                    confirm_msg += f"\nİşlem Kasa/Banka: {kb_bilgi.get('hesap_adi')}"
            except Exception as e:
                logging.error(f"Kasa/Banka bilgisi çekilirken hata: {e}")
                confirm_msg += "\nİşlem Kasa/Banka: Bilgi çekilemedi"

        reply = QMessageBox.question(self, "Faturaya Dönüştür Onayı", confirm_msg, QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.No:
            return
        from hizmetler import FaturaService
        fatura_service = FaturaService(self.db)
        success, message = fatura_service.siparis_faturaya_donustur(
            self.siparis_id,
            self.app.current_user[0] if self.app and hasattr(self.app, 'current_user') and self.app.current_user else None,
            selected_odeme_turu,
            selected_kasa_banka_id,
            selected_vade_tarihi
        )
        if success:
            QMessageBox.information(self, "Başarılı", message)
            self.close()
            if hasattr(self.app, 'siparis_listesi_sayfasi'):
                self.app.siparis_listesi_sayfasi.siparis_listesini_yukle()
            if hasattr(self.app, 'fatura_listesi_sayfasi'):
                if hasattr(self.app.fatura_listesi_sayfasi, 'satis_fatura_frame'):
                    self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
                if hasattr(self.app.fatura_listesi_sayfasi.alis_fatura_frame, 'fatura_listesini_yukle'):
                    self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()
        else:
            QMessageBox.critical(self, "Hata", message)

    def _siparisi_duzenle(self):
        from pencereler import SiparisPenceresi
        siparis_tipi_db = 'SATIŞ_SIPARIS' if self.siparis_ana.get('cari_tip') == 'MUSTERI' else 'ALIŞ_SIPARIS'
        dialog = SiparisPenceresi(
            parent=self.app, 
            db_manager=self.db,
            app_ref=self.app,
            siparis_tipi=siparis_tipi_db,
            siparis_id_duzenle=self.siparis_id,
            yenile_callback=self.yenile_callback
        )
        dialog.exec()
        self.close()

    def on_dialog_finished(self, result):
        if self.yenile_callback:
            self.yenile_callback()
            
    def clear_layout(self, layout):
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
            else:
                self.clear_layout(item.layout())

class YoneticiAyarlariPenceresi(QDialog):
    def __init__(self, parent=None, db_manager=None):
        super().__init__(parent)
        self.db = db_manager # Değişken adını koruyoruz
        self.setWindowTitle("Yönetici Ayarları - Firma Bilgileri")
        self.setMinimumWidth(500)

        self.layout = QGridLayout(self)
        self.layout.addWidget(QLabel("<b>Firma Bilgileri</b>"), 0, 0, 1, 2)

        self.layout.addWidget(QLabel("Firma Ünvanı:"), 1, 0)
        self.unvan_entry = QLineEdit()
        self.layout.addWidget(self.unvan_entry, 1, 1)
        
        self.layout.addWidget(QLabel("Vergi Dairesi:"), 2, 0)
        self.vergi_dairesi_entry = QLineEdit()
        self.layout.addWidget(self.vergi_dairesi_entry, 2, 1)
        
        self.layout.addWidget(QLabel("Vergi No:"), 3, 0)
        self.vergi_no_entry = QLineEdit()
        self.layout.addWidget(self.vergi_no_entry, 3, 1)
        
        self.layout.addWidget(QLabel("Adres:"), 4, 0)
        self.adres_entry = QLineEdit()
        self.layout.addWidget(self.adres_entry, 4, 1)

        # --- YENİ EKLENEN BÖLÜM ---
        # Araya yeni bir ayırıcı ve buton ekliyoruz.
        self.layout.addWidget(QLabel("<hr>"), 5, 0, 1, 2) # Yatay çizgi ile ayırıcı
        
        self.personel_yonetimi_button = QPushButton("Personel Yönetimi")
        self.personel_yonetimi_button.clicked.connect(self._personel_yonetimi_ac)
        # Yeni butonu 6. satıra ekliyoruz.
        self.layout.addWidget(self.personel_yonetimi_button, 6, 0, 1, 2)
        # --- YENİ BÖLÜM SONU ---

        self.kaydet_button = QPushButton("Bilgileri Kaydet")
        self.kaydet_button.clicked.connect(self.bilgileri_kaydet)
        # Mevcut kaydet butonunu bir alt satıra (7. satır) taşıyoruz.
        self.layout.addWidget(self.kaydet_button, 7, 0, 1, 2)

        self._mevcut_bilgileri_yukle()

    def _personel_yonetimi_ac(self):
        """Personel Yönetimi penceresini açar."""
        # db_manager'ı (self.db) yeni pencereye iletiyoruz.
        dialog = PersonelYonetimiPenceresi(self, self.db)
        dialog.exec()

    def _mevcut_bilgileri_yukle(self):
        """Mevcut firma bilgilerini API'den çeker ve form alanlarını doldurur."""
        try:
            success, data = self.db.sirket_bilgilerini_getir()
            if success and data:
                self.unvan_entry.setText(data.get("unvan", ""))
                self.vergi_dairesi_entry.setText(data.get("vergi_dairesi", ""))
                self.vergi_no_entry.setText(data.get("vergi_no", ""))
                self.adres_entry.setText(data.get("adres", ""))
            elif not success:
                QMessageBox.warning(self, "Hata", f"Firma bilgileri yüklenemedi:\n{data}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Firma bilgileri yüklenirken bir hata oluştu:\n{e}")

    def bilgileri_kaydet(self):
        """Formdaki bilgileri API'ye göndererek günceller."""
        guncel_data = {
            "unvan": self.unvan_entry.text(),
            "vergi_dairesi": self.vergi_dairesi_entry.text(),
            "vergi_no": self.vergi_no_entry.text(),
            "adres": self.adres_entry.text()
        }
        
        try:
            success, message = self.db.sirket_bilgilerini_guncelle(guncel_data)
            if success:
                QMessageBox.information(self, "Başarılı", "Firma bilgileri başarıyla güncellendi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", f"Bilgiler güncellenemedi:\n{message}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Bilgiler güncellenirken bir hata oluştu:\n{e}")
                        
class StokHareketiPenceresi(QDialog):
    def __init__(self, parent, db_manager, urun_id, urun_adi, mevcut_stok, hareket_yonu, yenile_callback):
        super().__init__(parent)
        self.db = db_manager
        self.urun_id = urun_id
        self.yenile_callback = yenile_callback
        
        # UI elemanları burada tanımlandığından, init metodunun kalanı aynı kalır.

        title = "Stok Girişi" if hareket_yonu == "EKLE" else "Stok Çıkışı"
        self.setWindowTitle(f"{title}: {urun_adi}")
        self.setMinimumWidth(400)
        self.setModal(True)

        self.main_layout = QVBoxLayout(self)
        self.form_layout = QGridLayout()

        self.main_layout.addWidget(QLabel(f"<b>{title}</b><br>Ürün: {urun_adi}<br>Mevcut Stok: {mevcut_stok:.2f}"), alignment=Qt.AlignCenter)
        self.main_layout.addLayout(self.form_layout)

        self.entries = {}
        self.form_layout.addWidget(QLabel("İşlem Tipi:"), 0, 0)
        self.entries['islem_tipi'] = QComboBox()
        if hareket_yonu == "EKLE": self.entries['islem_tipi'].addItems(["Giriş (Manuel)", "Sayım Fazlası", "İade Girişi"])
        else: self.entries['islem_tipi'].addItems(["Çıkış (Manuel)", "Sayım Eksiği", "Zayiat"])
        self.form_layout.addWidget(self.entries['islem_tipi'], 0, 1)

        self.form_layout.addWidget(QLabel("Miktar:"), 1, 0)
        self.entries['miktar'] = QLineEdit("0,00")
        self.entries['miktar'].setValidator(QDoubleValidator(0.01, 999999.0, 2))
        self.form_layout.addWidget(self.entries['miktar'], 1, 1)

        self.form_layout.addWidget(QLabel("Tarih:"), 2, 0)
        self.entries['tarih'] = QLineEdit(datetime.now().strftime('%Y-%m-%d'))
        self.form_layout.addWidget(self.entries['tarih'], 2, 1)

        self.form_layout.addWidget(QLabel("Açıklama:"), 3, 0, alignment=Qt.AlignTop)
        self.entries['aciklama'] = QTextEdit()
        self.form_layout.addWidget(self.entries['aciklama'], 3, 1)

        button_layout = QHBoxLayout()
        button_layout.addStretch()
        kaydet_button = QPushButton("Kaydet")
        kaydet_button.clicked.connect(self.kaydet)
        iptal_button = QPushButton("İptal")
        iptal_button.clicked.connect(self.reject)
        button_layout.addWidget(kaydet_button)
        button_layout.addWidget(iptal_button)
        self.main_layout.addLayout(button_layout)

    def kaydet(self):
        try:
            miktar = float(self.entries['miktar'].text().replace(',', '.'))
            if miktar <= 0: raise ValueError("Miktar pozitif bir değer olmalıdır.")
        except (ValueError, TypeError):
            QMessageBox.warning(self, "Geçersiz Değer", "Lütfen miktar alanına geçerli bir sayı girin.")
            return

        islem_tipi = self.entries['islem_tipi'].currentText()
        if islem_tipi == "Giriş (Manuel)":
            islem_tipi = self.db.STOK_ISLEM_TIP_GIRIS_MANUEL
        elif islem_tipi == "Çıkış (Manuel)":
            islem_tipi = self.db.STOK_ISLEM_TIP_CIKIS_MANUEL
        elif islem_tipi == "Sayım Fazlası":
            islem_tipi = self.db.STOK_ISLEM_TIP_SAYIM_FAZLASI
        elif islem_tipi == "Sayım Eksiği":
            islem_tipi = self.db.STOK_ISLEM_TIP_SAYIM_EKSIGI
        elif islem_tipi == "İade Girişi":
            islem_tipi = self.db.STOK_ISLEM_TIP_IADE_GIRIS
        elif islem_tipi == "Zayiat":
            islem_tipi = self.db.STOK_ISLEM_TIP_ZAYIAT

        data = {
            "islem_tipi": islem_tipi,
            "miktar": miktar, 
            "tarih": self.entries['tarih'].text(),
            "aciklama": self.entries['aciklama'].toPlainText().strip()
        }

        try:
            success, message = self.db.stok_hareket_ekle(self.urun_id, data)
            if success:
                QMessageBox.information(self, "Başarılı", message)
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "API Hatası", f"Stok hareketi kaydedilirken bir hata oluştu:\n{error_detail}")

class IlgiliFaturalarDetayPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, urun_id, urun_adi):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.urun_id = urun_id
        self.urun_adi = urun_adi
        self.setWindowTitle(f"{self.urun_adi} - İlgili Faturalar")
        self.setMinimumSize(1000, 600)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel(f"{self.urun_adi} Ürününün Yer Aldığı Faturalar")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        filter_frame = QFrame(self)
        filter_layout = QHBoxLayout(filter_frame)
        main_layout.addWidget(filter_frame)

        filter_layout.addWidget(QLabel("Fatura Tipi:"))
        self.fatura_tipi_filter_cb = QComboBox()
        self.fatura_tipi_filter_cb.addItems(["TÜMÜ", "ALIŞ", "SATIŞ"])
        self.fatura_tipi_filter_cb.currentIndexChanged.connect(self._load_ilgili_faturalar)
        filter_layout.addWidget(self.fatura_tipi_filter_cb)
        filter_layout.addStretch() # Sağa yaslama için

        # Filtreleme butonu kaldırıldı, combobox değişince tetikleniyor.
        # btn_filter = QPushButton("Filtrele")
        # btn_filter.clicked.connect(self._load_ilgili_faturalar)
        # filter_layout.addWidget(btn_filter)

        cols_fatura = ("ID", "Fatura No", "Tarih", "Tip", "Cari/Misafir", "KDV Hariç Top.", "KDV Dahil Top.")
        self.ilgili_faturalar_tree = QTreeWidget(self)
        self.ilgili_faturalar_tree.setHeaderLabels(cols_fatura)
        self.ilgili_faturalar_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ilgili_faturalar_tree.setSortingEnabled(True)

        from PySide6.QtWidgets import QHeaderView
        col_defs_fatura = [
            ("ID", 40, Qt.AlignCenter), # Sağa hizala
            ("Fatura No", 120, Qt.AlignCenter),
            ("Tarih", 85, Qt.AlignCenter),
            ("Tip", 70, Qt.AlignCenter),
            ("Cari/Misafir", 200, Qt.AlignCenter),
            ("KDV Hariç Top.", 120, Qt.AlignCenter),
            ("KDV Dahil Top.", 120, Qt.AlignCenter)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs_fatura):
            self.ilgili_faturalar_tree.setColumnWidth(i, width)
            self.ilgili_faturalar_tree.headerItem().setTextAlignment(i, alignment)
            self.ilgili_faturalar_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))
        self.ilgili_faturalar_tree.header().setStretchLastSection(False)
        self.ilgili_faturalar_tree.header().setSectionResizeMode(4, QHeaderView.Stretch) # Cari/Misafir genişlesin

        main_layout.addWidget(self.ilgili_faturalar_tree)

        self.ilgili_faturalar_tree.itemDoubleClicked.connect(self._on_fatura_double_click)

        self._load_ilgili_faturalar() # İlk yükleme

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignCenter)

    def _load_ilgili_faturalar(self, index=None): # index parametresi QComboBox'tan gelir, kullanılmıyor
        self.ilgili_faturalar_tree.clear()

        if not self.urun_id:
            item_qt = QTreeWidgetItem(self.ilgili_faturalar_tree)
            item_qt.setText(4, "Ürün seçili değil.")
            return

        fatura_tipi_filtre = self.fatura_tipi_filter_cb.currentText()
        if fatura_tipi_filtre == "TÜMÜ":
            fatura_tipi_filtre = None # API'ye tüm tipleri çekmesi için None gönder

        # API'den veri çek
        try:
            params = {'urun_id': self.urun_id}
            if fatura_tipi_filtre:
                params['fatura_tipi'] = fatura_tipi_filtre

            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            # API endpoint'i: /faturalar/urun_faturalari şeklinde olmalı
            response_data = self.db.get_urun_faturalari(params.get('urun_id'), params.get('fatura_tipi'))

            faturalar = []
            if isinstance(response_data, dict) and "items" in response_data:
                faturalar = response_data["items"]
            elif isinstance(response_data, list): # Eğer API doğrudan liste dönüyorsa
                faturalar = response_data
                self.app.set_status_message("Uyarı: İlgili faturalar API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else: # Beklenmeyen bir format gelirse
                self.app.set_status_message("Hata: İlgili faturalar API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"İlgili faturalar API'den beklenen formatta gelmedi: {type(response_data)} - {response_data}")
                return # Hata durumunda fonksiyonu sonlandır

            if not faturalar:
                item_qt = QTreeWidgetItem(self.ilgili_faturalar_tree)
                item_qt.setText(4, "Bu ürüne ait fatura bulunamadı.")
                return

            for fatura_item in faturalar:
                item_qt = QTreeWidgetItem(self.ilgili_faturalar_tree)

                fatura_id = fatura_item.get('id')
                fatura_no = fatura_item.get('fatura_no')
                tarih_str = fatura_item.get('tarih')
                fatura_tip = fatura_item.get('fatura_turu') # 'tip' yerine 'fatura_turu' kullanıldı
                cari_adi = fatura_item.get('cari_adi') # API'den gelmesi beklenir
                misafir_adi = fatura_item.get('misafir_adi') # API'den gelmesi beklenir
                toplam_kdv_haric = fatura_item.get('toplam_kdv_haric')
                toplam_kdv_dahil = fatura_item.get('toplam_kdv_dahil')

                try:
                    formatted_tarih = datetime.strptime(tarih_str, '%Y-%m-%d').strftime('%d.%m.%Y')
                except ValueError:
                    formatted_tarih = tarih_str

                display_cari_info = cari_adi
                if fatura_tip == self.db.FATURA_TIP_SATIS and misafir_adi: # 'SATIŞ' sabiti kullanıldı
                    display_cari_info = f"Perakende ({misafir_adi})"

                item_qt.setText(0, str(fatura_id))
                item_qt.setText(1, fatura_no)
                item_qt.setText(2, formatted_tarih)
                item_qt.setText(3, fatura_tip)
                item_qt.setText(4, display_cari_info)
                item_qt.setText(5, self.db._format_currency(toplam_kdv_haric))
                item_qt.setText(6, self.db._format_currency(toplam_kdv_dahil))

                self.app.set_status_message(f"Ürün '{self.urun_adi}' için {len(faturalar)} fatura listelendi.")

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"İlgili faturalar çekilirken hata: {e}")
            logging.error(f"İlgili faturalar yükleme hatası: {e}")

    def _on_fatura_double_click(self, item, column): # item and column from QTreeWidget signal
        fatura_id = item.text(0) # ID ilk sütunda
        if fatura_id:
            from pencereler import FaturaDetayPenceresi
            FaturaDetayPenceresi(self.app, self.db, int(fatura_id)).exec() # fatura_id int olmalı

class KategoriMarkaYonetimiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, refresh_callback=None):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.refresh_callback = refresh_callback
        self.setWindowTitle("Kategori & Marka Yönetimi")
        self.setMinimumSize(800, 500)
        self.setModal(True)
        main_layout = QVBoxLayout(self)
        title_label = QLabel("Kategori & Marka Yönetimi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        main_frame = QWidget(self)
        main_frame_layout = QHBoxLayout(main_frame)
        main_layout.addWidget(main_frame)
        main_frame_layout.setStretch(0, 1)
        main_frame_layout.setStretch(1, 1)
        kategori_frame = QGroupBox("Kategori Yönetimi", main_frame)
        kategori_frame_layout = QGridLayout(kategori_frame)
        main_frame_layout.addWidget(kategori_frame)
        kategori_frame_layout.setColumnStretch(1, 1)
        kategori_frame_layout.addWidget(QLabel("Kategori Adı:"), 0, 0)
        self.kategori_entry = QLineEdit()
        kategori_frame_layout.addWidget(self.kategori_entry, 0, 1)
        kategori_frame_layout.addWidget(QPushButton("Ekle", clicked=self._kategori_ekle_ui), 0, 2)
        kategori_frame_layout.addWidget(QPushButton("Güncelle", clicked=self._kategori_guncelle_ui), 0, 3)
        kategori_frame_layout.addWidget(QPushButton("Sil", clicked=self._kategori_sil_ui), 0, 4)
        self.kategori_tree = QTreeWidget()
        self.kategori_tree.setHeaderLabels(["ID", "Kategori Adı"])
        self.kategori_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kategori_tree.setColumnWidth(0, 50)
        self.kategori_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        self.kategori_tree.itemSelectionChanged.connect(self._on_kategori_select)
        self.kategori_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.kategori_tree.customContextMenuRequested.connect(self._open_kategori_context_menu)
        kategori_frame_layout.addWidget(self.kategori_tree, 1, 0, 1, 5)
        marka_frame = QGroupBox("Marka Yönetimi", main_frame)
        marka_frame_layout = QGridLayout(marka_frame)
        main_frame_layout.addWidget(marka_frame)
        marka_frame_layout.setColumnStretch(1, 1)
        marka_frame_layout.addWidget(QLabel("Marka Adı:"), 0, 0)
        self.marka_entry = QLineEdit()
        marka_frame_layout.addWidget(self.marka_entry, 0, 1)
        marka_frame_layout.addWidget(QPushButton("Ekle", clicked=self._marka_ekle_ui), 0, 2)
        marka_frame_layout.addWidget(QPushButton("Güncelle", clicked=self._marka_guncelle_ui), 0, 3)
        marka_frame_layout.addWidget(QPushButton("Sil", clicked=self._marka_sil_ui), 0, 4)
        self.marka_tree = QTreeWidget()
        self.marka_tree.setHeaderLabels(["ID", "Marka Adı"])
        self.marka_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.marka_tree.setColumnWidth(0, 50)
        self.marka_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        self.marka_tree.itemSelectionChanged.connect(self._on_marka_select)
        self.marka_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.marka_tree.customContextMenuRequested.connect(self._open_marka_context_menu)
        marka_frame_layout.addWidget(self.marka_tree, 1, 0, 1, 5)
        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self._on_close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignCenter)
        self._kategori_listesini_yukle()
        self._marka_listesini_yukle()
        
    def _on_close(self):
        if self.refresh_callback:
            self.refresh_callback()
        self.close()

    def _kategori_listesini_yukle(self):
        self.kategori_tree.clear()
        try:
            kategoriler_response = self.db.kategori_listele()
            kategoriler_list = kategoriler_response.get("items", [])
            for kat_item in kategoriler_list:
                item_qt = QTreeWidgetItem(self.kategori_tree)
                item_qt.setText(0, str(kat_item.get('id')))
                item_qt.setText(1, kat_item.get('ad'))
                item_qt.setData(0, Qt.UserRole, kat_item.get('id'))
            self.kategori_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Kategori listesi çekilirken hata: {e}")
            logging.error(f"Kategori listesi yükleme hatası: {e}", exc_info=True)

    def _on_kategori_select(self):
        selected_items = self.kategori_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.kategori_entry.setText(values)
        else:
            self.kategori_entry.clear()

    def _kategori_ekle_ui(self):
        kategori_adi = self.kategori_entry.text().strip()
        if not kategori_adi:
            QMessageBox.warning(self.app, "Uyarı", "Kategori adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            success, message = self.db.nitelik_ekle("kategoriler", {"ad": kategori_adi})
            
            if success:
                QMessageBox.information(self.app, "Başarılı", f"'{kategori_adi}' kategorisi başarıyla eklendi.")
                self.kategori_entry.clear()
                self._kategori_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(f"'{kategori_adi}' kategorisi başarıyla eklendi.")
            else:
                QMessageBox.critical(self.app, "Hata", message)
                self.app.set_status_message(f"Kategori ekleme başarısız: {message}")
        except Exception as e:
            QMessageBox.critical(self.app, "Beklenmeyen Hata", f"Kategori eklenirken beklenmeyen bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Kategori eklenirken hata: {e}")

    def _kategori_guncelle_ui(self):
        selected_items = self.kategori_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir kategori seçin.")
            return

        selected_item = selected_items[0]
        kategori_id = selected_item.data(0, Qt.UserRole)
        yeni_kategori_adi = self.kategori_entry.text().strip()

        if not yeni_kategori_adi:
            QMessageBox.warning(self.app, "Uyarı", "Kategori adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            success, message = self.db.nitelik_guncelle("kategoriler", kategori_id, {"ad": yeni_kategori_adi})
            
            if success:
                QMessageBox.information(self.app, "Başarılı", f"'{yeni_kategori_adi}' kategorisi başarıyla güncellendi.")
                self.kategori_entry.clear()
                self._kategori_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(f"'{yeni_kategori_adi}' kategorisi başarıyla güncellendi.")
            else:
                QMessageBox.critical(self.app, "Hata", message)
                self.app.set_status_message(f"Kategori güncelleme başarısız: {message}")
        except Exception as e:
            QMessageBox.critical(self.app, "Beklenmeyen Hata", f"Kategori güncellenirken beklenmeyen bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Kategori güncellenirken hata: {e}")

    def _kategori_sil_ui(self):
        selected_items = self.kategori_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir kategori seçin.")
            return

        selected_item = selected_items[0]
        kategori_id = selected_item.data(0, Qt.UserRole)
        kategori_adi = selected_item.text(1)

        reply = QMessageBox.question(self.app, "Onay", f"'{kategori_adi}' kategorisini silmek istediğinizden emin misiniz?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            try:
                # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
                success, message = self.db.nitelik_sil("kategoriler", kategori_id)
                
                if success:
                    QMessageBox.information(self.app, "Başarılı", f"'{kategori_adi}' kategorisi başarıyla silindi.")
                    self.kategori_entry.clear()
                    self._kategori_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                    self.app.set_status_message(f"'{kategori_adi}' kategorisi başarıyla silindi.")
                else:
                    QMessageBox.critical(self.app, "Hata", message)
                    self.app.set_status_message(f"Kategori silme başarısız: {message}")
            except Exception as e:
                QMessageBox.critical(self.app, "Beklenmeyen Hata", f"Kategori silinirken beklenmeyen bir hata oluştu:\n{e}")
                self.app.set_status_message(f"Kategori silinirken hata: {e}")

    def _open_kategori_context_menu(self, pos):
        item = self.kategori_tree.itemAt(pos)
        if not item: return
        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._kategori_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._kategori_sil_ui)
        context_menu.exec(self.kategori_tree.mapToGlobal(pos))

    def _marka_listesini_yukle(self):
        self.marka_tree.clear()
        try:
            markalar_response = self.db.marka_listele()
            markalar = markalar_response.get("items", [])
            for mar in markalar:
                item_qt = QTreeWidgetItem(self.marka_tree)
                item_qt.setText(0, str(mar.get('id')))
                item_qt.setText(1, mar.get('ad'))
                item_qt.setData(0, Qt.UserRole, mar.get('id'))
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Marka listesi çekilirken hata: {e}")
            logging.error(f"Marka listesi yükleme hatası: {e}")

    def _on_marka_select(self):
        selected_items = self.marka_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.marka_entry.setText(values)
        else:
            self.marka_entry.clear()

    def _marka_ekle_ui(self):
        marka_adi = self.marka_entry.text().strip()
        if not marka_adi:
            QMessageBox.warning(self.app, "Uyarı", "Marka adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            success, message = self.db.nitelik_ekle("markalar", {"ad": marka_adi})
            
            if success:
                QMessageBox.information(self.app, "Başarılı", f"'{marka_adi}' markası başarıyla eklendi.")
                self.marka_entry.clear()
                self._marka_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(f"'{marka_adi}' markası başarıyla eklendi.")
            else:
                QMessageBox.critical(self.app, "Hata", message)
                self.app.set_status_message(f"Marka ekleme başarısız: {message}")
        except Exception as e:
            QMessageBox.critical(self.app, "Beklenmeyen Hata", f"Marka eklenirken beklenmeyen bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Marka eklenirken hata: {e}")

    def _marka_guncelle_ui(self):
        selected_items = self.marka_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir marka seçin.")
            return

        selected_item = selected_items[0]
        marka_id = selected_item.data(0, Qt.UserRole)
        yeni_marka_adi = self.marka_entry.text().strip()

        if not yeni_marka_adi:
            QMessageBox.warning(self.app, "Uyarı", "Marka adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            success, message = self.db.nitelik_guncelle("markalar", marka_id, {"ad": yeni_marka_adi})

            if success:
                QMessageBox.information(self.app, "Başarılı", f"'{yeni_marka_adi}' markası başarıyla güncellendi.")
                self.marka_entry.clear()
                self._marka_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(f"'{yeni_marka_adi}' markası başarıyla güncellendi.")
            else:
                QMessageBox.critical(self.app, "Hata", message)
                self.app.set_status_message(f"Marka güncelleme başarısız: {message}")
        except Exception as e:
            QMessageBox.critical(self.app, "Beklenmeyen Hata", f"Marka güncellenirken beklenmeyen bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Marka güncellenirken hata: {e}")

    def _marka_sil_ui(self):
        selected_items = self.marka_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir marka seçin.")
            return

        selected_item = selected_items[0]
        marka_id = selected_item.data(0, Qt.UserRole)
        marka_adi = selected_item.text(1)

        reply = QMessageBox.question(self.app, "Onay", f"'{marka_adi}' markasını silmek istediğinizden emin misiniz?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            try:
                # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
                success, message = self.db.nitelik_sil("markalar", marka_id)
                
                if success:
                    QMessageBox.information(self.app, "Başarılı", f"'{marka_adi}' markası başarıyla silindi.")
                    self.marka_entry.clear()
                    self._marka_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                    self.app.set_status_message(f"'{marka_adi}' markası başarıyla silindi.")
                else:
                    QMessageBox.critical(self.app, "Hata", message)
                    self.app.set_status_message(f"Marka silme başarısız: {message}")
            except Exception as e:
                QMessageBox.critical(self.app, "Beklenmeyen Hata", f"Marka silinirken beklenmeyen bir hata oluştu:\n{e}")
                self.app.set_status_message(f"Marka silinirken hata: {e}")

    def _open_marka_context_menu(self, pos):
        item = self.marka_tree.itemAt(pos)
        if not item: return
        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._marka_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._marka_sil_ui)
        context_menu.exec(self.marka_tree.mapToGlobal(pos))

class UrunNitelikYonetimiPenceresi(QDialog):
    def __init__(self, parent_notebook, db_manager, app_ref, refresh_callback=None):
        super().__init__(parent_notebook)
        self.db = db_manager
        self.app = app_ref
        self.refresh_callback = refresh_callback

        self.setWindowTitle("Ürün Grubu, Birimi ve Menşe Ülke Yönetimi")
        self.setMinimumSize(800, 600)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Ürün Grubu, Birimi ve Menşe Ülke Yönetimi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # Ana içerik çerçevesi (grid gibi düzenleme için)
        main_frame = QWidget(self)
        main_frame_layout = QGridLayout(main_frame)
        main_layout.addWidget(main_frame)
        main_frame_layout.setColumnStretch(0, 1)
        main_frame_layout.setColumnStretch(1, 1)
        main_frame_layout.setRowStretch(0, 1)
        main_frame_layout.setRowStretch(1, 1)


        # --- Ürün Grubu Yönetimi ---
        urun_grubu_frame = QGroupBox("Ürün Grubu Yönetimi", main_frame)
        urun_grubu_frame_layout = QGridLayout(urun_grubu_frame)
        main_frame_layout.addWidget(urun_grubu_frame, 0, 0)
        urun_grubu_frame_layout.setColumnStretch(1, 1)

        urun_grubu_frame_layout.addWidget(QLabel("Grup Adı:"), 0, 0)
        self.urun_grubu_entry = QLineEdit()
        urun_grubu_frame_layout.addWidget(self.urun_grubu_entry, 0, 1)
        urun_grubu_frame_layout.addWidget(QPushButton("Ekle", clicked=self._urun_grubu_ekle_ui), 0, 2)
        urun_grubu_frame_layout.addWidget(QPushButton("Sil", clicked=self._urun_grubu_sil_ui), 0, 3)

        self.urun_grubu_tree = QTreeWidget()
        self.urun_grubu_tree.setHeaderLabels(["ID", "Grup Adı"])
        self.urun_grubu_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.urun_grubu_tree.setColumnWidth(0, 50)
        self.urun_grubu_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        urun_grubu_frame_layout.addWidget(self.urun_grubu_tree, 1, 0, 1, 4)
        self.urun_grubu_tree.itemSelectionChanged.connect(self._on_urun_grubu_select)
        
        self.urun_grubu_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.urun_grubu_tree.customContextMenuRequested.connect(self._open_urun_grubu_context_menu) 
        self._urun_grubu_listesini_yukle()

        # --- Ürün Birimi Yönetimi ---
        urun_birimi_frame = QGroupBox("Ürün Birimi Yönetimi", main_frame)
        urun_birimi_frame_layout = QGridLayout(urun_birimi_frame)
        main_frame_layout.addWidget(urun_birimi_frame, 0, 1)
        urun_birimi_frame_layout.setColumnStretch(1, 1)

        urun_birimi_frame_layout.addWidget(QLabel("Birim Adı:"), 0, 0)
        self.urun_birimi_entry = QLineEdit()
        urun_birimi_frame_layout.addWidget(self.urun_birimi_entry, 0, 1)
        urun_birimi_frame_layout.addWidget(QPushButton("Ekle", clicked=self._urun_birimi_ekle_ui), 0, 2)
        urun_birimi_frame_layout.addWidget(QPushButton("Sil", clicked=self._urun_birimi_sil_ui), 0, 3)

        self.urun_birimi_tree = QTreeWidget()
        self.urun_birimi_tree.setHeaderLabels(["ID", "Birim Adı"])
        self.urun_birimi_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.urun_birimi_tree.setColumnWidth(0, 50)
        self.urun_birimi_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        urun_birimi_frame_layout.addWidget(self.urun_birimi_tree, 1, 0, 1, 4)
        self.urun_birimi_tree.itemSelectionChanged.connect(self._on_urun_birimi_select)
        
        self.urun_birimi_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.urun_birimi_tree.customContextMenuRequested.connect(self._open_birim_context_menu)
        self._urun_birimi_listesini_yukle()

        # --- Ülke (Menşe) Yönetimi ---
        ulke_frame = QGroupBox("Menşe Ülke Yönetimi", main_frame)
        ulke_frame_layout = QGridLayout(ulke_frame)
        main_frame_layout.addWidget(ulke_frame, 1, 0, 1, 2) # İki sütuna yay
        ulke_frame_layout.setColumnStretch(1, 1)

        ulke_frame_layout.addWidget(QLabel("Ülke Adı:"), 0, 0)
        self.ulke_entry = QLineEdit()
        ulke_frame_layout.addWidget(self.ulke_entry, 0, 1)
        ulke_frame_layout.addWidget(QPushButton("Ekle", clicked=self._ulke_ekle_ui), 0, 2)
        ulke_frame_layout.addWidget(QPushButton("Sil", clicked=self._ulke_sil_ui), 0, 3)

        self.ulke_tree = QTreeWidget()
        self.ulke_tree.setHeaderLabels(["ID", "Ülke Adı"])
        self.ulke_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ulke_tree.setColumnWidth(0, 50)
        self.ulke_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        ulke_frame_layout.addWidget(self.ulke_tree, 1, 0, 1, 4)
        self.ulke_tree.itemSelectionChanged.connect(self._on_ulke_select)
        
        self.ulke_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ulke_tree.customContextMenuRequested.connect(self._open_ulke_context_menu)
        self._ulke_listesini_yukle()

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self._on_close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignCenter)

    def _on_close(self):
        if self.refresh_callback:
            self.refresh_callback() # Ürün kartı combobox'larını yenile
        self.close()

    # Ürün Grubu Yönetimi Metotları
    def _urun_grubu_listesini_yukle(self):
        self.urun_grubu_tree.clear()
        try:
            urun_gruplari_response = self.db.urun_grubu_listele() # API'den gelen tam yanıt
            urun_gruplari_list = urun_gruplari_response.get("items", []) # "items" listesini alıyoruz

            for grup_item in urun_gruplari_list: # urun_gruplari_list üzerinde döngü
                item_qt = QTreeWidgetItem(self.urun_grubu_tree)
                item_qt.setText(0, str(grup_item.get('id'))) # .get() ile güvenli erişim
                item_qt.setText(1, grup_item.get('ad')) # .get() ile güvenli erişim
                item_qt.setData(0, Qt.UserRole, grup_item.get('id'))
            self.urun_grubu_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Ürün grubu listesi çekilirken hata: {e}")
            logging.error(f"Ürün grubu listesi yükleme hatası: {e}", exc_info=True)

    def _on_urun_grubu_select(self):
        selected_items = self.urun_grubu_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.urun_grubu_entry.setText(values)
        else:
            self.urun_grubu_entry.clear()

    def _urun_grubu_ekle_ui(self):
        grup_adi = self.urun_grubu_entry.text().strip()
        if not grup_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ürün grubu adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("urun_gruplari", {"ad": grup_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_grubu_entry.clear()
                self._urun_grubu_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün grubu eklenirken hata: {e}")
            logging.error(f"Ürün grubu eklenirken hata: {e}", exc_info=True)

    def _urun_grubu_guncelle_ui(self):
        selected_items = self.urun_grubu_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir ürün grubu seçin.")
            return
        grup_id = selected_items[0].data(0, Qt.UserRole)
        yeni_grup_adi = self.urun_grubu_entry.text().strip()
        if not yeni_grup_adi:
            QMessageBox.warning(self.app, "Uyarı", "Grup adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("urun_gruplari", grup_id, {"ad": yeni_grup_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_grubu_entry.clear()
                self._urun_grubu_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün grubu güncellenirken hata: {e}")
            logging.error(f"Ürün grubu güncellenirken hata: {e}", exc_info=True)

    def _urun_grubu_sil_ui(self):
        selected_items = self.urun_grubu_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir ürün grubu seçin.")
            return
        grup_id = selected_items[0].data(0, Qt.UserRole)
        grup_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{grup_adi}' ürün grubunu silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("urun_gruplari", grup_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.urun_grubu_entry.clear()
                    self._urun_grubu_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                else:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Ürün grubu silinirken hata: {e}")
                logging.error(f"Ürün grubu silinirken hata: {e}", exc_info=True)
                
    # Ürün Birimi Yönetimi Metotları
    def _urun_birimi_listesini_yukle(self):
        self.urun_birimi_tree.clear()
        try:
            urun_birimleri_response = self.db.urun_birimi_listele()
            urun_birimleri_list = urun_birimleri_response.get("items", []) # "items" listesini alıyoruz

            for birim_item in urun_birimleri_list: # urun_birimleri_list üzerinde döngü
                item_qt = QTreeWidgetItem(self.urun_birimi_tree)
                item_qt.setText(0, str(birim_item.get('id'))) # .get() ile güvenli erişim
                item_qt.setText(1, birim_item.get('ad')) # .get() ile güvenli erişim
                item_qt.setData(0, Qt.UserRole, birim_item.get('id'))
            self.urun_birimi_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Ürün birimi listesi çekilirken hata: {e}")
            logging.error(f"Ürün birimi listesi yükleme hatası: {e}", exc_info=True)

    def _on_urun_birimi_select(self):
        selected_items = self.urun_birimi_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.urun_birimi_entry.setText(values)
        else:
            self.urun_birimi_entry.clear()

    def _urun_birimi_ekle_ui(self):
        birim_adi = self.urun_birimi_entry.text().strip()
        if not birim_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ürün birimi adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("urun_birimleri", {"ad": birim_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_birimi_entry.clear()
                self._urun_birimi_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün birimi eklenirken hata: {e}")
            logging.error(f"Ürün birimi eklenirken hata: {e}", exc_info=True)

    def _urun_birimi_guncelle_ui(self):
        selected_items = self.urun_birimi_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir ürün birimi seçin.")
            return
        birim_id = selected_items[0].data(0, Qt.UserRole)
        yeni_birim_adi = self.urun_birimi_entry.text().strip()
        if not yeni_birim_adi:
            QMessageBox.warning(self.app, "Uyarı", "Birim adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("urun_birimleri", birim_id, {"ad": yeni_birim_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_birimi_entry.clear()
                self._urun_birimi_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün birimi güncellenirken hata: {e}")
            logging.error(f"Ürün birimi güncellenirken hata: {e}", exc_info=True)

    def _urun_birimi_sil_ui(self):
        selected_items = self.urun_birimi_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir ürün birimi seçin.")
            return
        birim_id = selected_items[0].data(0, Qt.UserRole)
        birim_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{birim_adi}' ürün birimini silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("urun_birimleri", birim_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.urun_birimi_entry.clear()
                    self._urun_birimi_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                elif message:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Ürün birimi silinirken hata: {e}")
                logging.error(f"Ürün birimi silinirken hata: {e}", exc_info=True)

    def _open_urun_grubu_context_menu(self, pos):
        item = self.urun_grubu_tree.itemAt(pos)
        if not item: return

        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._urun_grubu_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._urun_grubu_sil_ui)
        context_menu.exec(self.urun_grubu_tree.mapToGlobal(pos))

    def _open_birim_context_menu(self, pos):
        item = self.urun_birimi_tree.itemAt(pos)
        if not item: return

        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._urun_birimi_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._urun_birimi_sil_ui)
        context_menu.exec(self.urun_birimi_tree.mapToGlobal(pos))

    def _open_ulke_context_menu(self, pos):
        item = self.ulke_tree.itemAt(pos)
        if not item: return

        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._ulke_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._ulke_sil_ui)
        context_menu.exec(self.ulke_tree.mapToGlobal(pos))

    # Ülke (Menşe) Yönetimi Metotları
    def _ulke_listesini_yukle(self):
        self.ulke_tree.clear()
        try:
            ulkeler_response = self.db.ulke_listele()
            ulkeler_list = ulkeler_response.get("items", []) # "items" listesini alıyoruz

            for ulke_item in ulkeler_list: # ulkeler_list üzerinde döngü
                item_qt = QTreeWidgetItem(self.ulke_tree)
                item_qt.setText(0, str(ulke_item.get('id'))) # .get() ile güvenli erişim
                item_qt.setText(1, ulke_item.get('ad')) # .get() ile güvenli erişim
                item_qt.setData(0, Qt.UserRole, ulke_item.get('id'))
            self.ulke_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Ülke listesi çekilirken hata: {e}")
            logging.error(f"Ülke listesi yükleme hatası: {e}", exc_info=True)
            
    def _on_ulke_select(self):
        selected_items = self.ulke_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.ulke_entry.setText(values)
        else:
            self.ulke_entry.clear()

    def _ulke_ekle_ui(self):
        ulke_adi = self.ulke_entry.text().strip()
        if not ulke_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ülke adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("ulkeler", {"ad": ulke_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.ulke_entry.clear()
                self._ulke_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ülke eklenirken hata: {e}")
            logging.error(f"Ülke eklenirken hata: {e}", exc_info=True)
            
    def _ulke_guncelle_ui(self):
        selected_items = self.ulke_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir ülke seçin.")
            return
        ulke_id = selected_items[0].data(0, Qt.UserRole)
        yeni_ulke_adi = self.ulke_entry.text().strip()
        if not yeni_ulke_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ülke adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("ulkeler", ulke_id, {"ad": yeni_ulke_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.ulke_entry.clear()
                self._ulke_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ülke güncellenirken hata: {e}")
            logging.error(f"Ülke güncellenirken hata: {e}", exc_info=True)

    def _ulke_sil_ui(self):
        selected_items = self.ulke_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir ülke seçin.")
            return
        ulke_id = selected_items[0].data(0, Qt.UserRole)
        ulke_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{ulke_adi}' ülkesini silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("ulkeler", ulke_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.ulke_entry.clear()
                    self._ulke_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                else:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Ülke silinirken hata: {e}")
                logging.error(f"Ülke silinirken hata: {e}", exc_info=True)
                
    def _yukle_kategori_marka_comboboxlari(self):
        # Kategoriler
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            kategoriler_response = self.db.kategori_listele() # <-- BURASI GÜNCELLENDİ
            kategoriler = kategoriler_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.kategoriler_map = {"Seçim Yok": None}
            # kategori_display_values artık kullanılmadığı için kaldırıldı.
            for k in kategoriler:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.kategoriler_map[k.get('ad')] = k.get('id') # <-- 'kategori_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Kategoriler combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Kategoriler yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

        # Markalar
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            markalar_response = self.db.marka_listele() # <-- BURASI GÜNCELLENDİ
            markalar = markalar_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.markalar_map = {"Seçim Yok": None}
            # marka_display_values artık kullanılmadığı için kaldırıldı.
            for m in markalar:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.markalar_map[m.get('ad')] = m.get('id') # <-- 'marka_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Markalar combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Markalar yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

    def _yukle_urun_grubu_birimi_ulke_comboboxlari(self):
        # Ürün Grupları
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            urun_gruplari_response = self.db.urun_grubu_listele() # <-- BURASI GÜNCELLENDİ
            urun_gruplari = urun_gruplari_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.urun_gruplari_map = {"Seçim Yok": None}
            # urun_grubu_display_values artık kullanılmadığı için kaldırıldı.
            for g in urun_gruplari:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.urun_gruplari_map[g.get('ad')] = g.get('id') # <-- 'grup_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Ürün grupları combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Ürün grupları yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

        # Ürün Birimleri
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            urun_birimleri_response = self.db.urun_birimi_listele() # <-- BURASI GÜNCELLENDİ
            urun_birimleri = urun_birimleri_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.urun_birimleri_map = {"Seçim Yok": None}
            # urun_birimi_display_values artık kullanılmadığı için kaldırıldı.
            for b in urun_birimleri:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.urun_birimleri_map[b.get('ad')] = b.get('id') # <-- 'birim_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Ürün birimleri combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Ürün birimleri yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

        # Ülkeler (Menşe)
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            ulkeler_response = self.db.ulke_listele() # <-- BURASI GÜNCELLENDİ
            ulkeler = ulkeler_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.ulkeler_map = {"Seçim Yok": None}
            # ulke_display_values artık kullanılmadığı için kaldırıldı.
            for u in ulkeler:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.ulkeler_map[u.get('ad')] = u.get('id') # <-- 'ulke_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Ülkeler combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Ülkeler yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

class StokKartiPenceresi(QDialog):
    data_updated = Signal()

    def __init__(self, parent, db_manager, yenile_callback=None, urun_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.parent_window = parent
        self.yenile_callback = yenile_callback
        self.app = app_ref
        self.urun_duzenle = urun_duzenle
        self.app_ref = app_ref
        self.duzenleme_modu = urun_duzenle is not None
        self.yeni_urun_resmi_yolu = None
        self.mevcut_urun_resmi_yolu = urun_duzenle.get('urun_resmi_yolu') if urun_duzenle and 'urun_resmi_yolu' in urun_duzenle else None
        
        self.stok_id = urun_duzenle.get('id') if urun_duzenle and 'id' in urun_duzenle else None

        self.original_pixmap = None 

        logger.info(f"StokKartiPenceresi başlatılıyor. Düzenleme modu: {self.duzenleme_modu}")

        self.setWindowTitle("Yeni Ürün Ekle" if not self.duzenleme_modu else "Ürün Kartı Düzenle")
        self.setModal(True)
        self.resize(800, 700)

        # UI elemanlarını burada tanımlıyoruz
        self.kod_e = QLineEdit()
        self.ad_e = QLineEdit()
        self.miktar_e = QLineEdit()
        self.alis_fiyat_e = QLineEdit()
        self.satis_fiyat_e = QLineEdit()
        self.kdv_e = QLineEdit()
        self.min_stok_e = QLineEdit()
        self.aktif_cb = QCheckBox()
        self.detay_e = QTextEdit()
        self.resim_label = QLabel("Resim Yok")
        self.kategori_combo = QComboBox()
        self.marka_combo = QComboBox()
        self.urun_grubu_combo = QComboBox()
        self.birim_combo = QComboBox()
        self.mensei_ulke_combo = QComboBox()
        
        # Bu nesneler, alt sekmeleri yönetecek
        self.stok_hareketleri_sekmesi = None
        self.ilgili_faturalar_sekmesi = None

        self._setup_ui() 
        
        if self.duzenleme_modu:
            self._mevcut_urunu_yukle()
        else:
            self._formu_sifirla()

        self._load_combobox_data()
        
    def _setup_ui(self):
        """Tüm UI elemanlarını tek bir metotta oluşturur."""
        main_container = QWidget(self)
        self.setLayout(QVBoxLayout(main_container))
        
        # Ana içerik ve resim için üst yatay çerçeve
        top_main_frame = QFrame(self)
        top_main_layout = QHBoxLayout(top_main_frame)
        self.layout().addWidget(top_main_frame)
        
        # Sol Panel (Giriş Alanları)
        input_panel_frame = QFrame(top_main_frame)
        input_panel_layout = QGridLayout(input_panel_frame)
        top_main_layout.addWidget(input_panel_frame, 2)
        
        input_panel_layout.addWidget(QLabel("Ürün Kodu (*):"), 0, 0)
        self.kod_e = QLineEdit()
        self.kod_e.setPlaceholderText("Benzersiz ürün kodu...")
        input_panel_layout.addWidget(self.kod_e, 0, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Ürün Adı (*):"), 1, 0)
        self.ad_e = QLineEdit()
        self.ad_e.setPlaceholderText("Ürün adı...")
        input_panel_layout.addWidget(self.ad_e, 1, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Miktar:"), 2, 0)
        self.miktar_e = QLineEdit()
        self.miktar_e.setReadOnly(True)
        input_panel_layout.addWidget(self.miktar_e, 2, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Alış Fiyatı (KDV Dahil):"), 3, 0)
        self.alis_fiyat_e = QLineEdit()
        setup_numeric_entry(self.app, self.alis_fiyat_e)
        input_panel_layout.addWidget(self.alis_fiyat_e, 3, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Satış Fiyatı (KDV Dahil):"), 4, 0)
        self.satis_fiyat_e = QLineEdit()
        setup_numeric_entry(self.app, self.satis_fiyat_e)
        input_panel_layout.addWidget(self.satis_fiyat_e, 4, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("KDV Oranı (%):"), 5, 0)
        self.kdv_e = QLineEdit()
        setup_numeric_entry(self.app, self.kdv_e, decimal_places=0)
        input_panel_layout.addWidget(self.kdv_e, 5, 1)
        
        input_panel_layout.addWidget(QLabel("Min. Stok Seviyesi:"), 6, 0)
        self.min_stok_e = QLineEdit()
        setup_numeric_entry(self.app, self.min_stok_e)
        input_panel_layout.addWidget(self.min_stok_e, 6, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Aktif:"), 7, 0)
        self.aktif_cb = QCheckBox()
        input_panel_layout.addWidget(self.aktif_cb, 7, 1, 1, 2)

        input_panel_layout.addWidget(QLabel("Kategori:"), 8, 0)
        self.kategori_combo = QComboBox()
        input_panel_layout.addWidget(self.kategori_combo, 8, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Marka:"), 9, 0)
        self.marka_combo = QComboBox()
        input_panel_layout.addWidget(self.marka_combo, 9, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Ürün Grubu:"), 10, 0)
        self.urun_grubu_combo = QComboBox()
        input_panel_layout.addWidget(self.urun_grubu_combo, 10, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Birim:"), 11, 0)
        self.birim_combo = QComboBox()
        input_panel_layout.addWidget(self.birim_combo, 11, 1, 1, 2)
        
        input_panel_layout.addWidget(QLabel("Menşei Ülke:"), 12, 0)
        self.mensei_ulke_combo = QComboBox()
        input_panel_layout.addWidget(self.mensei_ulke_combo, 12, 1, 1, 2)

        input_panel_layout.addWidget(QLabel("Detay:"), 13, 0, Qt.AlignTop)
        self.detay_e = QTextEdit()
        input_panel_layout.addWidget(self.detay_e, 13, 1, 1, 2)
        
        # Sağ Panel (Resim)
        image_panel_frame = QFrame(top_main_frame)
        image_panel_layout = QVBoxLayout(image_panel_frame)
        top_main_layout.addWidget(image_panel_frame, 1)
        
        self.resim_label = QLabel("Resim Yok")
        self.resim_label.setScaledContents(True)
        self.resim_label.setMinimumSize(250, 250)
        self.resim_label.setAlignment(Qt.AlignCenter)
        self.resim_label.setFrameShape(QFrame.StyledPanel)
        image_panel_layout.addWidget(self.resim_label)
        
        resim_button_frame = QFrame(image_panel_frame)
        resim_button_layout = QHBoxLayout(resim_button_frame)
        image_panel_layout.addWidget(resim_button_frame)
        
        btn_resim_sec = QPushButton("Resim Seç")
        btn_resim_sec.clicked.connect(self._resim_sec)
        resim_button_layout.addWidget(btn_resim_sec)
        
        btn_resim_sil = QPushButton("Resmi Kaldır")
        btn_resim_sil.clicked.connect(self._resim_sil)
        resim_button_layout.addWidget(btn_resim_sil)

        # Alt Sekmeler
        self.bottom_tab_widget = QTabWidget(self)
        self.layout().addWidget(self.bottom_tab_widget)
        
        # KRİTİK DÜZELTME: Hatalı 'StokHareketiSekmesi' adı, doğru olan 'StokHareketleriSekmesi' ile düzeltildi.
        from arayuz import StokHareketleriSekmesi, IlgiliFaturalarSekmesi 
        
        self.stok_hareketleri_sekmesi = StokHareketleriSekmesi(
            self.bottom_tab_widget, self.db, self.stok_id,
            self.urun_duzenle.get('ad', '-') if self.urun_duzenle else '',
            self.app)
        self.bottom_tab_widget.addTab(self.stok_hareketleri_sekmesi, "Stok Hareketleri")

        self.ilgili_faturalar_sekmesi = IlgiliFaturalarSekmesi(
            self.bottom_tab_widget, self.db, self.stok_id,
            self.urun_duzenle.get('ad', '-') if self.urun_duzenle else '',
            self.app)
        self.bottom_tab_widget.addTab(self.ilgili_faturalar_sekmesi, "İlgili Faturalar")

        # Butonlar
        self._add_bottom_buttons()

    def _add_bottom_buttons(self):
        """Pencerenin alt kısmındaki butonları oluşturur ve yerleştirir."""
        button_layout = QHBoxLayout()
        self.layout().addLayout(button_layout)

        self.btn_kaydet = QPushButton("Kaydet")
        self.btn_kaydet.clicked.connect(self.kaydet_urun)
        button_layout.addWidget(self.btn_kaydet)

        self.btn_iptal = QPushButton("İptal")
        self.btn_iptal.clicked.connect(self.reject)
        button_layout.addWidget(self.btn_iptal)

        self.btn_manuel_stok_giris = QPushButton("Manuel Stok Girişi")
        self.btn_manuel_stok_giris.clicked.connect(self._manuel_stok_giris_penceresi_ac)
        button_layout.addWidget(self.btn_manuel_stok_giris)

        self.btn_manuel_stok_cikis = QPushButton("Manuel Stok Çıkışı")
        self.btn_manuel_stok_cikis.clicked.connect(self._manuel_stok_cikis_penceresi_ac)
        button_layout.addWidget(self.btn_manuel_stok_cikis)

        self.btn_sil = QPushButton("Stoku Sil")
        self.btn_sil.clicked.connect(self._stok_sil)
        self.btn_sil.setVisible(bool(self.stok_id))
        button_layout.addWidget(self.btn_sil)

        if not self.duzenleme_modu:
            self.btn_manuel_stok_giris.setEnabled(False)
            self.btn_manuel_stok_cikis.setEnabled(False)
            self.bottom_tab_widget.setEnabled(False)

    def _manuel_stok_giris_penceresi_ac(self):
        """Stok ekleme penceresini açar."""
        if not self.stok_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce stoku kaydedin.")
            return
        
        # Güncel stok miktarını API'den çekerek al
        mevcut_stok = 0.0 # Varsayılan değer
        try:
            current_stok_data = self.db.stok_getir_by_id(self.stok_id)
            if current_stok_data:
                mevcut_stok = current_stok_data.get('miktar', 0.0)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Mevcut stok miktarı alınırken hata oluştu: {e}")
            logging.error(f"Stok miktarı alınırken hata: {e}", exc_info=True)
            return

        from pencereler import StokHareketiPenceresi
        dialog = StokHareketiPenceresi(
            self,
            self.db, # <-- db_manager parametresi eklendi
            self.stok_id,
            self.ad_e.text() if self.ad_e.text() else self.kod_e.text(),
            mevcut_stok,
            "GIRIŞ",
            self.refresh_data_and_ui
        )
        dialog.exec()

    def _manuel_stok_cikis_penceresi_ac(self):
        """Stok eksiltme penceresini açar."""
        if not self.stok_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce stoku kaydedin.")
            return

        # Güncel stok miktarını API'den çekerek al
        mevcut_stok = 0.0 # Varsayılan değer
        try:
            current_stok_data = self.db.stok_getir_by_id(self.stok_id)
            if current_stok_data:
                mevcut_stok = current_stok_data.get('miktar', 0.0)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Mevcut stok miktarı alınırken hata oluştu: {e}")
            logging.error(f"Stok miktarı alınırken hata: {e}", exc_info=True)
            return
        
        from pencereler import StokHareketiPenceresi
        dialog = StokHareketiPenceresi(
            self,
            self.db, # <-- db_manager parametresi eklendi
            self.stok_id,
            self.ad_e.text() if self.ad_e.text() else self.kod_e.text(),
            mevcut_stok,
            "CIKIS",
            self.refresh_data_and_ui
        )
        dialog.exec()
        
    def _stok_sil(self):
        reply = QMessageBox.question(self, "Ürün Silme Onayı", "Ürünü silmek istediğinizden emin misiniz? Bu işlem geri alınamaz.", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.stok_sil(self.stok_id)
                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.accept()
                    if self.refresh_callback:
                        self.refresh_callback()
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Ürün silinirken bir hata oluştu: {e}")

    def _resim_sec(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Stok Resmi Seç", "", "Resim Dosyaları (*.png *.jpg *.jpeg)")
        if file_path:
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                data_dir = os.path.join(base_dir, 'data')
                urun_resimleri_klasoru = os.path.join(data_dir, "urun_resimleri")
                
                os.makedirs(urun_resimleri_klasoru, exist_ok=True)
                yeni_path = os.path.join(urun_resimleri_klasoru, os.path.basename(file_path))
                shutil.copy2(file_path, yeni_path)
                self.yeni_urun_resmi_yolu = yeni_path
                self.mevcut_urun_resmi_yolu = None
                logger.info(f"Resim kopyalandı: {yeni_path}")
            except Exception as e:
                QMessageBox.warning(self, "Hata", f"Resim kopyalanamadı: {e}")
                logger.error(f"Resim kopyalama hatası: {e}", exc_info=True)
                self.yeni_urun_resmi_yolu = None
            self._load_urun_resmi()

    def _resim_sil(self):
        self.yeni_urun_resmi_yolu = ""
        self.mevcut_urun_resmi_yolu = None
        self._load_urun_resmi()
        logger.info("Ürün resmi silindi.")
    
    def _load_urun_resmi(self):
        resim_yolu = self.yeni_urun_resmi_yolu or self.mevcut_urun_resmi_yolu
        if resim_yolu and os.path.exists(resim_yolu):
            self.original_pixmap = QPixmap(resim_yolu)
            self._resize_image()
            self.resim_label.setText("")
        else:
            self.original_pixmap = None
            self.resim_label.setText("Resim Yok")
            self.resim_label.setPixmap(QPixmap())

    def _load_combobox_data(self):
        """Kategori, Marka vb. açılır listelerin verilerini yükler."""
        try:
            # Kullanıcı ID'si, yalnızca API metotlarında kullanılması gereken yerler için çekilir.
            kullanici_id = getattr(self.app_ref, 'current_user_id', None)
            
            if not kullanici_id:
                QMessageBox.warning(self, "Uyarı", "Nitelik verileri yüklenemedi: Kullanıcı ID'si mevcut değil.")
                return

            # --- Kategori Yükleme ---
            kategoriler = self.db.kategori_listele() 
            if isinstance(kategoriler, dict) and "items" in kategoriler:
                self.kategori_combo.addItem("Seçiniz...", -1)
                for item in kategoriler["items"]:
                    self.kategori_combo.addItem(item.get('ad'), item.get('id'))
            
            # --- Marka Yükleme ---
            markalar = self.db.marka_listele()
            if isinstance(markalar, dict) and "items" in markalar:
                self.marka_combo.addItem("Seçiniz...", -1)
                for item in markalar["items"]:
                    self.marka_combo.addItem(item.get('ad'), item.get('id'))
            
            # --- Ürün Grubu Yükleme ---
            urun_gruplari = self.db.urun_grubu_listele()
            if isinstance(urun_gruplari, dict) and "items" in urun_gruplari:
                self.urun_grubu_combo.addItem("Seçiniz...", -1)
                for item in urun_gruplari["items"]:
                    self.urun_grubu_combo.addItem(item.get('ad'), item.get('id'))
            
            # --- Birim Yükleme ---
            urun_birimleri = self.db.urun_birimi_listele(kullanici_id=kullanici_id)
            if isinstance(urun_birimleri, dict) and "items" in urun_birimleri:
                self.birim_combo.addItem("Seçiniz...", -1)
                for item in urun_birimleri["items"]:
                    self.birim_combo.addItem(item.get('ad'), item.get('id'))
            
            # --- Ülke Yükleme ---
            # KRİTİK DÜZELTME: kullanilci_id -> kullanici_id olarak düzeltildi
            ulkeler = self.db.ulke_listele(kullanici_id=kullanici_id) 
            if isinstance(ulkeler, dict) and "items" in ulkeler:
                self.mensei_ulke_combo.addItem("Seçiniz...", -1)
                for item in ulkeler["items"]:
                    self.mensei_ulke_combo.addItem(item.get('ad'), item.get('id'))
            
        except Exception as e:
            logger.error(f"StokKartiPenceresi: Nitelik verileri yüklenirken hata: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Nitelik verileri yüklenirken hata oluştu: {e}")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        QTimer.singleShot(50, self._resize_image)

    def _resize_image(self):
        if self.original_pixmap and not self.original_pixmap.isNull():
            scaled_pixmap = self.original_pixmap.scaled(self.resim_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.resim_label.setPixmap(scaled_pixmap)
            
    def refresh_data_and_ui(self):
        if not self.stok_id: return

        try:
            updated_stok_data = self.db.stok_getir_by_id(self.stok_id)

            self.miktar_e.setText(self.db._format_numeric(updated_stok_data.get('miktar', 0.0), 2))
            self.alis_fiyat_e.setText(self.db._format_numeric(updated_stok_data.get('alis_fiyati', 0.0), 2))
            self.satis_fiyat_e.setText(self.db._format_numeric(updated_stok_data.get('satis_fiyati', 0.0), 2))
            self.kdv_e.setText(self.db._format_numeric(updated_stok_data.get('kdv_orani', 0.0), 0))
            self.min_stok_e.setText(self.db._format_numeric(updated_stok_data.get('min_stok_seviyesi', 0.0), 2))
            self.aktif_cb.setChecked(updated_stok_data.get('aktif', True))

            self.data_updated.emit()
            self.stok_hareketleri_sekmesi._load_stok_hareketleri()
            self.ilgili_faturalar_sekmesi._load_ilgili_faturalar()
            
            logger.info(f"Stok kartı verileri yenilendi: ID {self.stok_id}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Stok verileri yenilenirken hata oluştu:\n{e}")
            logger.error(f"StokKartiPenceresi refresh_data_and_ui hatası: {e}", exc_info=True)

    def _mevcut_urunu_yukle(self):
        self.kod_e.setText(self.urun_duzenle.get('kod', ''))
        self.ad_e.setText(self.urun_duzenle.get('ad', ''))
        self.miktar_e.setText(self.db._format_numeric(self.urun_duzenle.get('miktar', 0.0), 2))
        self.alis_fiyat_e.setText(self.db._format_numeric(self.urun_duzenle.get('alis_fiyati', 0.0), 2))
        self.satis_fiyat_e.setText(self.db._format_numeric(self.urun_duzenle.get('satis_fiyati', 0.0), 2))
        self.kdv_e.setText(self.db._format_numeric(self.urun_duzenle.get('kdv_orani', 0.0), 0))
        self.min_stok_e.setText(self.db._format_numeric(self.urun_duzenle.get('min_stok_seviyesi', 0.0), 2))
        self.aktif_cb.setChecked(self.urun_duzenle.get('aktif', True))
        self.detay_e.setPlainText(self.urun_duzenle.get('detay', ''))

        self.kategori_combo.setCurrentIndex(self.kategori_combo.findData(self.urun_duzenle.get('kategori_id')))
        self.marka_combo.setCurrentIndex(self.marka_combo.findData(self.urun_duzenle.get('marka_id')))
        self.urun_grubu_combo.setCurrentIndex(self.urun_grubu_combo.findData(self.urun_duzenle.get('urun_grubu_id')))
        self.birim_combo.setCurrentIndex(self.birim_combo.findData(self.urun_duzenle.get('birim_id')))
        self.mensei_ulke_combo.setCurrentIndex(self.mensei_ulke_combo.findData(self.urun_duzenle.get('mense_id')))

        self.mevcut_urun_resmi_yolu = self.urun_duzenle.get('urun_resmi_yolu')
        self._load_urun_resmi()

        urun_adi_for_tabs = self.ad_e.text() if self.ad_e.text() else self.kod_e.text()

        if self.stok_id:
            self.stok_hareketleri_sekmesi.urun_id = self.stok_id
            self.stok_hareketleri_sekmesi.urun_adi = urun_adi_for_tabs
            self.ilgili_faturalar_sekmesi.urun_id = self.stok_id
            self.ilgili_faturalar_sekmesi.urun_adi = urun_adi_for_tabs
            self.stok_hareketleri_sekmesi._load_stok_hareketleri()
            self.ilgili_faturalar_sekmesi._load_ilgili_faturalar()
            self.btn_manuel_stok_giris.setEnabled(True)
            self.btn_manuel_stok_cikis.setEnabled(True)
            self.bottom_tab_widget.setEnabled(True)
        else:
            logger.warning("Mevcut ürünü yüklerken ID bulunamadı, sekmeler devre dışı bırakıldı.")
            self.btn_manuel_stok_giris.setEnabled(False)
            self.btn_manuel_stok_cikis.setEnabled(False)
            self.bottom_tab_widget.setEnabled(False)


        logger.info(f"Ürün ID {self.stok_id} için mevcut ürün verileri yüklendi.")

    def _formu_sifirla(self):
        try:
            generated_stok_kodu = self.db.get_next_stok_kodu()
            self.kod_e.setText(generated_stok_kodu)
            self.kod_e.setReadOnly(True)
            if generated_stok_kodu.startswith("MANUEL"):
                self.kod_e.setReadOnly(False)
        except Exception as e:
            QMessageBox.warning(self, "Kod Üretme Hatası", f"Otomatik stok kodu alınırken bir hata oluştu: {e}. Lütfen manuel olarak giriniz.")
            self.kod_e.setReadOnly(False)
            
        self.ad_e.clear()
        self.miktar_e.setText("0,00")
        self.alis_fiyat_e.setText("0,00")
        self.satis_fiyat_e.setText("0,00")
        self.kdv_e.setText("20")
        self.min_stok_e.setText("0,00")
        self.aktif_cb.setChecked(True)
        self.detay_e.clear()
        self.resim_label.setText("Resim Yok")
        self.resim_label.setPixmap(QPixmap())
        self.yeni_urun_resmi_yolu = None
        self.mevcut_urun_resmi_yolu = None

        self.kategori_combo.setCurrentIndex(0)
        self.marka_combo.setCurrentIndex(0)
        self.urun_grubu_combo.setCurrentIndex(0)
        self.birim_combo.setCurrentIndex(0)
        self.mensei_ulke_combo.setCurrentIndex(0)

        self.stok_id = None
        self.duzenleme_modu = False
        self.setWindowTitle("Yeni Ürün Ekle")
        self.btn_sil.setVisible(False)
        self.btn_manuel_stok_giris.setEnabled(False)
        self.btn_manuel_stok_cikis.setEnabled(False)
        
        # Sekmeler sadece düzenleme modunda etkin olmalı
        self.stok_hareketleri_sekmesi.urun_id = None
        self.stok_hareketleri_sekmesi.urun_adi = ""
        self.stok_hareketleri_sekmesi._load_stok_hareketleri()

        self.ilgili_faturalar_sekmesi.urun_id = None
        self.ilgili_faturalar_sekmesi.urun_adi = ""
        self.ilgili_faturalar_sekmesi._load_ilgili_faturalar()
        
        self.bottom_tab_widget.setEnabled(False)

        logger.info("Stok Kartı formu sıfırlandı.")

    def kaydet_urun(self):
        kod = self.kod_e.text().strip()
        ad = self.ad_e.text().strip()
        
        # Miktar readonly olduğu için direkt API'den gelen değeri kullanıyoruz.
        # Bu metot sadece kart bilgilerini günceller, miktar hareketlerle değişir.
        miktar = self.urun_duzenle.get('miktar', 0.0) if self.duzenleme_modu else 0.0
        
        # self.db.safe_float() metodu ile değerleri alıyoruz
        alis_fiyati = self.db.safe_float(self.alis_fiyat_e.text())
        satis_fiyati = self.db.safe_float(self.satis_fiyat_e.text())
        kdv_orani = self.db.safe_float(self.kdv_e.text())
        min_stok = self.db.safe_float(self.min_stok_e.text())
        
        aktif = self.aktif_cb.isChecked()
        detay = self.detay_e.toPlainText().strip()
        
        kategori_id = self.kategori_combo.currentData()
        marka_id = self.marka_combo.currentData()
        urun_grubu_id = self.urun_grubu_combo.currentData()
        birim_id = self.birim_combo.currentData()
        mense_id = self.mensei_ulke_combo.currentData()

        if not ad:
            QMessageBox.critical(self, "Eksik Bilgi", "Ürün Adı boş olamaz.")
            return

        if not self.duzenleme_modu: # Yeni ürün eklenirken kod da zorunlu
            if not kod:
                QMessageBox.critical(self, "Eksik Bilgi", "Yeni ürün için Ürün Kodu boş olamaz.")
                return

        if miktar < 0 or alis_fiyati < 0 or satis_fiyati < 0 or kdv_orani < 0 or min_stok < 0:
            QMessageBox.critical(self, "Geçersiz Değer", "Miktar, fiyatlar, KDV oranı ve minimum stok negatif olamaz.")
            return
        
        urun_data = {
            "kod": kod,
            "ad": ad,
            "miktar": miktar,
            "alis_fiyati": alis_fiyati,
            "satis_fiyati": satis_fiyati,
            "kdv_orani": kdv_orani,
            "min_stok_seviyesi": min_stok,
            "aktif": aktif,
            "detay": detay if detay else None,
            "kategori_id": kategori_id,
            "marka_id": marka_id,
            "urun_grubu_id": urun_grubu_id,
            "birim_id": birim_id,
            "mense_id": mense_id,
            "urun_resmi_yolu": self.yeni_urun_resmi_yolu if self.yeni_urun_resmi_yolu else self.mevcut_urun_resmi_yolu
        }
        
        # KRİTİK DÜZELTME: API'nin zorunlu tuttuğu kullanici_id'yi request body'ye (urun_data) ekle
        urun_data["kullanici_id"] = self.app.current_user.get("id")

        # KRİTİK EKSİK ARGÜMAN DÜZELTMESİ: kullanici_id'yi argüman olarak çekiyoruz
        kullanici_id_arg = self.app.current_user.get("id")

        try:
            if self.duzenleme_modu and self.stok_id:
                # Stok güncelleme işlemi
                # KRİTİK DÜZELTME: Eksik olan 'kullanici_id' argümanı eklendi.
                success, message = self.db.stok_guncelle(self.stok_id, urun_data, kullanici_id_arg)
            else:
                # Yeni stok ekleme işlemi
                success, message = self.db.stok_ekle(urun_data)

            if success:
                QMessageBox.information(self, "Başarılı", message)
                self.data_updated.emit()
                self.accept()
                if self.yenile_callback:
                    self.yenile_callback()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            logger.error(f"Ürün kaydedilirken hata oluştu: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Ürün kaydedilirken bir hata oluştu:\n{e}")

class YeniKasaBankaEklePenceresi(QDialog):
    def __init__(self, parent, db_manager, yenile_callback, hesap_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.app = app_ref
        self.yenile_callback = yenile_callback
        self.hesap_duzenle_data = hesap_duzenle

        self.hesap_duzenle_id = self.hesap_duzenle_data.get('id') if self.hesap_duzenle_data else None

        title = "Yeni Kasa/Banka Hesabı" if not self.hesap_duzenle_id else "Hesap Düzenle"
        self.setWindowTitle(title)
        self.setFixedSize(430, 250) 
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        
        # Değişiklik: Başlık etrafındaki gereksiz boşlukları kaldırmak için
        # title_label'ı direk main_layout'a ekliyoruz ve stretch faktörü kullanmıyoruz
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        # Form için QGridLayout kullanılıyor
        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)
        
        self.entries = {}
        self.odeme_turleri = ["YOK", "NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET", "AÇIK HESAP"]
        
        # Form elemanları
        form_layout.addWidget(QLabel("Hesap Adı (*):"), 0, 0)
        self.entries['hesap_adi'] = QLineEdit()
        form_layout.addWidget(self.entries['hesap_adi'], 0, 1)

        form_layout.addWidget(QLabel("Hesap Tipi (*):"), 1, 0)
        self.entries['tip'] = QComboBox()
        self.entries['tip'].addItems(["KASA", "BANKA"])
        self.entries['tip'].currentTextChanged.connect(self._tip_degisince_banka_alanlarini_ayarla)
        form_layout.addWidget(self.entries['tip'], 1, 1)

        self.banka_labels = {
            'banka_adi': QLabel("Banka Adı:"),
            'sube_adi': QLabel("Şube Adı:"),
            'hesap_no': QLabel("Hesap No/IBAN:")
        }
        form_layout.addWidget(self.banka_labels['banka_adi'], 2, 0)
        self.entries['banka_adi'] = QLineEdit()
        form_layout.addWidget(self.entries['banka_adi'], 2, 1)
        
        form_layout.addWidget(self.banka_labels['sube_adi'], 3, 0)
        self.entries['sube_adi'] = QLineEdit()
        form_layout.addWidget(self.entries['sube_adi'], 3, 1)

        form_layout.addWidget(self.banka_labels['hesap_no'], 4, 0)
        self.entries['hesap_no'] = QLineEdit()
        form_layout.addWidget(self.entries['hesap_no'], 4, 1)
        
        form_layout.addWidget(QLabel("Açılış Bakiyesi:"), 5, 0)
        self.entries['bakiye'] = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.entries['bakiye'])
        form_layout.addWidget(self.entries['bakiye'], 5, 1)

        form_layout.addWidget(QLabel("Para Birimi:"), 6, 0)
        self.entries['para_birimi'] = QLineEdit("TL")
        form_layout.addWidget(self.entries['para_birimi'], 6, 1)

        form_layout.addWidget(QLabel("Varsayılan Ödeme Türü:"), 7, 0)
        self.entries['varsayilan_odeme_turu'] = QComboBox()
        self.entries['varsayilan_odeme_turu'].addItems(self.odeme_turleri)
        form_layout.addWidget(self.entries['varsayilan_odeme_turu'], 7, 1)
        
        # Butonlar için bir layout
        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()
        kaydet_button = QPushButton("Kaydet")
        kaydet_button.clicked.connect(self.kaydet)
        button_layout.addWidget(kaydet_button)
        iptal_button = QPushButton("İptal")
        iptal_button.clicked.connect(self.reject)
        button_layout.addWidget(iptal_button)
        
        self._verileri_yukle()
        self._tip_degisince_banka_alanlarini_ayarla()

    def _tip_degisince_banka_alanlarini_ayarla(self):
        is_banka = self.entries['tip'].currentText() == "BANKA"
        for key, widget in self.banka_labels.items():
            widget.setVisible(is_banka)
        for key in ['banka_adi', 'sube_adi', 'hesap_no']:
            self.entries[key].setVisible(is_banka)
        if not is_banka:
            for key in ['banka_adi', 'sube_adi', 'hesap_no']:
                self.entries[key].clear()

    def _verileri_yukle(self):
        if self.hesap_duzenle_data:
            self.entries['hesap_adi'].setText(self.hesap_duzenle_data.get('hesap_adi', ''))
            self.entries['tip'].setCurrentText(self.hesap_duzenle_data.get('tip', 'KASA'))
            self.entries['banka_adi'].setText(self.hesap_duzenle_data.get('banka_adi', ''))
            self.entries['sube_adi'].setText(self.hesap_duzenle_data.get('sube_adi', ''))
            self.entries['hesap_no'].setText(self.hesap_duzenle_data.get('hesap_no', ''))
            bakiye = self.hesap_duzenle_data.get('bakiye', 0.0)
            self.entries['bakiye'].setText(f"{bakiye:.2f}".replace('.', ','))
            self.entries['para_birimi'].setText(self.hesap_duzenle_data.get('para_birimi', 'TL'))
            varsayilan_odeme_turu = self.hesap_duzenle_data.get('varsayilan_odeme_turu')
            self.entries['varsayilan_odeme_turu'].setCurrentText(varsayilan_odeme_turu if varsayilan_odeme_turu else "YOK")
            self.entries['bakiye'].setReadOnly(True) # Açılış bakiyesi düzenlemede değiştirilemez

    def kaydet(self):
        hesap_adi = self.entries['hesap_adi'].text().strip()
        if not hesap_adi:
            QMessageBox.warning(self, "Eksik Bilgi", "Hesap Adı alanı boş bırakılamaz.")
            return

        bakiye_str = self.entries['bakiye'].text().replace(',', '.')
        
        data = {
            "hesap_adi": hesap_adi,
            "tip": self.entries['tip'].currentText(),
            "bakiye": float(bakiye_str) if bakiye_str else 0.0,
            "banka_adi": self.entries['banka_adi'].text().strip(),
            "sube_adi": self.entries['sube_adi'].text().strip(),
            "hesap_no": self.entries['hesap_no'].text().strip(),
            "para_birimi": self.entries['para_birimi'].text().strip(),
            "varsayilan_odeme_turu": self.entries['varsayilan_odeme_turu'].currentText()
        }
        if data["varsayilan_odeme_turu"] == "YOK":
            data["varsayilan_odeme_turu"] = None

        try:
            if self.hesap_duzenle_id:
                # GÜNCELLEME (PUT isteği)
                success = self.db.kasa_banka_guncelle(self.hesap_duzenle_id, data)
            else:
                # YENİ KAYIT (POST isteği)
                success = self.db.kasa_banka_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Kasa/Banka hesabı başarıyla kaydedildi.")
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", "Kasa/Banka hesabı kaydedilirken bir hata oluştu.")

        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "Hata", f"Hesap kaydedilirken bir hata oluştu:\n{error_detail}")
            logging.error(f"Kasa/Banka kaydetme hatası: {error_detail}", exc_info=True)

class YeniTedarikciEklePenceresi(QDialog):
    def __init__(self, parent, db_manager, yenile_callback, tedarikci_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.app = app_ref
        self.yenile_callback = yenile_callback
        self.tedarikci_duzenle_data = tedarikci_duzenle

        self.tedarikci_duzenle_id = self.tedarikci_duzenle_data.get('id') if self.tedarikci_duzenle_data else None

        # CariService örneğini burada oluştur
        from hizmetler import CariService # CariService'i burada import ediyoruz
        self.cari_service = CariService(self.db) # <-- CariService BAŞLATILDI

        title = "Yeni Tedarikçi Ekle" if not self.tedarikci_duzenle_id else "Tedarikçi Düzenle"
        self.setWindowTitle(title)
        self.setFixedSize(500, 330)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)
        
        self.entries = {}
        labels_entries = {
            "Tedarikçi Kodu:": "entry_kod",
            "Ad Soyad (*):": "entry_ad",
            "Telefon:": "entry_tel",
            "Adres:": "entry_adres",
            "Vergi Dairesi:": "entry_vd",
            "Vergi No:": "entry_vn"
        }

        for i, (label_text, entry_name) in enumerate(labels_entries.items()):
            form_layout.addWidget(QLabel(label_text), i, 0, alignment=Qt.AlignCenter)
            if entry_name == "entry_adres":
                widget = QTextEdit()
                widget.setFixedHeight(80)
                # QTextEdit için karakter sınırı (255)
                widget.textChanged.connect(lambda w=widget: self._limit_text_length(w, 255))
            else:
                widget = QLineEdit()
                
            self.entries[entry_name] = widget
            form_layout.addWidget(widget, i, 1)

        # KRİTİK GÜNCELLEME: QLineEdit'lar için karakter kısıtlamaları
        self.entries["entry_kod"].setMaxLength(50)  # Kod
        self.entries["entry_ad"].setMaxLength(100) # Ad Soyad
        self.entries["entry_tel"].setMaxLength(11)  # Telefon (İstenen kısıtlama)
        self.entries["entry_vd"].setMaxLength(100) # Vergi Dairesi
        self.entries["entry_vn"].setMaxLength(20)  # Vergi No/TCKN

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()
        self.kaydet_button = QPushButton("Kaydet")
        self.kaydet_button.clicked.connect(self.kaydet)
        button_layout.addWidget(self.kaydet_button)
        self.iptal_button = QPushButton("İptal")
        self.iptal_button.clicked.connect(self.reject)
        button_layout.addWidget(self.iptal_button)
        
        self._verileri_yukle()

    def _limit_text_length(self, text_edit_widget, max_length):
        """QTextEdit için metin uzunluğunu sınırlayan yardımcı metot."""
        if len(text_edit_widget.toPlainText()) > max_length:
            cursor = text_edit_widget.textCursor()
            cursor.movePosition(cursor.End)
            text_edit_widget.setText(text_edit_widget.toPlainText()[:max_length])
            text_edit_widget.setTextCursor(cursor)

    def _verileri_yukle(self):
        """Mevcut tedarikçi verilerini düzenleme modunda forma yükler ve yeni kod üretir."""
        kullanici_id = self.app.current_user.get("id")
        
        if self.tedarikci_duzenle_data:
            self.entries["entry_kod"].setText(self.tedarikci_duzenle_data.get('kod', ''))
            self.entries["entry_ad"].setText(self.tedarikci_duzenle_data.get('ad', ''))
            self.entries["entry_tel"].setText(self.tedarikci_duzenle_data.get('telefon', ''))
            self.entries["entry_adres"].setPlainText(self.tedarikci_duzenle_data.get('adres', ''))
            self.entries["entry_vd"].setText(self.tedarikci_duzenle_data.get('vergi_dairesi', ''))
            self.entries["entry_vn"].setText(self.tedarikci_duzenle_data.get('vergi_no', ''))
            self.entries["entry_kod"].setReadOnly(True)
        else:
            generated_code = self.db.get_next_tedarikci_kodu(kullanici_id=kullanici_id)
            self.entries["entry_kod"].setText(generated_code)
            self.entries["entry_kod"].setReadOnly(True)

    def kaydet(self):
        ad = self.entries["entry_ad"].text().strip()
        if not ad:
            QMessageBox.warning(self, "Eksik Bilgi", "Tedarikçi Adı alanı boş bırakılamaz.")
            return

        data = {
            "ad": ad,
            "kod": self.entries["entry_kod"].text().strip(),
            "telefon": self.entries["entry_tel"].text().strip(),
            "adres": self.entries["entry_adres"].toPlainText().strip(),
            "vergi_dairesi": self.entries["entry_vd"].text().strip(),
            "vergi_no": self.entries["entry_vn"].text().strip()
        }

        try:
            if self.tedarikci_duzenle_id:
                success, message = self.db.tedarikci_guncelle(self.tedarikci_duzenle_id, data)
            else:
                success, message = self.db.tedarikci_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Tedarikçi bilgileri başarıyla kaydedildi.")
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", "Tedarikçi kaydedilirken bir hata oluştu.")

        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "Hata", f"Tedarikçi kaydedilirken bir hata oluştu:\n{error_detail}")
            logging.error(f"Tedarikçi kaydetme hatası: {error_detail}", exc_info=True)

class KalemDuzenlePenceresi(QDialog):
    def __init__(self, parent_page, db_manager, kalem_index, kalem_verisi, islem_tipi, fatura_id_duzenle=None):
        super().__init__(parent_page)
        self.parent_page = parent_page # FaturaPenceresi objesi
        self.db = db_manager # db_manager artık direkt parametre olarak alınıyor
        self.kalem_index = kalem_index
        self.islem_tipi = islem_tipi
        self.fatura_id_duzenle = fatura_id_duzenle
        
        # DEĞİŞİKLİK: app referansını parent_page üzerinden alıyoruz.
        self.app = parent_page.app

        self.urun_id = kalem_verisi[0]
        self.urun_adi = kalem_verisi[1]
        self.mevcut_miktar = self.db.safe_float(kalem_verisi[2])
        self.orijinal_birim_fiyat_kdv_haric = self.db.safe_float(kalem_verisi[3])
        self.kdv_orani = self.db.safe_float(kalem_verisi[4])
        self.mevcut_alis_fiyati_fatura_aninda = self.db.safe_float(kalem_verisi[8])
        
        self_initial_iskonto_yuzde_1 = self.db.safe_float(kalem_verisi[10])
        self_initial_iskonto_yuzde_2 = self.db.safe_float(kalem_verisi[11])

        self.orijinal_birim_fiyat_kdv_dahil = self.orijinal_birim_fiyat_kdv_haric * (1 + self.kdv_orani / 100)

        self.setWindowTitle(f"Kalem Düzenle: {self.urun_adi}")
        self.setFixedSize(450, 550) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame) # Izgara düzenleyici
        
        main_frame_layout.addWidget(QLabel(f"Ürün: <b>{self.urun_adi}</b>", font=QFont("Segoe UI", 12, QFont.Bold)), 0, 0, 1, 3, Qt.AlignCenter)
        main_frame_layout.setColumnStretch(1, 1) # İkinci sütun genişlesin

        current_row = 1
        main_frame_layout.addWidget(QLabel("Miktar:"), current_row, 0)
        self.miktar_e = QLineEdit()
        setup_numeric_entry(self.app, self.miktar_e) # <-- decimal_places=2 kaldırıldı
        self.miktar_e.setText(f"{self.mevcut_miktar:.2f}".replace('.',','))
        self.miktar_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.miktar_e, current_row, 1)

        current_row += 1
        main_frame_layout.addWidget(QLabel("Birim Fiyat (KDV Dahil):"), current_row, 0)
        self.fiyat_e = QLineEdit()
        setup_numeric_entry(self.app, self.fiyat_e) # <-- decimal_places=2 kaldırıldı
        self.fiyat_e.setText(f"{self.orijinal_birim_fiyat_kdv_dahil:.2f}".replace('.',','))
        self.fiyat_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.fiyat_e, current_row, 1)

        current_row += 1
        self.alis_fiyati_aninda_e = None
        if self.islem_tipi in [self.db.FATURA_TIP_SATIS, self.db.SIPARIS_TIP_SATIS]:
            main_frame_layout.addWidget(QLabel("Fatura Anı Alış Fiyatı (KDV Dahil):"), current_row, 0)
            self.alis_fiyati_aninda_e = QLineEdit()
            setup_numeric_entry(self.app, self.alis_fiyati_aninda_e) # <-- decimal_places=2 kaldırıldı
            self.alis_fiyati_aninda_e.setText(f"{self.mevcut_alis_fiyati_fatura_aninda:.2f}".replace('.',','))
            self.alis_fiyati_aninda_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
            main_frame_layout.addWidget(self.alis_fiyati_aninda_e, current_row, 1)
            current_row += 1
        
        main_frame_layout.addWidget(QFrame(), current_row, 0, 1, 3) # Separator yerine boş QFrame
        current_row += 1

        main_frame_layout.addWidget(QLabel("İsk 1 (%):"), current_row, 0)
        self.iskonto_yuzde_1_e = QLineEdit()
        setup_numeric_entry(self.app, self.iskonto_yuzde_1_e) # <-- decimal_places=2 kaldırıldı
        self.iskonto_yuzde_1_e.setText(f"{self_initial_iskonto_yuzde_1:.2f}".replace('.',','))
        self.iskonto_yuzde_1_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.iskonto_yuzde_1_e, current_row, 1)
        main_frame_layout.addWidget(QLabel("%"), current_row, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("İsk 2 (%):"), current_row, 0)
        self.iskonto_yuzde_2_e = QLineEdit()
        setup_numeric_entry(self.app, self.iskonto_yuzde_2_e) # <-- decimal_places=2, max_value=100 kaldırıldı
        self.iskonto_yuzde_2_e.setText(f"{self_initial_iskonto_yuzde_2:.2f}".replace('.',','))
        self.iskonto_yuzde_2_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.iskonto_yuzde_2_e, current_row, 1)
        main_frame_layout.addWidget(QLabel("%"), current_row, 2)
        current_row += 1

        main_frame_layout.addWidget(QFrame(), current_row, 0, 1, 3) # Separator yerine boş QFrame
        current_row += 1

        main_frame_layout.addWidget(QLabel("Toplam İsk Yüzdesi:", font=QFont("Segoe UI", 9, QFont.Bold)), current_row, 0)
        self.lbl_toplam_iskonto_yuzdesi = QLabel("0,00 %", font=QFont("Segoe UI", 9))
        main_frame_layout.addWidget(self.lbl_toplam_iskonto_yuzdesi, current_row, 1, 1, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("Uygulanan İsk Tutarı (KDV Dahil):", font=QFont("Segoe UI", 9, QFont.Bold)), current_row, 0)
        self.lbl_uygulanan_iskonto_dahil = QLabel("0,00 TL", font=QFont("Segoe UI", 9))
        main_frame_layout.addWidget(self.lbl_uygulanan_iskonto_dahil, current_row, 1, 1, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("İskontolu Birim Fiyat (KDV Dahil):", font=QFont("Segoe UI", 9, QFont.Bold)), current_row, 0)
        self.lbl_iskontolu_bf_dahil = QLabel("0,00 TL", font=QFont("Segoe UI", 9))
        main_frame_layout.addWidget(self.lbl_iskontolu_bf_dahil, current_row, 1, 1, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("Kalem Toplam (KDV Dahil):", font=QFont("Segoe UI", 10, QFont.Bold)), current_row, 0)
        self.lbl_kalem_toplam_dahil = QLabel("0,00 TL", font=QFont("Segoe UI", 10, QFont.Bold))
        main_frame_layout.addWidget(self.lbl_kalem_toplam_dahil, current_row, 1, 1, 2)
        current_row += 1

        btn_f = QFrame(self)
        btn_layout = QHBoxLayout(btn_f)
        main_layout.addWidget(btn_f, alignment=Qt.AlignCenter)
        
        btn_guncelle = QPushButton("Güncelle")
        btn_guncelle.clicked.connect(self._kalemi_kaydet)
        btn_layout.addWidget(btn_guncelle)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        btn_layout.addWidget(btn_iptal)

        self._anlik_hesaplama_ve_guncelleme()
        self.miktar_e.setFocus()
        self.miktar_e.selectAll()

    def _anlik_hesaplama_ve_guncelleme(self):
        try:
            miktar = self.db.safe_float(self.miktar_e.text())
            birim_fiyat_kdv_dahil_orijinal = self.db.safe_float(self.fiyat_e.text())

            yuzde_iskonto_1 = self.db.safe_float(self.iskonto_yuzde_1_e.text())
            yuzde_iskonto_2 = self.db.safe_float(self.iskonto_yuzde_2_e.text())

            if not (0 <= yuzde_iskonto_1 <= 100):
                self.iskonto_yuzde_1_e.setText("0,00")
                yuzde_iskonto_1 = 0.0

            if not (0 <= yuzde_iskonto_2 <= 100):
                self.iskonto_yuzde_2_e.setText("0,00")
                yuzde_iskonto_2 = 0.0

            fiyat_iskonto_1_sonrasi_dahil = birim_fiyat_kdv_dahil_orijinal * (1 - yuzde_iskonto_1 / 100)
            iskontolu_birim_fiyat_dahil = fiyat_iskonto_1_sonrasi_dahil * (1 - yuzde_iskonto_2 / 100)
            
            if iskontolu_birim_fiyat_dahil < 0:
                iskontolu_birim_fiyat_dahil = 0.0

            toplam_uygulanan_iskonto_dahil = birim_fiyat_kdv_dahil_orijinal - iskontolu_birim_fiyat_dahil
            
            kalem_toplam_dahil = miktar * iskontolu_birim_fiyat_dahil

            if birim_fiyat_kdv_dahil_orijinal > 0:
                toplam_iskonto_yuzdesi = (toplam_uygulanan_iskonto_dahil / birim_fiyat_kdv_dahil_orijinal) * 100
            else:
                toplam_iskonto_yuzdesi = 0.0 

            self.lbl_toplam_iskonto_yuzdesi.setText(f"{toplam_iskonto_yuzdesi:,.2f} %")
            self.lbl_uygulanan_iskonto_dahil.setText(self.db._format_currency(toplam_uygulanan_iskonto_dahil))
            self.lbl_iskontolu_bf_dahil.setText(self.db._format_currency(iskontolu_birim_fiyat_dahil))
            self.lbl_kalem_toplam_dahil.setText(self.db._format_currency(kalem_toplam_dahil))

        except ValueError:
            self.lbl_toplam_iskonto_yuzdesi.setText("0,00 %")
            self.lbl_uygulanan_iskonto_dahil.setText("0,00 TL")
            self.lbl_iskontolu_bf_dahil.setText("0,00 TL")
            self.lbl_kalem_toplam_dahil.setText("0,00 TL")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Hesaplama sırasında beklenmeyen bir hata oluştu: {e}")
            logging.error(f"Anlık hesaplama hatası: {e}", exc_info=True)


    def _kalemi_kaydet(self):
        yeni_miktar = 0.0
        yeni_fiyat_kdv_dahil_orijinal = 0.0
        yuzde_iskonto_1 = 0.0
        yuzde_iskonto_2 = 0.0
        yeni_alis_fiyati_aninda = self.mevcut_alis_fiyati_fatura_aninda

        try:
            yeni_miktar = self.db.safe_float(self.miktar_e.text())
            yeni_fiyat_kdv_dahil_orijinal = self.db.safe_float(self.fiyat_e.text())
            
            yuzde_iskonto_1 = self.db.safe_float(self.iskonto_yuzde_1_e.text())
            yuzde_iskonto_2 = self.db.safe_float(self.iskonto_yuzde_2_e.text())
            
            if (self.islem_tipi == self.db.FATURA_TIP_SATIS or self.islem_tipi == self.db.SIPARIS_TIP_SATIS) and self.alis_fiyati_aninda_e:
                yeni_alis_fiyati_aninda = self.db.safe_float(self.alis_fiyati_aninda_e.text())

            if yeni_miktar <= 0:
                QMessageBox.critical(self, "Geçersiz Miktar", "Miktar pozitif bir sayı olmalıdır.")
                return
            if yeni_fiyat_kdv_dahil_orijinal < 0:
                QMessageBox.critical(self, "Geçersiz Fiyat", "Birim fiyat negatif olamaz.")
                return
            if not (0 <= yuzde_iskonto_1 <= 100):
                QMessageBox.critical(self, "Geçersiz İsk 1 Yüzdesi", "İsk 1 yüzdesi 0 ile 100 arasında olmalıdır.")
                return
            if not (0 <= yuzde_iskonto_2 <= 100):
                QMessageBox.critical(self, "Geçersiz İsk 2 Yüzdesi", "İsk 2 yüzdesi 0 ile 100 arasında olmalıdır.")
                return
            if (self.islem_tipi == self.db.FATURA_TIP_SATIS or self.islem_tipi == self.db.SIPARIS_TIP_SATIS) and self.alis_fiyati_aninda_e and yeni_alis_fiyati_aninda < 0:
                QMessageBox.critical(self, "Geçersiz Fiyat", "Fatura anı alış fiyatı negatif olamaz.")
                return
            
            self.parent_page.kalem_guncelle( # _kalem_guncelle yerine kalem_guncelle oldu
                self.kalem_index, 
                yeni_miktar, 
                yeni_fiyat_kdv_dahil_orijinal, 
                yuzde_iskonto_1,       
                yuzde_iskonto_2,       
                yeni_alis_fiyati_aninda 
            )
            self.accept() # QDialog'u kapat.

        except ValueError as ve:
            QMessageBox.critical(self, "Giriş Hatası", f"Sayısal alanlarda geçersiz değerler var: {ve}")
            logging.error(f"Kalem Guncelle ValueError: {ve}", exc_info=True)
        except IndexError as ie:
            QMessageBox.critical(self, "Hata", f"Güncellenecek kalem bulunamadı (indeks hatası): {ie}")
            logging.error(f"Kalem Guncelle IndexError: {ie}", exc_info=True)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kalem güncellenirken beklenmeyen bir hata oluştu: {e}")
            logging.error(f"Kalem Guncelle Genel Hata: {e}", exc_info=True)
            
class FiyatGecmisiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, cari_id, urun_id, fatura_tipi, update_callback, current_kalem_index):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.cari_id = cari_id
        self.urun_id = urun_id
        self.fatura_tipi = fatura_tipi
        self.update_callback = update_callback
        self.current_kalem_index = current_kalem_index

        self.setWindowTitle("Fiyat Geçmişi Seç")
        self.setFixedSize(600, 400)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Geçmiş Fiyat Listesi")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        tree_frame = QFrame(self)
        tree_layout = QVBoxLayout(tree_frame)
        main_layout.addWidget(tree_frame)

        cols = ("Fatura No", "Tarih", "Fiyat (KDV Dahil)", "İsk 1 (%)", "İsk 2 (%)")
        self.price_history_tree = QTreeWidget()
        self.price_history_tree.setHeaderLabels(cols)
        self.price_history_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.price_history_tree.setSortingEnabled(True)

        from PySide6.QtWidgets import QHeaderView
        col_defs = [
            ("Fatura No", 120, Qt.AlignCenter),
            ("Tarih", 90, Qt.AlignCenter),
            ("Fiyat (KDV Dahil)", 120, Qt.AlignCenter),
            ("İsk 1 (%)", 90, Qt.AlignCenter),
            ("İsk 2 (%)", 90, Qt.AlignCenter)
        ]

        for i, (col_name, width, alignment) in enumerate(col_defs):
            self.price_history_tree.setColumnWidth(i, width)
            self.price_history_tree.headerItem().setTextAlignment(i, alignment)
            self.price_history_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))
        
        self.price_history_tree.header().setStretchLastSection(True)

        tree_layout.addWidget(self.price_history_tree)
        self.price_history_tree.itemDoubleClicked.connect(self._on_price_selected_double_click)
        self._load_price_history()

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_onayla = QPushButton("Seç ve Uygula")
        btn_onayla.clicked.connect(self._on_price_selected_button)
        button_layout.addWidget(btn_onayla)
        
        button_layout.addStretch()

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        button_layout.addWidget(btn_kapat)

    def _load_price_history(self):
        """Veritabanından geçmiş fiyat bilgilerini çeker ve Treeview'e doldurur."""
        self.price_history_tree.clear()
        
        # db.get_gecmis_fatura_kalemi_bilgileri metodunu çağırıyoruz
        try:
            history_data = self.db.get_gecmis_fatura_kalemi_bilgileri(self.cari_id, self.urun_id, self.fatura_tipi)

            if not history_data:
                item_qt = QTreeWidgetItem(self.price_history_tree)
                item_qt.setText(2, "Geçmiş Fiyat Yok")
                return

            for item in history_data:
                # API'den dönen veri Pydantic modeline uygun bir sözlük olmalı
                fatura_no = item.get('fatura_no', '-')
                tarih = item.get('tarih', '-')
                fiyat_kdv_dahil = item.get('nihai_iskontolu_kdv_dahil_bf', 0.0)
                iskonto_1 = item.get('iskonto_yuzde_1', 0.0)
                iskonto_2 = item.get('iskonto_yuzde_2', 0.0)

                # Tarih objesi ise formatla
                if isinstance(tarih, (datetime, date)):
                    tarih = tarih.strftime('%d.%m.%Y')

                item_qt = QTreeWidgetItem(self.price_history_tree)
                item_qt.setText(0, fatura_no)
                item_qt.setText(1, tarih)
                item_qt.setText(2, self.db._format_currency(fiyat_kdv_dahil))
                item_qt.setText(3, f"%{iskonto_1:.2f}".replace('.', ',').rstrip('0').rstrip(','))
                item_qt.setText(4, f"%{iskonto_2:.2f}".replace('.', ',').rstrip('0').rstrip(','))
                
                # Sütunları ortala
                for i in range(self.price_history_tree.columnCount()):
                    item_qt.setTextAlignment(i, Qt.AlignCenter)

        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Geçmiş fiyat listesi çekilirken hata: {e}")
            logging.error(f"Geçmiş fiyat listesi yükleme hatası: {e}", exc_info=True)

    def _on_price_selected_double_click(self, item, column):
        self._on_price_selected_button()

    def _on_price_selected_button(self):
        selected_items = self.price_history_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen uygulamak için bir geçmiş fiyat seçin.")
            return

        item_values = [selected_items[0].text(i) for i in range(self.price_history_tree.columnCount())]
        
        selected_price_str = item_values[2] 
        selected_iskonto1_str = item_values[3] 
        selected_iskonto2_str = item_values[4] 

        try:
            cleaned_price_str = selected_price_str.replace(' TL', '').replace('₺', '').strip()
            cleaned_iskonto1_str = selected_iskonto1_str.replace('%', '').strip()
            cleaned_iskonto2_str = selected_iskonto2_str.replace('%', '').strip()

            selected_price = self.db.safe_float(cleaned_price_str)
            selected_iskonto1 = self.db.safe_float(cleaned_iskonto1_str)
            selected_iskonto2 = self.db.safe_float(cleaned_iskonto2_str)

        except ValueError:
            QMessageBox.critical(self, "Hata", "Seçilen fiyat verisi geçersiz. (Dönüştürme hatası)")
            return
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Fiyat geçmişi verisi işlenirken beklenmeyen bir hata oluştu: {e}")
            logging.error(f"Fiyat geçmişi verisi işleme hatası: {e}", exc_info=True)
            return

        self.update_callback(self.current_kalem_index, selected_price, selected_iskonto1, selected_iskonto2)
        self.close()
        
class KullaniciYonetimiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app # Ana App referansı
        self.setWindowTitle("Kullanıcı Yönetimi")
        self.setMinimumSize(600, 650)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Kullanıcı Listesi ve Yönetimi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # Kullanıcı Listesi
        list_frame = QFrame(self)
        list_layout = QHBoxLayout(list_frame)
        main_layout.addWidget(list_frame)
        
        cols_kul = ("ID", "Kullanıcı Adı", "Yetki")
        self.tree_kul = QTreeWidget()
        self.tree_kul.setHeaderLabels(cols_kul)
        self.tree_kul.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tree_kul.setSortingEnabled(True) # Sıralama özelliği

        self.tree_kul.setColumnWidth(0, 50)
        self.tree_kul.headerItem().setTextAlignment(0, Qt.AlignCenter)
        self.tree_kul.headerItem().setTextAlignment(2, Qt.AlignCenter)
        self.tree_kul.header().setSectionResizeMode(1, QHeaderView.Stretch) # Kullanıcı Adı genişlesin

        list_layout.addWidget(self.tree_kul)
        
        self.kullanici_listesini_yenile() # İlk yüklemede listeyi doldur

        # Yeni Kullanıcı Ekleme Formu
        form_frame = QGroupBox("Yeni Kullanıcı Ekle / Güncelle", self)
        form_layout = QGridLayout(form_frame)
        main_layout.addWidget(form_frame)

        form_layout.addWidget(QLabel("Kullanıcı Adı:"), 0, 0, Qt.AlignCenter)
        self.k_adi_yeni_e = QLineEdit()
        form_layout.addWidget(self.k_adi_yeni_e, 0, 1)
        form_layout.setColumnStretch(1, 1) # Genişlesin

        form_layout.addWidget(QLabel("Yeni Şifre:"), 1, 0, Qt.AlignCenter)
        self.sifre_yeni_e = QLineEdit()
        self.sifre_yeni_e.setEchoMode(QLineEdit.Password) # Şifre gizleme
        form_layout.addWidget(self.sifre_yeni_e, 1, 1)

        form_layout.addWidget(QLabel("Yetki:"), 0, 2, Qt.AlignCenter)
        self.yetki_yeni_cb = QComboBox()
        self.yetki_yeni_cb.addItems(["kullanici", "admin"])
        self.yetki_yeni_cb.setCurrentText("kullanici") # Varsayılan
        form_layout.addWidget(self.yetki_yeni_cb, 0, 3)

        # Butonlar
        button_frame_kul = QFrame(self)
        button_layout_kul = QHBoxLayout(button_frame_kul)
        main_layout.addWidget(button_frame_kul)
        
        self.ekle_guncelle_btn = QPushButton("Ekle / Güncelle")
        self.ekle_guncelle_btn.clicked.connect(self.yeni_kullanici_ekle)
        button_layout_kul.addWidget(self.ekle_guncelle_btn)
        
        btn_sil_kul = QPushButton("Seçili Kullanıcıyı Sil")
        btn_sil_kul.clicked.connect(self.secili_kullanici_sil)
        button_layout_kul.addWidget(btn_sil_kul)
        
        button_layout_kul.addStretch() # Sağa yaslama
        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        button_layout_kul.addWidget(btn_kapat)

        self.tree_kul.itemSelectionChanged.connect(self.secili_kullaniciyi_forma_yukle) # Seçim değiştiğinde formu doldur

    def kullanici_listesini_yenile(self):
        self.tree_kul.clear()
        try:
            # API'den kullanıcı listesini çekmek için uygun bir endpoint varsayımı
            # Eğer API'de böyle bir endpoint yoksa, doğrudan db_manager kullanılmalıdır.
            # Şimdilik db_manager'dan çekiliyor.
            kullanicilar = self.db.kullanici_listele()
            
            for kul in kullanicilar:
                item_qt = QTreeWidgetItem(self.tree_kul)
                item_qt.setText(0, str(kul.get('id'))) # 'id' alanı
                item_qt.setText(1, kul.get('kullanici_adi')) # 'kullanici_adi' alanı
                item_qt.setText(2, kul.get('yetki')) # 'yetki' alanı
                item_qt.setData(0, Qt.UserRole, kul.get('id')) # ID'yi UserRole olarak sakla
                
            self.app.set_status_message(f"{len(kullanicilar)} kullanıcı listelendi.")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kullanıcı listesi çekilirken hata: {e}")
            logging.error(f"Kullanıcı listesi yükleme hatası: {e}", exc_info=True)
    
    def secili_kullaniciyi_forma_yukle(self):
        selected_items = self.tree_kul.selectedItems()
        if selected_items:
            item = selected_items[0]
            kullanici_adi = item.text(1)
            yetki = item.text(2)
            self.k_adi_yeni_e.setText(kullanici_adi)
            self.yetki_yeni_cb.setCurrentText(yetki)
            self.sifre_yeni_e.clear() # Şifre alanı temizlensin
            self.ekle_guncelle_btn.setText("Güncelle")
        else: # Seçim yoksa formu temizle
            self.k_adi_yeni_e.clear()
            self.sifre_yeni_e.clear()
            self.yetki_yeni_cb.setCurrentText("kullanici")
            self.ekle_guncelle_btn.setText("Ekle / Güncelle")

    def yeni_kullanici_ekle(self):
        k_adi = self.k_adi_yeni_e.text().strip()
        sifre = self.sifre_yeni_e.text().strip()
        yetki = self.yetki_yeni_cb.currentText()

        if not (k_adi and yetki):
            QMessageBox.critical(self, "Eksik Bilgi", "Kullanıcı adı ve yetki boş bırakılamaz.")
            return

        selected_items = self.tree_kul.selectedItems()
        
        if selected_items: # Güncelleme
            user_id = selected_items[0].data(0, Qt.UserRole)
            mevcut_k_adi = selected_items[0].text(1)

            success_name_update = True
            message_name_update = ""

            if k_adi != mevcut_k_adi:
                try:
                    # API endpoint'i üzerinden kullanıcı adını güncelleme (varsayalım mevcut)
                    # response = requests.put(f"{API_BASE_URL}/kullanicilar/{user_id}/kullanici_adi", json={"kullanici_adi": k_adi})
                    # response.raise_for_status()
                    # success_name_update, message_name_update = True, "Kullanıcı adı güncellendi."
                    success_name_update, message_name_update = self.db.kullanici_adi_guncelle(user_id, k_adi)

                except Exception as e:
                    success_name_update = False
                    message_name_update = f"Kullanıcı adı güncellenirken hata: {e}"
                    logging.error(f"Kullanıcı adı güncelleme hatası: {e}", exc_info=True)
                
                if not success_name_update:
                    QMessageBox.critical(self, "Hata", message_name_update)
                    return

            sifre_to_hash = None
            if sifre:
                sifre_to_hash = self.db._hash_sifre(sifre)
            else: # Şifre boş bırakılırsa mevcut şifreyi koru
                try:
                    # API'den şifre çekme veya doğrudan db_manager'dan çekme
                    # response = requests.get(f"{API_BASE_URL}/kullanicilar/{user_id}/sifre_hash")
                    # response.raise_for_status()
                    # sifre_to_hash = response.json().get('sifre_hash')
                    self.db.c.execute("SELECT sifre FROM kullanicilar WHERE id=?", (user_id,))
                    sifre_to_hash = self.db.c.fetchone()[0]
                except Exception as e:
                    QMessageBox.critical(self, "Hata", f"Mevcut şifre alınırken bir hata oluştu: {e}")
                    logging.error(f"Mevcut şifre alma hatası: {e}", exc_info=True)
                    return

            try:
                # API endpoint'i üzerinden kullanıcıyı güncelleme
                # response = requests.put(f"{API_BASE_URL}/kullanicilar/{user_id}", json={"sifre": sifre_to_hash, "yetki": yetki})
                # response.raise_for_status()
                # success, message = True, "Kullanıcı başarıyla güncellendi."
                success, message = self.db.kullanici_guncelle_sifre_yetki(user_id, sifre_to_hash, yetki)

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kullanıcı güncellenirken hata: {e}")
                logging.error(f"Kullanıcı güncelleme hatası: {e}", exc_info=True)

            self.kullanici_listesini_yenile()
            self.k_adi_yeni_e.clear()
            self.sifre_yeni_e.clear()
            self.tree_kul.clearSelection()
            self.secili_kullaniciyi_forma_yukle() # Formu temizle (butonu da "Ekle / Güncelle" yapar)

        else: # Yeni kullanıcı ekleme
            if not sifre:
                QMessageBox.critical(self, "Eksik Bilgi", "Yeni kullanıcı eklerken şifre boş bırakılamaz.")
                return

            try:
                # API endpoint'i üzerinden yeni kullanıcı ekleme
                # response = requests.post(f"{API_BASE_URL}/kullanicilar/", json={"kullanici_adi": k_adi, "sifre": sifre, "yetki": yetki})
                # response.raise_for_status()
                # success, message = True, "Yeni kullanıcı başarıyla eklendi."
                success, message = self.db.kullanici_ekle(k_adi, sifre, yetki)

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Yeni kullanıcı eklenirken hata: {e}")
                logging.error(f"Yeni kullanıcı ekleme hatası: {e}", exc_info=True)

            self.kullanici_listesini_yenile()
            self.k_adi_yeni_e.clear()
            self.sifre_yeni_e.clear()
            self.tree_kul.clearSelection()
            self.secili_kullaniciyi_forma_yukle()

    def secili_kullanici_sil(self):
        selected_items = self.tree_kul.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Seçim Yok", "Lütfen silmek istediğiniz kullanıcıyı seçin.")
            return
        
        k_adi_secili = selected_items[0].text(1)
        user_id_to_delete = selected_items[0].data(0, Qt.UserRole)

        if k_adi_secili == self.app.current_user[1]: 
             QMessageBox.warning(self, "Engellendi", "Aktif olarak giriş yapmış olduğunuz kendi kullanıcı hesabınızı silemezsiniz.")
             return

        reply = QMessageBox.question(self, "Onay", f"'{k_adi_secili}' kullanıcısını silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                # API endpoint'i üzerinden kullanıcı silme
                # response = requests.delete(f"{API_BASE_URL}/kullanicilar/{user_id_to_delete}")
                # response.raise_for_status()
                # success, message = True, "Kullanıcı başarıyla silindi."
                success, message = self.db.kullanici_sil(user_id_to_delete)

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.kullanici_listesini_yenile()
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kullanıcı silinirken hata: {e}")
                logging.error(f"Kullanıcı silme hatası: {e}", exc_info=True)

class YeniGelirGiderEklePenceresi(QDialog):
    def __init__(self, parent, db_manager, yenile_callback, initial_tip=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.yenile_callback = yenile_callback
        # DEĞİŞİKLİK: app_ref yerine parent'tan app referansını al
        self.app = app_ref if app_ref else parent
        
        self.kasa_banka_map = {}
        self.gelir_siniflandirma_map = {}
        self.gider_siniflandirma_map = {}

        self.setWindowTitle("Yeni Manuel Gelir/Gider Kaydı")
        self.setFixedSize(420, 265)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        entry_frame = QFrame(self)
        main_layout.addWidget(entry_frame)
        entry_frame_layout = QGridLayout(entry_frame)
        
        current_row = 0

        entry_frame_layout.addWidget(QLabel("Tarih (YYYY-AA-GG):"), current_row, 0, Qt.AlignCenter)
        self.tarih_entry = QLineEdit()
        self.tarih_entry.setText(datetime.now().strftime('%Y-%m-%d'))
        self.tarih_entry.setPlaceholderText("YYYY-AA-GG")
        entry_frame_layout.addWidget(self.tarih_entry, current_row, 1)
        btn_date = QPushButton("🗓️")
        btn_date.setFixedWidth(30)
        btn_date.clicked.connect(lambda: DatePickerDialog(self.app, self.tarih_entry))
        entry_frame_layout.addWidget(btn_date, current_row, 2)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("İşlem Tipi:"), current_row, 0, Qt.AlignCenter)
        self.tip_combo = QComboBox()
        self.tip_combo.addItems(["GELİR", "GİDER"])
        if initial_tip and initial_tip in ["GELİR", "GİDER"]:
            self.tip_combo.setCurrentText(initial_tip)
        else:
            self.tip_combo.setCurrentIndex(0)
        self.tip_combo.currentIndexChanged.connect(self._on_tip_changed)
        entry_frame_layout.addWidget(self.tip_combo, current_row, 1)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("Sınıflandırma:"), current_row, 0, Qt.AlignCenter)
        self.siniflandirma_combo = QComboBox()
        entry_frame_layout.addWidget(self.siniflandirma_combo, current_row, 1)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("Tutar (TL):"), current_row, 0, Qt.AlignCenter)
        self.tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.tutar_entry)
        entry_frame_layout.addWidget(self.tutar_entry, current_row, 1)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("İşlem Kasa/Banka (*):"), current_row, 0, Qt.AlignCenter)
        self.kasa_banka_combobox = QComboBox()
        entry_frame_layout.addWidget(self.kasa_banka_combobox, current_row, 1)
        current_row += 1
        
        entry_frame_layout.addWidget(QLabel("Açıklama:"), current_row, 0, Qt.AlignCenter)
        self.aciklama_entry = QLineEdit()
        entry_frame_layout.addWidget(self.aciklama_entry, current_row, 1)
        current_row += 1
        
        entry_frame_layout.setColumnStretch(1, 1)

        main_layout.addStretch()

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignCenter)

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close)
        button_layout.addWidget(btn_iptal)
        
        self._yukle_siniflandirmalar_comboboxlari_ve_ayarla()
        self.tarih_entry.setFocus()
        self.adjustSize()

    def _yukle_siniflandirmalar_comboboxlari_ve_ayarla(self):
        self._yukle_kasa_banka_hesaplarini() 

        # API'den sınıflandırmaları çek
        try:
            gelir_siniflandirmalar_response = self.db.gelir_siniflandirma_listele()
            
            # DÜZELTME: API yanıtının doğru formatta olup olmadığını kontrol et
            if isinstance(gelir_siniflandirmalar_response, dict) and "items" in gelir_siniflandirmalar_response:
                gelir_siniflandirmalar = gelir_siniflandirmalar_response["items"]
            else:
                gelir_siniflandirmalar = []
                # Hata loglanıyor ancak uygulamanın çökmesi engelleniyor
                logging.warning("API'den geçersiz gelir sınıflandırma yanıtı alındı. Liste boş olarak işleniyor.")
                
            self.gelir_siniflandirma_map = {item.get('ad'): item.get('id') for item in gelir_siniflandirmalar}

            gider_siniflandirmalar_response = self.db.gider_siniflandirma_listele()
            
            # DÜZELTME: API yanıtının doğru formatta olup olmadığını kontrol et
            if isinstance(gider_siniflandirmalar_response, dict) and "items" in gider_siniflandirmalar_response:
                gider_siniflandirmalar = gider_siniflandirmalar_response["items"]
            else:
                gider_siniflandirmalar = []
                logging.warning("API'den geçersiz gider sınıflandırma yanıtı alındı. Liste boş olarak işleniyor.")
                
            self.gider_siniflandirma_map = {item.get('ad'): item.get('id') for item in gider_siniflandirmalar}

        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırmalar yüklenirken hata: {e}")
            logging.error(f"Sınıflandırma yükleme hatası: {e}", exc_info=True)
            self.siniflandirma_combo.addItem("Hata", None)
            return

        self._on_tip_changed()

    def _on_tip_changed(self):
        selected_tip = self.tip_combo.currentText()
        display_values = ["Seçim Yok"]
        selected_map = {}

        if selected_tip == "GELİR":
            selected_map = self.gelir_siniflandirma_map
        elif selected_tip == "GİDER":
            selected_map = self.gider_siniflandirma_map

        display_values.extend(sorted(selected_map.keys()))
        self.siniflandirma_combo.clear()
        self.siniflandirma_combo.addItems(display_values)
        self.siniflandirma_combo.setCurrentText("Seçim Yok")
        # combobox'ın state'i QComboBox'ta otomatik olarak readonly'dir.

    def _yukle_kasa_banka_hesaplarini(self):
        self.kasa_banka_combobox.clear()
        self.kasa_banka_map.clear()

        try:
            hesaplar_response = self.db.kasa_banka_listesi_al(limit=10000)

            hesaplar = []
            if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                hesaplar = hesaplar_response["items"]
            elif isinstance(hesaplar_response, list):
                hesaplar = hesaplar_response
                self.app.set_status_message("Uyarı: Kasa/Banka listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                hesaplar = []
                self.app.set_status_message("Hata: Kasa/Banka listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Kasa/Banka listesi API'den beklenen formatta gelmedi: {type(hesaplar_response)} - {hesaplar_response}")
                self.kasa_banka_combobox.addItem("Hesap Yok", None)
                self.kasa_banka_combobox.setEnabled(False)
                return

            if hesaplar:
                for h in hesaplar:
                    display_text = f"{h.get('hesap_adi')} ({h.get('tip')})"
                    if h.get('tip') == "BANKA" and h.get('banka_adi'):
                        display_text += f" - {h.get('banka_adi')}"
                    if h.get('bakiye') is not None:
                        display_text += f" (Bakiye: {self.db._format_currency(h.get('bakiye'))})"

                    self.kasa_banka_map[display_text] = h.get('id')
                    self.kasa_banka_combobox.addItem(display_text, h.get('id'))

                default_hesap_text = None
                for text in self.kasa_banka_map.keys():
                    if text.strip().startswith("NAKİT KASA"):
                        default_hesap_text = text
                        break

                if default_hesap_text:
                    self.kasa_banka_combobox.setCurrentText(default_hesap_text)
                elif self.kasa_banka_combobox.count() > 0:
                    self.kasa_banka_combobox.setCurrentIndex(0)
                else:
                    self.kasa_banka_combobox.clear()
                    self.kasa_banka_combobox.addItem("Hesap Yok", None)
                    self.kasa_banka_combobox.setEnabled(False)

            self.app.set_status_message(f"{len(hesaplar)} kasa/banka hesabı API'den yüklendi.")

        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Kasa/Banka hesapları yüklenirken hata: {e}")
            logging.error(f"Kasa/Banka yükleme hatası: {e}", exc_info=True)
            self.kasa_banka_combobox.addItem("Hesap Yok", None)
            self.kasa_banka_combobox.setEnabled(False)

    def _kaydet(self):
        tarih_str = self.tarih_entry.text().strip()
        tip_str = self.tip_combo.currentText()
        tutar_str = self.tutar_entry.text().strip()
        aciklama_str = self.aciklama_entry.text().strip()

        secili_hesap_id = self.kasa_banka_combobox.currentData()

        secili_siniflandirma_adi = self.siniflandirma_combo.currentText()
        gelir_siniflandirma_id_val = None
        gider_siniflandirma_id_val = None

        if secili_siniflandirma_adi != "Seçim Yok" and secili_siniflandirma_adi:
            if tip_str == "GELİR":
                gelir_siniflandirma_id_val = self.gelir_siniflandirma_map.get(secili_siniflandirma_adi)
            elif tip_str == "GİDER":
                gider_siniflandirma_id_val = self.gider_siniflandirma_map.get(secili_siniflandirma_adi)
        
        if secili_hesap_id is None:
            QMessageBox.critical(self, "Eksik Bilgi", "Lütfen bir İşlem Kasa/Banka hesabı seçin.")
            return

        # Tarih formatı kontrolü
        try:
            datetime.strptime(tarih_str, '%Y-%m-%d')
        except ValueError:
            QMessageBox.critical(self, "Hata", "Tarih formatı 'YYYY-AA-GG' şeklinde olmalıdır.")
            return
        
        # YENİ KOD: Açıklama veya Sınıflandırma alanlarından birinin dolu olmasını zorunlu kıl
        if not (aciklama_str or (secili_siniflandirma_adi and secili_siniflandirma_adi != "Seçim Yok")):
             QMessageBox.critical(self, "Eksik Bilgi", "Lütfen bir açıklama veya sınıflandırma girin.")
             return

        try:
            tutar_f = float(tutar_str.replace(',', '.'))
            if tutar_f <= 0:
                QMessageBox.critical(self, "Geçersiz Tutar", "Tutar pozitif bir sayı olmalıdır.")
                return
        except ValueError:
            QMessageBox.critical(self, "Giriş Hatası", "Tutar sayısal bir değer olmalıdır.")
            return

        try:
            data = {
                "tarih": tarih_str,
                "tip": tip_str,
                "tutar": tutar_f,
                "aciklama": aciklama_str,
                "kaynak": "MANUEL",
                "kasa_banka_id": secili_hesap_id,
                "gelir_siniflandirma_id": gelir_siniflandirma_id_val,
                "gider_siniflandirma_id": gider_siniflandirma_id_val
            }
            success = self.db.gelir_gider_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Gelir/Gider kaydı başarıyla eklendi.")
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", "Gelir/Gider kaydı eklenirken bir hata oluştu.")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kaydedilirken bir hata oluştu:\n{e}")
            logging.error(f"Gelir/Gider kaydetme hatası: {e}", exc_info=True)

class TarihAraligiDialog(QDialog): # simpledialog.Dialog yerine QDialog kullanıldı
    def __init__(self, parent_app, title=None, baslangic_gun_sayisi=30):
        super().__init__(parent_app)
        self.app = parent_app # Ana uygulama referansını tut
        self.bas_tarih_str = (datetime.now() - timedelta(days=baslangic_gun_sayisi)).strftime('%Y-%m-%d')
        self.bitis_tarih_str = datetime.now().strftime('%Y-%m-%d')
        self.sonuc = None # Kullanıcının seçtiği tarih aralığını tutacak

        self.setWindowTitle(title if title else "Tarih Aralığı Seçin")
        self.setFixedSize(350, 180) # Sabit boyut
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)

        form_layout.addWidget(QLabel("Başlangıç Tarihi (YYYY-AA-GG):"), 0, 0, Qt.AlignCenter)
        self.bas_tarih_entry_dialog = QLineEdit()
        self.bas_tarih_entry_dialog.setText(self.bas_tarih_str)
        form_layout.addWidget(self.bas_tarih_entry_dialog, 0, 1)
        btn_bas_tarih = QPushButton("🗓️")
        btn_bas_tarih.setFixedWidth(30)
        btn_bas_tarih.clicked.connect(lambda: DatePickerDialog(self.app, self.bas_tarih_entry_dialog)) # app referansı kullanıldı
        form_layout.addWidget(btn_bas_tarih, 0, 2)

        form_layout.addWidget(QLabel("Bitiş Tarihi (YYYY-AA-GG):"), 1, 0, Qt.AlignCenter)
        self.bit_tarih_entry_dialog = QLineEdit()
        self.bit_tarih_entry_dialog.setText(self.bitis_tarih_str)
        form_layout.addWidget(self.bit_tarih_entry_dialog, 1, 1)
        btn_bit_tarih = QPushButton("🗓️")
        btn_bit_tarih.setFixedWidth(30)
        btn_bit_tarih.clicked.connect(lambda: DatePickerDialog(self.app, self.bit_tarih_entry_dialog)) # app referansı kullanıldı
        form_layout.addWidget(btn_bit_tarih, 1, 2)

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()

        btn_ok = QPushButton("Onayla")
        btn_ok.clicked.connect(self._apply)
        button_layout.addWidget(btn_ok)

        btn_cancel = QPushButton("İptal")
        btn_cancel.clicked.connect(self.reject) # QDialog'u reject ile kapat
        button_layout.addWidget(btn_cancel)

        self.bas_tarih_entry_dialog.setFocus() # İlk odaklanılacak widget

    def _apply(self): 
        bas_t_str_dialog = self.bas_tarih_entry_dialog.text()
        bit_t_str_dialog = self.bit_tarih_entry_dialog.text()
        try:
            bas_dt_dialog = datetime.strptime(bas_t_str_dialog, '%Y-%m-%d')
            bit_dt_dialog = datetime.strptime(bit_t_str_dialog, '%Y-%m-%d')
            if bas_dt_dialog > bit_dt_dialog:
                QMessageBox.critical(self, "Tarih Hatası", "Başlangıç tarihi, bitiş tarihinden sonra olamaz.")
                self.sonuc = None 
                return
            self.sonuc = (bas_t_str_dialog, bit_t_str_dialog) 
            self.accept() # QDialog'u accept ile kapat
        except ValueError:
            QMessageBox.critical(self, "Format Hatası", "Tarih formatı YYYY-AA-GG olmalıdır (örn: 2023-12-31).")
            self.sonuc = None
            return

class OdemeTuruSecimDialog(QDialog):
    def __init__(self, parent_app, db_manager, fatura_tipi, initial_cari_id, callback_func):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.fatura_tipi = fatura_tipi
        self.initial_cari_id = initial_cari_id
        self.callback_func = callback_func
        self.setWindowTitle("Ödeme Türü Seçimi")
        self.setFixedSize(400, 300)
        self.setModal(True)
        self.kasa_banka_map = {}
        main_layout = QVBoxLayout(self)
        title_label = QLabel("Fatura Ödeme Türünü Seçin")
        title_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        form_frame = QFrame(self)
        form_layout = QGridLayout(form_frame)
        main_layout.addWidget(form_frame)
        form_layout.addWidget(QLabel("Ödeme Türü (*):"), 0, 0, Qt.AlignCenter)
        self.odeme_turu_cb = QComboBox()
        self._set_odeme_turu_values()
        form_layout.addWidget(self.odeme_turu_cb, 0, 1)
        self.odeme_turu_cb.currentIndexChanged.connect(self._odeme_turu_degisince_hesap_combobox_ayarla)
        self.odeme_turu_cb.setCurrentIndex(0)
        form_layout.addWidget(QLabel("İşlem Kasa/Banka (*):"), 1, 0, Qt.AlignCenter)
        self.islem_hesap_cb = QComboBox()
        self.islem_hesap_cb.setEnabled(False)
        form_layout.addWidget(self.islem_hesap_cb, 1, 1)
        self.lbl_vade_tarihi = QLabel("Vade Tarihi:")
        self.entry_vade_tarihi = QLineEdit()
        self.entry_vade_tarihi.setEnabled(False) 
        self.btn_vade_tarihi = QPushButton("🗓️")
        self.btn_vade_tarihi.setFixedWidth(30)
        self.btn_vade_tarihi.clicked.connect(lambda: DatePickerDialog(self.app, self.entry_vade_tarihi))
        self.btn_vade_tarihi.setEnabled(False)
        form_layout.addWidget(self.lbl_vade_tarihi, 2, 0, Qt.AlignCenter)
        form_layout.addWidget(self.entry_vade_tarihi, 2, 1)
        form_layout.addWidget(self.btn_vade_tarihi, 2, 2)
        self.lbl_vade_tarihi.hide()
        self.entry_vade_tarihi.hide()
        self.btn_vade_tarihi.hide()
        form_layout.setColumnStretch(1, 1)
        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)
        btn_onayla = QPushButton("Onayla")
        btn_onayla.clicked.connect(self._onayla)
        button_layout.addWidget(btn_onayla)
        button_layout.addStretch()
        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close)
        button_layout.addWidget(btn_iptal)
        self._yukle_kasa_banka_hesaplarini()
        self._odeme_turu_degisince_hesap_combobox_ayarla()

    def _set_odeme_turu_values(self):
        all_payment_values = [self.db.ODEME_TURU_NAKIT, self.db.ODEME_TURU_KART, 
                              self.db.ODEME_TURU_EFT_HAVALE, self.db.ODEME_TURU_CEK, 
                              self.db.ODEME_TURU_SENET, self.db.ODEME_TURU_ACIK_HESAP, 
                              self.db.ODEME_TURU_ETKISIZ_FATURA]
        
        # Seçili cari varsayılan bir cari mi diye kontrol et
        is_default_cari = (
            (self.fatura_tipi == self.db.FATURA_TIP_SATIS and self.initial_cari_id is not None and str(self.initial_cari_id) == str(self.db.get_perakende_musteri_id())) or
            (self.fatura_tipi == self.db.FATURA_TIP_ALIS and self.initial_cari_id is not None and str(self.initial_cari_id) == str(self.db.get_genel_tedarikci_id()))
        )

        if is_default_cari:
            # Varsayılan cari ise, açık hesap ve etkisiz faturayı listeden çıkar.
            self.odeme_turu_cb.addItems([p for p in all_payment_values if p != self.db.ODEME_TURU_ACIK_HESAP and p != self.db.ODEME_TURU_ETKISIZ_FATURA])
        else:
            # Diğer cariler için sadece etkisiz faturayı çıkar.
            self.odeme_turu_cb.addItems([p for p in all_payment_values if p != self.db.ODEME_TURU_ETKISIZ_FATURA])
            
    def _yukle_kasa_banka_hesaplarini(self):
        self.islem_hesap_cb.clear()
        self.kasa_banka_map.clear()
        try:
            hesaplar_response = self.db.kasa_banka_listesi_al(limit=10000)
            hesaplar = []
            if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                hesaplar = hesaplar_response["items"]
            elif isinstance(hesaplar_response, list):
                hesaplar = hesaplar_response
                self.app.set_status_message("Uyarı: Kasa/Banka listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                hesaplar = []
                self.app.set_status_message("Hata: Kasa/Banka listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Kasa/Banka listesi API'den beklenen formatta gelmedi: {type(hesaplar_response)} - {hesaplar_response}")
                self.islem_hesap_cb.addItem("Hesap Yok", None)
                self.islem_hesap_cb.setEnabled(False)
                return
            if hesaplar:
                for h in hesaplar:
                    display_text = f"{h.get('hesap_adi')} ({h.get('tip')})"
                    if h.get('tip') == "BANKA" and h.get('banka_adi'):
                        display_text += f" - {h.get('banka_adi')}"
                    if h.get('bakiye') is not None:
                        display_text += f" (Bakiye: {self.db._format_currency(h.get('bakiye'))})"
                    self.kasa_banka_map[display_text] = h.get('id')
                    self.islem_hesap_cb.addItem(display_text, h.get('id'))
                self.islem_hesap_cb.setEnabled(True)
                self.islem_hesap_cb.setCurrentIndex(0)
            else:
                self.islem_hesap_cb.addItem("Hesap Yok", None)
                self.islem_hesap_cb.setEnabled(False)
            self.app.set_status_message(f"{len(hesaplar)} kasa/banka hesabı API'den yüklendi.")
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Kasa/Banka hesapları yüklenirken hata: {e}")
            logging.error(f"Kasa/Banka yükleme hatası: {e}", exc_info=True)
            self.islem_hesap_cb.addItem("Hesap Yok", None)
            self.islem_hesap_cb.setEnabled(False)

    def _odeme_turu_degisince_hesap_combobox_ayarla(self):
        selected_odeme_turu = self.odeme_turu_cb.currentText()
        pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"]
        
        self.islem_hesap_cb.blockSignals(True)
        self.islem_hesap_cb.clear()
        self.kasa_banka_map.clear()
        
        if selected_odeme_turu == self.db.ODEME_TURU_ACIK_HESAP:
            self.islem_hesap_cb.addItem("Hesap Yok", userData=None)
            self.islem_hesap_cb.setEnabled(False)
        elif selected_odeme_turu in pesin_odeme_turleri:
            try:
                hesaplar_response = self.db.kasa_banka_listesi_al(limit=10000)
                hesaplar = []
                if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                    hesaplar = hesaplar_response["items"]
                elif isinstance(hesaplar_response, list):
                    hesaplar = hesaplar_response
                else:
                    hesaplar = []

                if hesaplar:
                    for h in hesaplar:
                        display_text = f"{h.get('hesap_adi')} ({h.get('tip')})"
                        if h.get('tip') == "BANKA" and h.get('banka_adi'):
                            display_text += f" - {h.get('banka_adi')}"
                        self.kasa_banka_map[display_text] = h.get('id')
                        self.islem_hesap_cb.addItem(display_text, h.get('id'))
                    self.islem_hesap_cb.setEnabled(True)
                    self.islem_hesap_cb.setCurrentIndex(0)
                else:
                    self.islem_hesap_cb.addItem("Hesap Yok", None)
                    self.islem_hesap_cb.setEnabled(False)
            except Exception:
                self.islem_hesap_cb.addItem("Hesap Yok", None)
                self.islem_hesap_cb.setEnabled(False)
        else:
            self.islem_hesap_cb.addItem("Hesap Yok", userData=None)
            self.islem_hesap_cb.setEnabled(False)
        
        self.islem_hesap_cb.blockSignals(False)

    def _onayla(self):
        secili_odeme_turu = self.odeme_turu_cb.currentText()
        secili_hesap_display = self.islem_hesap_cb.currentText()
        vade_tarihi_val = self.entry_vade_tarihi.text().strip()
        kasa_banka_id_val = None
        if secili_hesap_display and secili_hesap_display != "Hesap Yok":
            kasa_banka_id_val = self.kasa_banka_map.get(secili_hesap_display)
        if not secili_odeme_turu:
            QMessageBox.critical(self, "Eksik Bilgi", "Lütfen bir Ödeme Türü seçin.")
            return
        pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"]
        if secili_odeme_turu in pesin_odeme_turleri and kasa_banka_id_val is None:
            QMessageBox.critical(self, "Eksik Bilgi", "Peşin ödeme türleri için bir İşlem Kasa/Banka hesabı seçmelisiniz.")
            return
        if secili_odeme_turu == "AÇIK HESAP":
            if not vade_tarihi_val:
                QMessageBox.critical(self, "Eksik Bilgi", "Açık Hesap ödeme türü için Vade Tarihi boş olamaz.")
                return
            try:
                datetime.strptime(vade_tarihi_val, '%Y-%m-%d')
            except ValueError:
                QMessageBox.critical(self, "Tarih Formatı Hatası", "Vade Tarihi formatı (YYYY-AA-GG) olmalıdır.")
                return
        self.callback_func(secili_odeme_turu, kasa_banka_id_val, vade_tarihi_val)
        self.accept()

class TopluVeriEklePenceresi(QDialog):
    def __init__(self, parent=None, db_manager=None, app_ref=None, yenile_callback=None):
        super().__init__(parent)
        self.app = app_ref
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.setWindowTitle("Toplu Veri Ekleme (Excel)")
        self.setMinimumSize(600, 650)
        self.setModal(True)
        self.bekleme_penceresi = None
        self.process_timer = None
        self.process = None
        self.result_queue = multiprocessing.Queue()

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Toplu Veri Ekleme (Excel)")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        main_frame = QFrame(self)
        main_frame_layout = QGridLayout(main_frame)
        main_layout.addWidget(main_frame)

        main_frame_layout.addWidget(QLabel("Veri Tipi:"), 0, 0, Qt.AlignCenter)
        self.veri_tipi_combo = QComboBox()
        self.veri_tipi_combo.addItems(["Müşteri", "Tedarikçi", "Stok/Ürün Ekle/Güncelle"])
        self.veri_tipi_combo.setCurrentText("Müşteri")
        self.veri_tipi_combo.currentIndexChanged.connect(self._show_template_info_and_options)
        main_frame_layout.addWidget(self.veri_tipi_combo, 0, 1)

        main_frame_layout.addWidget(QLabel("Excel Dosyası:"), 1, 0, Qt.AlignCenter)
        self.dosya_yolu_entry = QLineEdit()
        main_frame_layout.addWidget(self.dosya_yolu_entry, 1, 1)
        btn_gozat = QPushButton("Gözat...")
        btn_gozat.clicked.connect(self._gozat_excel_dosyasi)
        main_frame_layout.addWidget(btn_gozat, 1, 2)

        self.stok_guncelleme_options_frame = QGroupBox("Stok/Ürün Güncelleme Seçenekleri", main_frame)
        self.stok_guncelleme_options_layout = QVBoxLayout(self.stok_guncelleme_options_frame)
        main_frame_layout.addWidget(self.stok_guncelleme_options_frame, 2, 0, 1, 3)
        self.stok_guncelleme_options_frame.hide()

        self.cb_vars = {}
        self.cb_vars['fiyat_bilgileri'] = QCheckBox("Fiyat Bilgileri (Alış/Satış/KDV)")
        self.stok_guncelleme_options_layout.addWidget(self.cb_vars['fiyat_bilgileri'])
        self.cb_vars['urun_nitelikleri'] = QCheckBox("Ürün Nitelikleri (Kategori/Marka/Grup/Birim/Menşe/Detay)")
        self.stok_guncelleme_options_layout.addWidget(self.cb_vars['urun_nitelikleri'])
        self.cb_vars['stok_miktari'] = QCheckBox("Stok Miktarı (Mevcut/Minimum)")
        self.stok_guncelleme_options_layout.addWidget(self.cb_vars['stok_miktari'])
        
        self.cb_tumu = QCheckBox("Tümü (Yukarıdakilerin hepsi)")
        self.cb_tumu.stateChanged.connect(self._toggle_all_checkboxes)
        self.stok_guncelleme_options_layout.addWidget(self.cb_tumu)

        # Yeni bir çerçeve oluşturarak etiketi ve butonu hizalıyoruz
        template_info_frame = QFrame(main_frame)
        template_info_layout = QHBoxLayout(template_info_frame)
        template_info_layout.setContentsMargins(0, 0, 0, 0)
        template_info_layout.setSpacing(5)
        
        self.template_info_label = QLabel()
        self.template_info_label.setWordWrap(True)
        self.template_info_label.setAlignment(Qt.AlignCenter)
        template_info_layout.addWidget(self.template_info_label, 1)

        self.detayli_aciklama_button = QPushButton("Detaylı Bilgi / Şablon Açıklaması")
        self.detayli_aciklama_button.clicked.connect(self._show_detayli_aciklama_penceresi)
        template_info_layout.addWidget(self.detayli_aciklama_button)
        self.detayli_aciklama_button.hide()

        main_frame_layout.addWidget(template_info_frame, 3, 0, 1, 3)
        
        main_frame_layout.setColumnStretch(1, 1)

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_yukle = QPushButton("Verileri Yükle")
        btn_yukle.clicked.connect(self._verileri_yukle)
        button_layout.addWidget(btn_yukle)
        
        btn_sablon_indir = QPushButton("Örnek Şablon İndir")
        btn_sablon_indir.clicked.connect(self._excel_sablonu_indir)
        button_layout.addWidget(btn_sablon_indir)
        
        button_layout.addStretch()
        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close)
        button_layout.addWidget(btn_iptal)

        self.analysis_results = None
        self._show_template_info_and_options()
        self.adjustSize()

    def _show_template_info_and_options(self):
        selected_type = self.veri_tipi_combo.currentText()
        short_info_text = ""
        if selected_type == "Stok/Ürün Ekle/Güncelle":
            self.stok_guncelleme_options_frame.show()
            self.detayli_aciklama_button.show()
        else:
            self.stok_guncelleme_options_frame.hide()
            self.detayli_aciklama_button.hide()
            self.cb_tumu.setChecked(False)
            self._toggle_all_checkboxes(Qt.Unchecked, force_off=True)
            
        if selected_type == "Müşteri":
            short_info_text = "Müşteri Excel dosyası:\n`Müşteri Kodu`, `Ad Soyad` (ZORUNLU) ve diğer detaylar."
        elif selected_type == "Tedarikçi":
            short_info_text = "Tedarikçi Excel dosyası:\n`Tedarikçi Kodu`, `Ad Soyad` (ZORUNLU) ve diğer detaylar."
        elif selected_type == "Stok/Ürün Ekle/Güncelle":
            short_info_text = "Stok/Ürün Excel dosyası:\n`Ürün Kodu`, `Ürün Adı` (ZORUNLU) ve diğer detaylar.\nGüncellemek istediğiniz alanları yukarıdan seçin. Detaylı şablon bilgisi için butona tıklayın."
        self.template_info_label.setText(short_info_text)

    def _excel_sablonu_indir(self):
        veri_tipi = self.veri_tipi_combo.currentText()
        if not veri_tipi:
            QMessageBox.warning(self, "Uyarı", "Lütfen şablon indirmek için bir veri tipi seçin.")
            return
        
        file_name_prefix, headers = "", []
        if veri_tipi == "Müşteri": file_name_prefix, headers = "Musteri_Sablonu", ["Müşteri Kodu", "Ad Soyad", "Telefon", "Adres", "Vergi Dairesi", "Vergi No"]
        elif veri_tipi == "Tedarikçi": file_name_prefix, headers = "Tedarikci_Sablonu", ["Tedarikçi Kodu", "Ad Soyad", "Telefon", "Adres", "Vergi Dairesi", "Vergi No"]
        elif veri_tipi == "Stok/Ürün Ekle/Güncelle": file_name_prefix, headers = "Stok_Urun_Sablonu", ["Ürün Kodu", "Ürün Adı", "Miktar", "Alış Fiyatı (KDV Dahil)", "Satış Fiyatı (KDV Dahil)", "KDV Oranı (%)", "Minimum Stok Seviyesi", "Kategori Adı", "Marka Adı", "Ürün Grubu Adı", "Ürün Birimi Adı", "Menşe Ülke Adı", "Ürün Detayı", "Ürün Resmi Yolu"]
        else:
            QMessageBox.critical(self, "Hata", "Geçersiz veri tipi seçimi.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Excel Şablonunu Kaydet",
                                                    f"{file_name_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                                    "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)")
        if file_path:
            try:
                workbook = openpyxl.Workbook(); sheet = workbook.active; sheet.title = "Veri Şablonu"; sheet.append(headers)
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_idx); cell.font = openpyxl.styles.Font(bold=True)
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) + 2, 15)
                workbook.save(file_path)
                QMessageBox.information(self, "Başarılı", f"'{veri_tipi}' şablonu başarıyla oluşturuldu:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Şablon oluşturulurken bir hata oluştu: {e}")

    def _show_detayli_aciklama_penceresi(self):
        selected_type = self.veri_tipi_combo.currentText()
        title = f"{selected_type} Şablon Açıklaması"
        message = ""
        if selected_type == "Müşteri": message = "Müşteri Veri Şablonu Detayları:\n\nExcel dosyasının ilk satırı başlık (header) olmalıdır. Veriler ikinci satırdan başlamalıdır.\n\nSütun Sırası ve Açıklamaları:\n1.  **Müşteri Kodu (ZORUNLU):** Müşterinin benzersiz kodu.\n2.  **Ad Soyad (ZORUNLU):** Müşterinin tam adı veya şirket adı.\n3.  **Telefon (İsteğe Bağlı)**\n4.  **Adres (İsteğe Bağlı)**\n5.  **Vergi Dairesi (İsteğe Bağlı)**\n6.  **Vergi No (İsteğe Bağlı)**"
        elif selected_type == "Tedarikçi": message = "Tedarikçi Veri Şablonu Detayları:\n\nExcel dosyasının ilk satırı başlık (header) olmalıdır. Veriler ikinci satırdan başlamalıdır.\n\nSütun Sırası ve Açıklamaları:\n1.  **Tedarikçi Kodu (ZORUNLU):** Tedarikçinin benzersiz kodu.\n2.  **Ad Soyad (ZORUNLU):** Tedarikçinin tam adı veya şirket adı.\n3.  **Telefon (İsteğe Bağlı)**\n4.  **Adres (İsteğe Bağlı)**\n5.  **Vergi Dairesi (İsteğe Bağlı)**\n6.  **Vergi No (İsteğe Bağlı)**"
        elif selected_type == "Stok/Ürün Ekle/Güncelle": message = "Stok/Ürün Veri Şablonu Detayları:\n\n'Ürün Kodu' eşleşirse güncelleme, eşleşmezse yeni kayıt yapılır.\n\nSütunlar:\n1.  **Ürün Kodu (ZORUNLU)**\n2.  **Ürün Adı (Yeni ürün için ZORUNLU)**\n3.  **Miktar (İsteğe Bağlı):** Pozitif girilirse, mevcut stoğa eklemek için bir 'ALIŞ' faturası oluşturulur.\n4.  **Alış Fiyatı (KDV Dahil) (İsteğe Bağlı)**\n5.  **Satış Fiyatı (KDV Dahil) (İsteğe Bağlı)**\n6.  **KDV Oranı (%) (İsteğe Bağlı)**\n7.  **Minimum Stok Seviyesi (İsteğe Bağlı)**\n8.  **Kategori Adı (İsteğe Bağlı)**\n9.  **Marka Adı (İsteğe Bağlı)**\n10. **Ürün Grubu Adı (İsteğe Bağlı)**\n11. **Ürün Birimi Adı (İsteğe Bağlı)**\n12. **Menşe Ülke Adı (İsteğe Bağlı)**\n13. **Ürün Detayı (İsteğe Bağlı)**\n14. **Ürün Resmi Yolu (İsteğe Bağlı):** Resim dosyasının tam yolu (ör: C:/resimler/urun1.png)."
        from pencereler import AciklamaDetayPenceresi
        AciklamaDetayPenceresi(self, title, message).exec()

    def _gozat_excel_dosyasi(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Excel Dosyası Seç", "", "Excel Dosyaları (*.xlsx;*.xls);;Tüm Dosyalar (*)")
        if file_path:
            self.dosya_yolu_entry.setText(file_path)

    def _toggle_all_checkboxes(self, state, force_off=False):
        is_checked = (state == Qt.Checked) if not force_off else False
        for key, checkbox in self.cb_vars.items():
            checkbox.setChecked(is_checked)

    def _verileri_yukle(self):
        dosya_yolu = self.dosya_yolu_entry.text().strip()
        veri_tipi = self.veri_tipi_combo.currentText()
        if not dosya_yolu or not os.path.exists(dosya_yolu):
            QMessageBox.critical(self, "Dosya Hatası", "Lütfen geçerli bir Excel dosyası seçin.")
            return

        from pencereler import BeklemePenceresi
        # Bağımsız pencere olarak açmak için parent parametresini None olarak ayarlayın.
        self.bekleme_penceresi = BeklemePenceresi(None, message="Excel okunuyor ve veriler analiz ediliyor...")
        self.bekleme_penceresi.show()
        
        QApplication.instance().processEvents()
        
        selected_update_fields = []
        if veri_tipi == "Stok/Ürün Ekle/Güncelle":
            if self.cb_tumu.isChecked():
                selected_update_fields = ['fiyat_bilgileri', 'urun_nitelikleri', 'stok_miktari']
            else:
                for key, cb in self.cb_vars.items():
                    if cb.isChecked():
                        selected_update_fields.append(key)
        
        self.result_queue = multiprocessing.Queue()
        
        self.process = multiprocessing.Process(
            target=self._analiz_et_ve_onizle_process,
            args=(self.db.api_base_url, dosya_yolu, veri_tipi, selected_update_fields, self.result_queue)
        )
        self.process.start()
        
        self.process_timer = QTimer(self)
        self.process_timer.timeout.connect(self._check_analysis_completion)
        self.process_timer.start(100)

    @staticmethod
    def _analiz_et_ve_onizle_process(api_base_url, dosya_yolu, veri_tipi, selected_update_fields, result_queue):
        """Excel dosyasını ayrı bir işlemde okur ve analiz sonuçlarını döndürür."""
        import openpyxl
        import logging
        import traceback
        import multiprocessing
        import csv

        logger = logging.getLogger(__name__)

        try:
            raw_data_from_excel_list = []
            file_extension = os.path.splitext(dosya_yolu)[1].lower()

            if file_extension in ['.xlsx', '.xls']:
                workbook = openpyxl.load_workbook(dosya_yolu, data_only=True)
                sheet = workbook.active
                
                for row_obj in sheet.iter_rows(min_row=2):
                    if any(cell.value is not None and str(cell.value).strip() != '' for cell in row_obj):
                        row_values = [cell.value for cell in row_obj]
                        raw_data_from_excel_list.append(row_values)
            
            elif file_extension == '.csv':
                with open(dosya_yolu, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    header_skipped = False
                    for row_values in reader:
                        if not header_skipped:
                            header_skipped = True
                            continue
                        if any(cell is not None and str(cell).strip() != '' for cell in row_values):
                            raw_data_from_excel_list.append(row_values)
            else:
                raise ValueError("Desteklenmeyen dosya formatı. Lütfen .xlsx, .xls veya .csv kullanın.")

            if not raw_data_from_excel_list:
                raise ValueError("Excel dosyasında okunacak geçerli veri bulunamadı.")
            
            from hizmetler import TopluIslemService
            from veritabani import OnMuhasebe
            
            local_db_manager = OnMuhasebe(api_base_url=api_base_url)
            local_toplu_islem_service = TopluIslemService(local_db_manager)

            if veri_tipi == "Müşteri":
                analysis_results = local_toplu_islem_service.toplu_musteri_analiz_et(raw_data_from_excel_list)
            elif veri_tipi == "Tedarikçi":
                analysis_results = local_toplu_islem_service.toplu_tedarikci_analiz_et(raw_data_from_excel_list)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                analysis_results = local_toplu_islem_service.toplu_stok_analiz_et(raw_data_from_excel_list, selected_update_fields)
            
            result_queue.put({"success": True, "results": analysis_results})

        except Exception as e:
            result_queue.put({"success": False, "error": str(e), "traceback": traceback.format_exc()})
            
    def _check_analysis_completion(self):
        if not self.result_queue.empty():
            self.process_timer.stop()
            self.bekleme_penceresi.hide()
            
            result = self.result_queue.get()
            
            def handle_completion():
                self.bekleme_penceresi.close()
                if result['success']:
                    self._onizleme_penceresini_ac(self.veri_tipi_combo.currentText(), result['results'])
                else:
                    QMessageBox.critical(self, "Hata", f"Veri analizi başarısız oldu:\n{result['error']}")
                    logging.error(f"Toplu veri analizi sürecinde hata: {result['error']}\n{result['traceback']}")
                self.process.join()

            QTimer.singleShot(0, handle_completion)
            
        elif not self.process.is_alive():
            self.process_timer.stop()
            self.bekleme_penceresi.close()
            QMessageBox.critical(self, "Hata", "Veri analizi süreci beklenmedik bir şekilde sonlandı.")
            self.process.join()

    def _onizleme_penceresini_ac(self, veri_tipi, analysis_results):
        try:
            dialog = TopluVeriOnizlemePenceresi(self, self.db, veri_tipi, analysis_results, 
                                                callback_on_confirm=self._yazma_islemi_from_onizleme)
            dialog.exec()
        except Exception as e:
            logging.error(f"Önizleme penceresi açılırken bir hata oluştu: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Önizleme penceresi açılırken beklenmedik bir hata oluştu:\n{e}")

    def _yazma_islemi_from_onizleme(self, veri_tipi, analysis_results):
        from pencereler import BeklemePenceresi
        # Bağımsız pencere olarak açmak için parent parametresini None olarak ayarlayın.
        self.bekleme_penceresi = BeklemePenceresi(None, message=f"Toplu {veri_tipi} veritabanına yazılıyor, lütfen bekleyiniz...")
        self.bekleme_penceresi.show()
        
        self.result_queue = multiprocessing.Queue()
        
        self.process = multiprocessing.Process(
            target=self._yazma_islemi_process,
            args=(self.db.api_base_url, veri_tipi, analysis_results, self.result_queue)
        )
        self.process.start()
        
        self.process_timer = QTimer(self)
        self.process_timer.timeout.connect(self._check_writing_completion)
        self.process_timer.start(100)

    @staticmethod
    def _yazma_islemi_process(api_base_url, veri_tipi, analysis_results, result_queue):
        try:
            from hizmetler import TopluIslemService
            from veritabani import OnMuhasebe
            
            local_db_manager = OnMuhasebe(api_base_url=api_base_url)
            local_toplu_islem_service = TopluIslemService(local_db_manager)

            result = {"yeni_eklenen_sayisi": 0, "guncellenen_sayisi": 0, "hata_sayisi": 0, "hatalar": []}
            
            # Düzeltme: API'ye gönderilecek veriyi, analiz sonucunda elde edilen
            # doğru formatta (sözlükler listesi) alıyoruz.
            data_to_import = analysis_results.get('new_records', []) + analysis_results.get('update_records', [])

            if veri_tipi == "Müşteri":
                result = local_toplu_islem_service.toplu_musteri_ice_aktar(data_to_import)
            elif veri_tipi == "Tedarikçi":
                result = local_toplu_islem_service.toplu_tedarikci_ice_aktar(data_to_import)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                try:
                    result = local_toplu_islem_service.toplu_stok_ice_aktar(data_to_import)
                except Exception as e:
                    result = {"yeni_eklenen_sayisi": 0, "guncellenen_sayisi": 0, "hata_sayisi": len(data_to_import), "hatalar": [f"Toplu işlem sırasında hata: {e}"]}
            
            result_queue.put({"success": True, "results": result})
            
        except Exception as e:
            result_queue.put({"success": False, "error": str(e), "traceback": traceback.format_exc()})

    def _check_writing_completion(self):
        if not self.result_queue.empty():
            self.process_timer.stop()
            self.bekleme_penceresi.hide()

            result = self.result_queue.get()

            def handle_completion():
                self.bekleme_penceresi.close()

                if result['success']:
                    final_results = result['results']
                    success_count = final_results.get('yeni_eklenen_sayisi', 0) + final_results.get('guncellenen_sayisi', 0)
                    error_count = final_results.get('hata_sayisi', 0)
                    
                    if error_count == 0:
                        message = f"Toplu {self.veri_tipi_combo.currentText()} işlemi başarıyla tamamlandı. {success_count} kayıt işlendi."
                        QMessageBox.information(self, "Başarılı", message)
                        self._refresh_related_lists(self.veri_tipi_combo.currentText())
                        self.accept()
                    else:
                        message = f"Toplu {self.veri_tipi_combo.currentText()} işlemi kısmen tamamlandı. {success_count} kayıt başarılı, {error_count} kayıt hata verdi."
                        QMessageBox.warning(self, "Uyarı", message)
                        self._refresh_related_lists(self.veri_tipi_combo.currentText())
                        self.accept()
                else:
                    QMessageBox.critical(self, "Kritik Hata", f"İşlem sırasında beklenmedik bir hata oluştu:\n{result['error']}")
                    logging.error(f"Toplu yazma işlemi sürecinde hata: {result['error']}\n{result['traceback']}")
                
                self.process.join()

            QTimer.singleShot(0, handle_completion)
        elif not self.process.is_alive():
            self.process_timer.stop()
            self.bekleme_penceresi.close()
            QMessageBox.critical(self, "Hata", "Veri yazma süreci beklenmedik bir şekilde sonlandı.")
            self.process.join()

    def _refresh_related_lists(self, veri_tipi):
        if veri_tipi == "Müşteri" and hasattr(self.app, 'musteri_yonetimi_sayfasi'):
            self.app.musteri_yonetimi_sayfasi.musteri_listesini_yenile()
        elif veri_tipi == "Tedarikçi" and hasattr(self.app, 'tedarikci_yonetimi_sayfasi'):
            self.app.tedarikci_yonetimi_sayfasi.tedarikci_listesini_yenile()
        elif veri_tipi == "Stok/Ürün Ekle/Güncelle" and hasattr(self.app, 'stok_yonetimi_sayfasi'):
            self.app.stok_yonetimi_sayfasi.stok_listesini_yenile()
        if hasattr(self.app, 'ana_sayfa'):
            self.app.ana_sayfa.guncelle_ozet_bilgiler()

class TopluVeriOnizlemePenceresi(QDialog):
    def __init__(self, parent_app, db_manager, veri_tipi, analysis_results, callback_on_confirm):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.veri_tipi = veri_tipi
        self.analysis_results = analysis_results
        self.callback_on_confirm = callback_on_confirm

        self.setWindowTitle(f"Toplu {veri_tipi} İçe Aktarım Önizlemesi")
        self.setFixedSize(1200, 750) # Daha fazla veri göstermek için boyutu büyüttük
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel(f"Toplu {veri_tipi} İçe Aktarım Önizlemesi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        info_label_text = f"<b>Yeni Eklenecek Kayıt Sayısı:</b> {len(self.analysis_results.get('new_records', []))}<br>" \
                          f"<b>Güncellenecek Kayıt Sayısı:</b> {len(self.analysis_results.get('update_records', []))}<br>" \
                          f"<b>Hata Veren Kayıt Sayısı:</b> {len(self.analysis_results.get('error_records', []))}"
        info_label = QLabel(info_label_text)
        main_layout.addWidget(info_label)
        
        # Sütun başlıklarını veri tipine göre dinamik olarak belirliyoruz
        if self.veri_tipi == "Müşteri" or self.veri_tipi == "Tedarikçi":
            headers = ["Durum", "Kod", "Ad Soyad", "Telefon", "Adres", "Vergi Dairesi", "Vergi No", "Hata Detayı"]
        elif self.veri_tipi == "Stok/Ürün Ekle/Güncelle":
            headers = ["Durum", "Kod", "Ad", "Miktar", "Alış Fyt", "Satış Fyt", "KDV %", "Min.Stok", "Hata Detayı"]

        self.tree_preview = QTreeWidget()
        self.tree_preview.setHeaderLabels(headers)
        self.tree_preview.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tree_preview.setSortingEnabled(True) # Sıralama özelliğini ekledik

        # Sütun genişliklerini ve hizalamalarını ayarladık
        self.tree_preview.header().setSectionResizeMode(0, QHeaderView.ResizeToContents) # Durum
        self.tree_preview.header().setSectionResizeMode(1, QHeaderView.ResizeToContents) # Kod
        self.tree_preview.header().setSectionResizeMode(2, QHeaderView.Stretch) # Ad
        self.tree_preview.header().setSectionResizeMode(len(headers) - 1, QHeaderView.Stretch) # Son sütun genişlesin

        main_layout.addWidget(self.tree_preview)

        self._populate_preview_tree()

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        btn_onayla = QPushButton("Onayla ve İçe Aktar")
        btn_onayla.clicked.connect(self._onayla)
        if len(self.analysis_results.get('error_records', [])) > 0:
            btn_onayla.setEnabled(False)
        button_layout.addWidget(btn_onayla)
        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.reject)
        button_layout.addWidget(btn_iptal)

    def _populate_preview_tree(self):
        """Önizleme verilerini QTreeWidget'a yükler."""
        self.tree_preview.clear()
        
        # Orijinal data'yı analysis_results'tan alıyoruz
        all_data = self.analysis_results.get('all_processed_data', [])
        
        # Hata veren kayıtların satır numaralarını tutan bir sözlük oluşturuyoruz
        error_rows = {r.get('satir')-2: r.get('hata') for r in self.analysis_results.get('error_records', [])}
        
        # Yeni eklenecek ve güncellenecek kayıtların kodlarını tutan set'ler oluşturuyoruz
        new_codes = {r.get('kod') for r in self.analysis_results.get('new_records', [])}
        update_codes = {r.get('kod') for r in self.analysis_results.get('update_records', [])}
        
        # Gelen verinin yapısını kontrol etmek için debug çıktısı
        logger.debug(f"Önizleme penceresi için gelen veri (all_data): {all_data}")

        for row_index, record in enumerate(all_data):
            item = QTreeWidgetItem(self.tree_preview)

            record_kod = ""
            if isinstance(record, dict):
                record_kod = record.get('kod', '')
            elif isinstance(record, list) and len(record) > 0:
                record_kod = str(record[0]) if record[0] is not None else ''

            # Hata kontrolü
            if row_index in error_rows:
                item.setText(0, "Hata")
                item.setForeground(0, QBrush(QColor("red")))
                item.setText(self.tree_preview.header().count() - 1, error_rows[row_index])
            # Durum belirleme
            elif record_kod in new_codes:
                item.setText(0, "Yeni")
                item.setForeground(0, QBrush(QColor("blue")))
            elif record_kod in update_codes:
                item.setText(0, "Güncelle")
                item.setForeground(0, QBrush(QColor("orange")))

            # Veri tipine göre sütunları doldurma
            if self.veri_tipi == "Müşteri" or self.veri_tipi == "Tedarikçi":
                if isinstance(record, dict):
                    item.setText(1, record.get('kod', ''))
                    item.setText(2, record.get('ad', ''))
                    item.setText(3, record.get('telefon', ''))
                    item.setText(4, record.get('adres', ''))
                    item.setText(5, record.get('vergi_dairesi', ''))
                    item.setText(6, record.get('vergi_no', ''))
                elif isinstance(record, list) and len(record) > 2:
                    item.setText(1, str(record[0]) if record[0] is not None else '')
                    item.setText(2, str(record[1]) if record[1] is not None else '')
                    item.setText(3, str(record[2]) if len(record) > 2 and record[2] is not None else '')
                    item.setText(4, str(record[3]) if len(record) > 3 and record[3] is not None else '')
                    item.setText(5, str(record[4]) if len(record) > 4 and record[4] is not None else '')
                    item.setText(6, str(record[5]) if len(record) > 5 and record[5] is not None else '')
            elif self.veri_tipi == "Stok/Ürün Ekle/Güncelle":
                if isinstance(record, dict):
                    item.setText(1, record.get('kod', ''))
                    item.setText(2, record.get('ad', ''))
                    item.setText(3, str(record.get('miktar', '')))
                    item.setText(4, str(record.get('alis_fiyati', '')))
                    item.setText(5, str(record.get('satis_fiyati', '')))
                    item.setText(6, str(record.get('kdv_orani', '')))
                    item.setText(7, str(record.get('min_stok_seviyesi', '')))
                elif isinstance(record, list) and len(record) > 2:
                    item.setText(1, str(record[0]) if record[0] is not None else '')
                    item.setText(2, str(record[1]) if record[1] is not None else '')
                    item.setText(3, str(record[2]) if len(record) > 2 and record[2] is not None else '')
                    item.setText(4, str(record[3]) if len(record) > 3 and record[3] is not None else '')
                    item.setText(5, str(record[4]) if len(record) > 4 and record[4] is not None else '')
                    item.setText(6, str(record[5]) if len(record) > 5 and record[5] is not None else '')
                    item.setText(7, str(record[6]) if len(record) > 6 and record[6] is not None else '')

    def _onayla(self):
        if self.callback_on_confirm:
            self.callback_on_confirm(self.veri_tipi, self.analysis_results)
        
        if len(self.analysis_results.get('error_records', [])) == 0:
            self.accept()

    def _analiz_et_ve_onizle_threaded(self, dosya_yolu, veri_tipi, selected_update_fields):
        analysis_results = {}
        try:
            workbook = openpyxl.load_workbook(dosya_yolu, data_only=True)
            sheet = workbook.active
            
            raw_data_from_excel_list = []
            for row_obj in sheet.iter_rows(min_row=2):
                if any(cell.value is not None and str(cell.value).strip() != '' for cell in row_obj):
                    row_values = [str(cell.value).strip() if cell.value is not None else None for cell in row_obj]
                    raw_data_from_excel_list.append(row_values)

            if not raw_data_from_excel_list:
                raise ValueError("Excel dosyasında okunacak geçerli veri bulunamadı.")
            
            from hizmetler import TopluIslemService
            local_db_manager = self.db.__class__(api_base_url=self.db.api_base_url)
            local_toplu_islem_service = TopluIslemService(local_db_manager)

            if veri_tipi == "Müşteri":
                analysis_results = local_toplu_islem_service.toplu_musteri_analiz_et(raw_data_from_excel_list)
            elif veri_tipi == "Tedarikçi":
                analysis_results = local_toplu_islem_service.toplu_tedarikci_analiz_et(raw_data_from_excel_list)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                analysis_results = local_toplu_islem_service.toplu_stok_analiz_et(raw_data_from_excel_list, selected_update_fields)
            
            # Ana iş parçacığında bekleme penceresini kapat ve önizlemeyi göster
            QTimer.singleShot(0, self.bekleme_penceresi.close)
            QTimer.singleShot(0, lambda: self._onizleme_penceresini_ac(veri_tipi, analysis_results))

        except Exception as e:
            # Ana iş parçacığında hata mesajını göster ve bekleme penceresini kapat
            QTimer.singleShot(0, self.bekleme_penceresi.close)
            QTimer.singleShot(0, lambda: QMessageBox.critical(self, "Hata", f"Veri analizi başarısız oldu:\n{e}"))
            logging.error(f"Toplu veri analizi thread'inde hata: {e}", exc_info=True)

    def _onizleme_penceresini_ac(self, veri_tipi, analysis_results):
        # Hatalı import satırı
        # from pencereler import TopluVeriOnizlemePenceresi
        dialog = TopluVeriOnizlemePenceresi(self.app, self.db, veri_tipi, analysis_results, 
                                            callback_on_confirm=self._yazma_islemi_threaded_from_onizleme)
        dialog.exec()

    def _yazma_islemi_threaded_from_onizleme(self, veri_tipi, analysis_results):
        from pencereler import BeklemePenceresi
        self.bekleme_penceresi_yazma = BeklemePenceresi(self, message=f"Toplu {veri_tipi} veritabanına yazılıyor, lütfen bekleyiniz...")
        self.bekleme_penceresi_yazma.show()

        threading.Thread(target=self._yazma_islemi_threaded, args=(
            veri_tipi, 
            analysis_results
        )).start()

    def _yazma_islemi_threaded(self, veri_tipi, analysis_results):
        success = False
        message = ""
        try:
            from hizmetler import TopluIslemService
            local_db_manager = self.db.__class__(api_base_url=self.db.api_base_url, app_ref=self.app)
            local_toplu_islem_service = TopluIslemService(local_db_manager)

            result = {"yeni_eklenen_sayisi": 0, "guncellenen_sayisi": 0, "hata_sayisi": 0, "hatalar": []}
            data_to_import = analysis_results.get('all_processed_data', [])

            if veri_tipi == "Müşteri":
                result = local_toplu_islem_service.toplu_musteri_ice_aktar(data_to_import)
            elif veri_tipi == "Tedarikçi":
                result = local_toplu_islem_service.toplu_tedarikci_ice_aktar(data_to_import)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                try:
                    response = local_db_manager.bulk_stok_upsert(data_to_import)
                    result = {
                        "yeni_eklenen_sayisi": response.get("yeni_eklenen_sayisi", 0),
                        "guncellenen_sayisi": response.get("guncellenen_sayisi", 0),
                        "hata_sayisi": response.get("hata_sayisi", 0),
                        "hatalar": response.get("hatalar", [])
                    }
                except Exception as e:
                    result = {"yeni_eklenen_sayisi": 0, "guncellenen_sayisi": 0, "hata_sayisi": len(data_to_import), "hatalar": [f"Toplu işlem sırasında hata: {e}"]}

            success = result["hata_sayisi"] == 0
            if success:
                message = f"{result['yeni_eklenen_sayisi']} yeni {veri_tipi} başarıyla eklendi, {result['guncellenen_sayisi']} güncellendi."
            else:
                message = f"Hata: {result['yeni_eklenen_sayisi']} yeni {veri_tipi} başarıyla eklendi, {result['guncellenen_sayisi']} güncellendi ancak {result['hata_sayisi']} kayıt hata verdi:\n" + "\n".join(result["hatalar"])

            # UI güncellemelerini ana iş parçacığına taşı
            QTimer.singleShot(0, lambda: self.bekleme_penceresi_yazma.close())
            if success:
                QTimer.singleShot(0, lambda: QMessageBox.information(self, "Başarılı", message))
                QTimer.singleShot(0, lambda: self._refresh_related_lists(veri_tipi))
                QTimer.singleShot(0, self.accept)
            else:
                QTimer.singleShot(0, lambda: QMessageBox.critical(self, "Hata", f"Toplu {veri_tipi} işlemi başarısız oldu:\n{message}"))

        except Exception as e:
            QTimer.singleShot(0, lambda: self.bekleme_penceresi_yazma.close())
            QTimer.singleShot(0, lambda: QMessageBox.critical(self, "Kritik Hata", f"İşlem sırasında beklenmedik bir hata oluştu: {e}"))
            logging.error(f"Toplu yazma işlemi thread'inde hata: {e}", exc_info=True)
        finally:
            pass

class AciklamaDetayPenceresi(QDialog):
    def __init__(self, parent_app, title="Detaylı Bilgi", message_text=""):
        super().__init__(parent_app)
        self.setWindowTitle(title)
        self.setFixedSize(600, 400) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        # Pencereyi ortalamak için
        self.move(parent_app.pos() + parent_app.rect().center() - self.rect().center())

        main_layout = QVBoxLayout(self)
        self.text_widget = QTextEdit() # tk.Text yerine QTextEdit kullanıldı
        self.text_widget.setPlainText(message_text)
        self.text_widget.setReadOnly(True) # config(state=tk.DISABLED) yerine setReadOnly
        
        main_layout.addWidget(self.text_widget)

        # QScrollArea içinde QTextEdit otomatik kaydırma çubuklarını yönetir, ek scrollbar gerekmez
        # tk.Text'teki vsb kısmı kaldırıldı

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close) # QDialog'u kapat
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignCenter) # Ortala

class CariSecimPenceresi(QDialog):
    # 'callback_func' parametresi artık init metodunda bulunmuyor.
    def __init__(self, parent_window, db_manager, cari_selection_type):
        super().__init__(parent_window)
        # DEĞİŞİKLİK: app referansını doğrudan parent_window olarak atıyoruz.
        # Çünkü FaturaOlusturmaSayfasi'ndan gelen parent_window zaten ana App nesnesidir.
        self.app = parent_window 
        self.db = db_manager
        
        self.cari_to_select_type = cari_selection_type
        # Seçilen cari verilerini saklamak için sınıf özelliklerini tanımlıyoruz.
        self.secili_cari_id = None
        self.secili_cari_adi = None

        self.setWindowTitle("Cari Seçimi")
        self.setFixedSize(600, 450)
        self.setModal(True)

        self.tum_cariler_cache_data = []

        if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI:
            baslik_text = "Müşteri Seçimi"
        elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI:
            baslik_text = "Tedarikçi Seçimi"
        else:
            baslik_text = "Cari Seçimi (Bilinmeyen Tip)"
            logging.warning(f"CariSecimPenceresi bilinmeyen tip ile başlatıldı: {cari_selection_type}")

        title_label = QLabel(baslik_text)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(title_label)

        # BURASI GÜNCELLENDİ: Arama çubuğu düzeni
        search_frame = QFrame(self)
        search_layout = QHBoxLayout(search_frame)
        main_layout.addWidget(search_frame)

        # 'Ara (Ad/Kod):' yazısını sola hizalı olarak yerleştiriyoruz.
        search_layout.addWidget(QLabel("Ara (Ad/Kod):")) 
        self.search_entry = QLineEdit()
        self.search_entry.textChanged.connect(self._filtre_liste)
        
        # Arama kutusunun metin hizalamasını sola ayarlıyoruz ve minimum genişlik veriyoruz.
        self.search_entry.setAlignment(Qt.AlignLeft)
        self.search_entry.setMinimumWidth(200)

        search_layout.addWidget(self.search_entry)
        search_layout.setStretchFactor(self.search_entry, 1)

        tree_frame = QFrame(self)
        tree_layout = QVBoxLayout(tree_frame)
        main_layout.addWidget(tree_frame)

        self.cari_tree = QTreeWidget()
        self.cari_tree.setHeaderLabels(["Cari Adı", "Kodu"])
        self.cari_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.cari_tree.setSortingEnabled(True)

        self.cari_tree.setColumnWidth(0, 300)
        self.cari_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.cari_tree.setColumnWidth(1, 100)
        self.cari_tree.headerItem().setTextAlignment(1, Qt.AlignCenter)

        tree_layout.addWidget(self.cari_tree)

        self.cari_tree.itemDoubleClicked.connect(self._sec)

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_sec = QPushButton("Seç")
        btn_sec.clicked.connect(self._sec)
        button_layout.addWidget(btn_sec)

        button_layout.addStretch()

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.reject)
        button_layout.addWidget(btn_iptal)

        self._yukle_carileri()
        self.search_entry.setFocus()

    def _sec(self):
        """Seçili cariyi kaydeder ve diyalog penceresini kapatır."""
        selected_items = self.cari_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Seçim Yok", "Lütfen bir cari seçin.")
            return

        selected_item = selected_items[0]
        # Seçilen verileri sınıf özelliklerine kaydet
        self.secili_cari_id = selected_item.data(0, Qt.UserRole)
        self.secili_cari_adi = selected_item.text(0)
        
        # Diyalogu kapat ve sonucu ACCEPTED olarak işaretle
        self.accept()

    def _yukle_carileri(self):
        self.tum_cariler_cache_data = []

        try:
            kod_anahtari_db = ''
            if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI:
                cariler_response = self.db.musteri_listesi_al(limit=10000)
                kod_anahtari_db = 'kod'
            elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI:
                cariler_response = self.db.tedarikci_listesi_al(limit=10000)
                kod_anahtari_db = 'tedarikci_kodu'
            else:
                self.app.set_status_message("Hata: CariSecimPenceresi için geçersiz tip belirtildi.", "red")
                logging.error(f"CariSecimPenceresi._yukle_carileri: Geçersiz cari_to_select_type: {self.cari_to_select_type}")
                return

            cariler = cariler_response.get("items", [])
            if isinstance(cariler_response, list):
                cariler = cariler_response
                self.app.set_status_message("Uyarı: Cari listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")

            for c in cariler:
                cari_id = c.get('id')
                cari_ad = c.get('ad')
                cari_kodu = c.get(kod_anahtari_db, "")
                self.tum_cariler_cache_data.append(c)

            self._filtre_liste()
            default_id_str = None
            if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI and self.db.get_perakende_musteri_id() is not None:
                default_id_str = str(self.db.get_perakende_musteri_id())
            elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI and self.db.get_genel_tedarikci_id() is not None:
                default_id_str = str(self.db.get_genel_tedarikci_id())

            if default_id_str:
                for i in range(self.cari_tree.topLevelItemCount()):
                    item = self.cari_tree.topLevelItem(i)
                    if item.data(0, Qt.UserRole) == int(default_id_str):
                        item.setSelected(True)
                        self.cari_tree.scrollToItem(item)
                        break

        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Cari listesi çekilirken hata: {e}")
            logging.error(f"Cari listesi yükleme hatası: {e}", exc_info=True)

    def _filtre_liste(self):
        """
        Arama kutusuna yazıldıkça cari listesini filtreler.
        """
        # Arama terimini al ve normalleştirerek küçük harfe dönüştür.
        arama_terimi = self.search_entry.text().strip()
        normalized_arama_terimi = normalize_turkish_chars(arama_terimi).lower()

        self.cari_tree.clear()

        for cari_row in self.tum_cariler_cache_data:
            cari_id = cari_row.get('id')
            cari_ad = cari_row.get('ad')
            cari_kodu = ""
            if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI:
                cari_kodu = cari_row.get('kod', '')
            elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI:
                # Tedarikçi kodu için `kod` anahtarını kullanıyoruz, `tedarikci_kodu` değil.
                cari_kodu = cari_row.get('kod', '')
            
            # Karşılaştırılacak verileri de normalleştirip küçük harfe dönüştürüyoruz.
            normalized_cari_ad = normalize_turkish_chars(cari_ad).lower() if cari_ad else ''
            normalized_cari_kodu = normalize_turkish_chars(cari_kodu).lower() if cari_kodu else ''
            
            # Arama filtresi mantığı
            if (not normalized_arama_terimi or
                (normalized_cari_ad and normalized_arama_terimi in normalized_cari_ad) or
                (normalized_cari_kodu and normalized_arama_terimi in normalized_cari_kodu)
               ):
                item_qt = QTreeWidgetItem(self.cari_tree)
                item_qt.setText(0, cari_ad)
                item_qt.setText(1, cari_kodu)
                item_qt.setData(0, Qt.UserRole, cari_id)

class TedarikciSecimDialog(QDialog):
    def __init__(self, parent_window, db_manager, callback_func):
        super().__init__(parent_window) 
        self.app = parent_window.app # parent_window'un içindeki app referansını al
        self.db = db_manager
        self.callback_func = callback_func

        self.setWindowTitle("Tedarikçi Seçimi")
        self.setFixedSize(600, 400) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        self.tum_tedarikciler_cache = [] # Data dict'lerini saklar

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Tedarikçi Seçimi")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # Arama Çerçevesi
        search_frame = QFrame(self)
        search_layout = QHBoxLayout(search_frame)
        main_layout.addWidget(search_frame)

        search_layout.addWidget(QLabel("Ara (Ad/Kod):"), Qt.AlignCenter)
        self.search_entry = QLineEdit()
        self.search_entry.textChanged.connect(self._filtre_liste)
        search_layout.addWidget(self.search_entry)
        search_layout.setStretchFactor(self.search_entry, 1) # Genişlemesi için

        # Tedarikçi Listesi Treeview
        tree_frame = QFrame(self)
        tree_layout = QVBoxLayout(tree_frame)
        main_layout.addWidget(tree_frame)

        self.tedarikci_tree = QTreeWidget()
        self.tedarikci_tree.setHeaderLabels(["Tedarikçi Adı", "Kodu"])
        self.tedarikci_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tedarikci_tree.setSortingEnabled(True)

        self.tedarikci_tree.setColumnWidth(0, 300) # Tedarikçi Adı sütun genişliği
        self.tedarikci_tree.header().setSectionResizeMode(0, QHeaderView.Stretch) # Tedarikçi Adı genişlesin
        self.tedarikci_tree.setColumnWidth(1, 100) # Kodu sütun genişliği
        self.tedarikci_tree.headerItem().setTextAlignment(1, Qt.AlignCenter) # Kodu sütununu ortala

        tree_layout.addWidget(self.tedarikci_tree)
        
        self.tedarikci_tree.itemDoubleClicked.connect(self._sec) # Çift tıklama ile seçim

        # Butonlar
        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_sec = QPushButton("Seç")
        btn_sec.clicked.connect(self._sec)
        button_layout.addWidget(btn_sec)
        
        button_layout.addStretch() # Sağ tarafa yaslamak için boşluk

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

        # Başlangıç yüklemesi
        self._yukle_tedarikcileri()
        self.search_entry.setFocus()
    
    def _yukle_tedarikcileri(self):
        """Tüm tedarikçileri API'den çeker ve listeler."""
        self.tum_tedarikciler_cache = [] 

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            tedarikciler_response = self.db.tedarikci_listesi_al(limit=10000)

            tedarikciler = []
            if isinstance(tedarikciler_response, dict) and "items" in tedarikciler_response:
                tedarikciler = tedarikciler_response["items"]
            elif isinstance(tedarikciler_response, list): # Eğer API doğrudan liste dönüyorsa
                tedarikciler = tedarikciler_response
                self.app.set_status_message("Uyarı: Tedarikçi listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                self.app.set_status_message("Hata: Tedarikçi listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Tedarikçi listesi API'den beklenen formatta gelmedi: {type(tedarikciler_response)} - {tedarikciler_response}")
                return # Hata durumunda fonksiyonu sonlandır

            self.tum_tedarikciler_cache = tedarikciler
            self._filtre_liste() 

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Tedarikçi listesi çekilirken hata: {e}")
            logging.error(f"Tedarikçi listesi yükleme hatası: {e}", exc_info=True)

    def _filtre_liste(self):
        """Arama kutusuna yazıldıkça tedarikçi listesini filtreler."""
        arama_terimi = self.search_entry.text().strip()
        normalized_arama_terimi = normalize_turkish_chars(arama_terimi).lower()
        
        self.tedarikci_tree.clear()
        
        for tedarikci_row in self.tum_tedarikciler_cache:
            tedarikci_id = tedarikci_row.get('id')
            tedarikci_kodu = tedarikci_row.get('kod', '')
            tedarikci_ad = tedarikci_row.get('ad')
            
            normalized_tedarikci_ad = normalize_turkish_chars(tedarikci_ad).lower() if tedarikci_ad else ''
            normalized_tedarikci_kodu = normalize_turkish_chars(tedarikci_kodu).lower() if tedarikci_kodu else ''
            
            if (not normalized_arama_terimi or
                (normalized_tedarikci_ad and normalized_arama_terimi in normalized_tedarikci_ad) or
                (normalized_tedarikci_kodu and normalized_arama_terimi in normalized_tedarikci_kodu)
               ):
                item_qt = QTreeWidgetItem(self.tedarikci_tree)
                item_qt.setText(0, tedarikci_ad)
                item_qt.setText(1, tedarikci_kodu)
                item_qt.setData(0, Qt.UserRole, tedarikci_id)
                
    def _sec(self, item=None, column=None): # item ve column QTreeWidget sinyalinden gelir
        """Seçili tedarikçiyi onaylar ve callback fonksiyonunu çağırır."""
        selected_items = self.tedarikci_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Seçim Yok", "Lütfen bir tedarikçi seçin.")
            return

        selected_tedarikci_id = selected_items[0].data(0, Qt.UserRole) # UserRole'dan ID'yi al
        selected_tedarikci_ad = selected_items[0].text(0) # Tedarikçi Adı sütunu
        
        self.callback_func(selected_tedarikci_id, selected_tedarikci_ad) # Callback'i çağır
        self.accept() # Pencereyi kapat  

class BeklemePenceresi(QDialog):
    def __init__(self, parent_app, title="İşlem Devam Ediyor...", message="Lütfen bekleyiniz..."):
        super().__init__(parent_app)
        self.setWindowTitle(title)
        self.setFixedSize(300, 120)
        self.setModal(True)

        # Pencereyi ana pencerenin ortasına konumlandır
        if parent_app:
            parent_rect = parent_app.geometry()
            x = parent_rect.x() + (parent_rect.width() - self.width()) // 2
            y = parent_rect.y() + (parent_rect.height() - self.height()) // 2
            self.move(x, y)

        main_layout = QVBoxLayout(self)
        message_label = QLabel(message)
        message_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
        message_label.setWordWrap(True)
        message_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(message_label)
        
        self.progressbar = QProgressBar()
        self.progressbar.setRange(0, 0)
        main_layout.addWidget(self.progressbar, alignment=Qt.AlignCenter)
        
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_DeleteOnClose)
        
        self.setWindowModality(Qt.ApplicationModal)
        self.closeEvent = self._do_nothing_close_event

    def _do_nothing_close_event(self, event):
        event.ignore()

    def kapat(self):
        self.close()
        
class GelirGiderSiniflandirmaYonetimiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, yenile_callback):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.yenile_callback = yenile_callback # Ana pencereyi yenilemek için

        self.setWindowTitle("Gelir/Gider Sınıflandırma Yönetimi")
        self.setMinimumSize(600, 450)
        self.setModal(True)

        main_layout = QVBoxLayout(self)

        # Notebook (Sekmeler) oluştur
        self.notebook = QTabWidget(self)
        main_layout.addWidget(self.notebook)

        # Gelir Sınıflandırmaları Sekmesi
        self.gelir_frame = QWidget()
        self.notebook.addTab(self.gelir_frame, "Gelir Sınıflandırmaları")
        self._setup_siniflandirma_sekmesi(self.gelir_frame, "GELİR")

        # Gider Sınıflandırmaları Sekmesi
        self.gider_frame = QWidget()
        self.notebook.addTab(self.gider_frame, "Gider Sınıflandırmaları")
        self._setup_siniflandirma_sekmesi(self.gider_frame, "GİDER")

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignCenter)

        # Sağ tık menüsü (Ortak olabilir)
        self.context_menu = QMenu(self)
        self.context_menu.addAction("Güncelle").triggered.connect(self._siniflandirma_guncelle)
        self.context_menu.addAction("Sil").triggered.connect(self._siniflandirma_sil)

    def _setup_siniflandirma_sekmesi(self, parent_frame, tip):
        frame_layout = QVBoxLayout(parent_frame) # Çerçeveye bir layout ata

        # Arama ve Ekleme alanı
        top_frame = QFrame(parent_frame)
        top_layout = QHBoxLayout(top_frame)
        frame_layout.addWidget(top_frame)

        top_layout.addWidget(QLabel("Sınıflandırma Adı:")) # "Yeni Sınıflandırma Adı:" yerine "Sınıflandırma Adı:"
        entry = QLineEdit()
        top_layout.addWidget(entry)

        add_button = QPushButton("Ekle")
        add_button.clicked.connect(lambda: self._siniflandirma_ekle(tip, entry.text().strip(), entry))
        top_layout.addWidget(add_button)

        # Yeni Eklendi: Güncelle butonu ve fonksiyonu
        update_button = QPushButton("Güncelle")
        update_button.clicked.connect(lambda: self._siniflandirma_guncelle_dogrudan(tip, entry.text().strip(), entry)) # Yeni doğrudan güncelleme metodu
        top_layout.addWidget(update_button)

        # Yeni Eklendi: Sil butonu ve fonksiyonu
        delete_button = QPushButton("Sil")
        delete_button.clicked.connect(lambda: self._siniflandirma_sil(tip)) # Sil metodu doğrudan çağrıldı
        top_layout.addWidget(delete_button)

        # Treeview alanı
        tree_frame = QFrame(parent_frame)
        tree_layout = QVBoxLayout(tree_frame)
        frame_layout.addWidget(tree_frame)

        tree = QTreeWidget()
        tree.setHeaderLabels(["ID", "Sınıflandırma Adı"])
        tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        tree.setColumnWidth(0, 50)
        tree.header().setSectionResizeMode(1, QHeaderView.Stretch) # Sınıflandırma Adı genişlesin
        tree_layout.addWidget(tree)

        # Treeview'i kaydet
        if tip == "GELİR":
            self.gelir_tree = tree
            self.gelir_entry = entry # Gelir giriş alanını sınıf özelliği olarak kaydet
        else:
            self.gider_tree = tree
            self.gider_entry = entry # Gider giriş alanını sınıf özelliği olarak kaydet

        # Seçim değiştiğinde giriş alanını doldurmak için bağlantı
        tree.itemSelectionChanged.connect(lambda: self._on_siniflandirma_select(tree, entry))

        # Sağ tık menüsünü treeview'e bağla
        tree.setContextMenuPolicy(Qt.CustomContextMenu)
        tree.customContextMenuRequested.connect(self._on_treeview_right_click)

        self._load_siniflandirmalar(tip)

    def _load_siniflandirmalar(self, tip):
        tree = self.gelir_tree if tip == "GELİR" else self.gider_tree

        tree.clear() # Mevcut öğeleri temizle

        siniflandirmalar = []
        try:
            if tip == "GELİR":
                # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
                siniflandirmalar = self.db.gelir_siniflandirma_listele()
            else:
                # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
                siniflandirmalar = self.db.gider_siniflandirma_listele()
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"{tip} sınıflandırmaları çekilirken hata: {e}")
            logging.error(f"{tip} sınıflandırma yükleme hatası: {e}", exc_info=True)
            return

        for s_item in siniflandirmalar:
            item_qt = QTreeWidgetItem(tree)
            item_qt.setText(0, str(s_item.get('id')))
            item_qt.setText(1, s_item.get('ad')) # 'siniflandirma_adi' yerine 'ad' kullanıldı
            item_qt.setData(0, Qt.UserRole, s_item.get('id')) # ID'yi UserRole olarak sakla

    def _siniflandirma_ekle(self, tip, siniflandirma_adi, entry_widget):
        if not siniflandirma_adi:
            QMessageBox.warning(self, "Uyarı", "Sınıflandırma adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle(f"{tip.lower()}_siniflandirmalari", {"ad": siniflandirma_adi})
            if success:
                QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla eklendi.")
                entry_widget.clear()
                self._load_siniflandirmalar(tip)
                if self.yenile_callback:
                    self.yenile_callback()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırma eklenirken hata: {e}")
            logging.error(f"Sınıflandırma ekleme hatası: {e}", exc_info=True)

    def _on_siniflandirma_select(self, tree, entry_widget):
        selected_items = tree.selectedItems()
        if selected_items:
            # Sınıflandırma adını al ve giriş alanına yerleştir
            siniflandirma_adi = selected_items[0].text(1) 
            entry_widget.setText(siniflandirma_adi)
        else:
            entry_widget.clear()

    def _on_treeview_right_click(self, pos):
        current_tab_text = self.notebook.tabText(self.notebook.currentIndex())

        tree = None
        entry_widget = None # Güncelleme için giriş alanını alacağız
        if "Gelir Sınıflandırmaları" in current_tab_text:
            tree = self.gelir_tree
            entry_widget = self.gelir_entry
        else:
            tree = self.gider_tree
            entry_widget = self.gider_entry

        item = tree.itemAt(pos) # Position'dan öğeyi al

        if item:
            tree.setCurrentItem(item) # Öğeyi seçili hale getir (sağ tıklama ile seçilmemiş olabilir)

            context_menu = QMenu(self)

            # Düzeltildi: _siniflandirma_guncelle metodunu doğrudan çağırıldı
            update_action = context_menu.addAction("Güncelle")
            update_action.triggered.connect(self._siniflandirma_guncelle)

            # Silme işlemi
            delete_action = context_menu.addAction("Sil")
            delete_action.triggered.connect(lambda: self._siniflandirma_sil(
                "GELİR" if "Gelir" in current_tab_text else "GİDER"
            ))

            context_menu.exec(tree.mapToGlobal(pos))
        else:
            # Boş alana tıklandığında menüyü gizle/kapat (eğer açıksa)
            if hasattr(self, 'context_menu') and self.context_menu.isVisible():
                self.context_menu.hide()

    def _siniflandirma_guncelle(self):
        current_tab_text = self.notebook.tabText(self.notebook.currentIndex())
        tree = self.gelir_tree if "Gelir Sınıflandırmaları" in current_tab_text else self.gider_tree
        entry_widget = self.gelir_entry if "Gelir Sınıflandırmaları" in current_tab_text else self.gider_entry
        tip = "GELİR" if "Gelir" in current_tab_text else "GİDER"

        selected_items = tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen güncellemek istediğiniz sınıflandırmayı seçin.")
            return

        siniflandirma_id = selected_items[0].data(0, Qt.UserRole)
        mevcut_siniflandirma_adi = selected_items[0].text(1)
        yeni_siniflandirma_adi = entry_widget.text().strip()

        if not yeni_siniflandirma_adi or yeni_siniflandirma_adi == mevcut_siniflandirma_adi:
            QMessageBox.warning(self, "Uyarı", "Yeni sınıflandırma adı boş olamaz veya mevcut adla aynı olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle(f"{tip.lower()}_siniflandirmalari", siniflandirma_id, {"ad": yeni_siniflandirma_adi})
            if success:
                QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla güncellendi.")
                entry_widget.clear()
                self._load_siniflandirmalar(tip)
                if self.yenile_callback:
                    self.yenile_callback()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırma güncellenirken hata: {e}")
            logging.error(f"Sınıflandırma güncellenirken hata: {e}", exc_info=True)

    def _siniflandirma_sil(self):
        current_tab_text = self.notebook.tabText(self.notebook.currentIndex())
        
        tree = None
        tip = ""
        if "Gelir Sınıflandırmaları" in current_tab_text:
            tree = self.gelir_tree
            tip = "GELİR"
        else:
            tree = self.gider_tree
            tip = "GİDER"

        selected_items = tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek istediğiniz sınıflandırmayı seçin.")
            return
        siniflandirma_id = selected_items[0].data(0, Qt.UserRole)
        siniflandirma_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self, "Onay", f"'{siniflandirma_adi}' sınıflandırmasını silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil(f"{tip.lower()}_siniflandirmalari", siniflandirma_id)
                if success:
                    QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla silindi.")
                    self._load_siniflandirmalar(tip)
                    if self.yenile_callback:
                        self.yenile_callback()
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "API Hatası", f"Sınıflandırma silinirken hata: {e}")
                logging.error(f"Sınıflandırma silme hatası: {e}", exc_info=True)

class BirimDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, birim_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.birim_id = birim_info['id']
        self.mevcut_birim_adi = birim_info['birim_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"Birim Düzenle: {self.mevcut_birim_adi}")
        self.setFixedSize(350, 200) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Birim Adı:"), 0, 0, Qt.AlignCenter)
        self.birim_adi_entry = QLineEdit()
        self.birim_adi_entry.setText(self.mevcut_birim_adi)
        main_frame_layout.addWidget(self.birim_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignCenter) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_birim_adi = self.birim_adi_entry.text().strip()
        if not yeni_birim_adi:
            QMessageBox.warning(self, "Uyarı", "Birim adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            response = self.db.nitelik_guncelle("urun_birimleri", self.birim_id, {"ad": yeni_birim_adi})

            QMessageBox.information(self, "Başarılı", "Birim başarıyla güncellendi.")
            self.yenile_callback() # Ana listedeki birimleri yenile
            self.accept() # Pencereyi kapat
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Birim güncellenirken hata: {e}")
            logging.error(f"Birim güncelleme hatası: {e}", exc_info=True)

class GrupDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, grup_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.grup_id = grup_info['id']
        self.mevcut_grup_adi = grup_info['grup_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"Grup Düzenle: {self.mevcut_grup_adi}")
        self.setFixedSize(350, 200) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Grup Adı:"), 0, 0, Qt.AlignCenter)
        self.grup_adi_entry = QLineEdit()
        self.grup_adi_entry.setText(self.mevcut_grup_adi)
        main_frame_layout.addWidget(self.grup_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignCenter) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_grup_adi = self.grup_adi_entry.text().strip()
        if not yeni_grup_adi:
            QMessageBox.warning(self, "Uyarı", "Grup adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            response = self.db.nitelik_guncelle("urun_gruplari", self.grup_id, {"ad": yeni_grup_adi})

            QMessageBox.information(self, "Başarılı", "Grup başarıyla güncellendi.")
            self.yenile_callback()
            self.accept()
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Grup güncellenirken hata: {e}")
            logging.error(f"Grup güncelleme hatası: {e}", exc_info=True)

# UlkeDuzenlePenceresi sınıfı
class UlkeDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, ulke_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.ulke_id = ulke_info['id']
        self.mevcut_ulke_adi = ulke_info['ulke_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"Ülke Düzenle: {self.mevcut_ulke_adi}")
        self.setFixedSize(350, 200) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Ülke Adı:"), 0, 0, Qt.AlignCenter)
        self.ulke_adi_entry = QLineEdit()
        self.ulke_adi_entry.setText(self.mevcut_ulke_adi)
        main_frame_layout.addWidget(self.ulke_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignCenter) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_ulke_adi = self.ulke_adi_entry.text().strip()
        if not yeni_ulke_adi:
            QMessageBox.warning(self, "Uyarı", "Ülke adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            response = self.db.nitelik_guncelle("ulkeler", self.ulke_id, {"ad": yeni_ulke_adi})

            QMessageBox.information(self, "Başarılı", "Ülke başarıyla güncellendi.")
            self.yenile_callback()
            self.accept()
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Ülke güncellenirken hata: {e}")
            logging.error(f"Ülke güncelleme hatası: {e}", exc_info=True)

class SiniflandirmaDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, tip, siniflandirma_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.tip = tip # "GELİR" veya "GİDER"
        self.siniflandirma_id = siniflandirma_info['id']
        self.mevcut_siniflandirma_adi = siniflandirma_info['siniflandirma_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"{tip.capitalize()} Sınıflandırma Düzenle: {self.mevcut_siniflandirma_adi}")
        self.setFixedSize(400, 220) # Boyut ayarı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Sınıflandırma Adı:"), 0, 0, Qt.AlignCenter)
        self.siniflandirma_adi_entry = QLineEdit()
        self.siniflandirma_adi_entry.setText(self.mevcut_siniflandirma_adi)
        main_frame_layout.addWidget(self.siniflandirma_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignCenter) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_siniflandirma_adi = self.siniflandirma_adi_entry.text().strip()
        if not yeni_siniflandirma_adi:
            QMessageBox.warning(self, "Uyarı", "Sınıflandırma adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            success, message = self.db.nitelik_guncelle(f"{self.tip.lower()}_siniflandirmalari", self.siniflandirma_id, {"ad": yeni_siniflandirma_adi})

            if success:
                QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla güncellendi.")
                self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırma güncellenirken hata: {error_detail}")
            logging.error(f"Sınıflandırma güncelleme hatası: {error_detail}", exc_info=True)

class KullaniciKayitPenceresi(QDialog):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.setWindowTitle("Yeni Kullanıcı Hesabı Oluştur")
        self.setFixedSize(350, 200)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Yeni Hesap Oluştur")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)

        form_layout.addWidget(QLabel("Kullanıcı Adı:"), 0, 0)
        self.kullanici_adi_entry = QLineEdit()
        self.kullanici_adi_entry.setPlaceholderText("Benzersiz kullanıcı adı")
        form_layout.addWidget(self.kullanici_adi_entry, 0, 1)

        form_layout.addWidget(QLabel("Şifre:"), 1, 0)
        self.sifre_entry = QLineEdit()
        self.sifre_entry.setEchoMode(QLineEdit.Password)
        self.sifre_entry.setPlaceholderText("Şifre")
        form_layout.addWidget(self.sifre_entry, 1, 1)

        form_layout.addWidget(QLabel("Şifre Tekrar:"), 2, 0)
        self.sifre_tekrar_entry = QLineEdit()
        self.sifre_tekrar_entry.setEchoMode(QLineEdit.Password)
        self.sifre_tekrar_entry.setPlaceholderText("Şifre tekrarı")
        form_layout.addWidget(self.sifre_tekrar_entry, 2, 1)
        
        main_layout.addStretch()

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()

        btn_kaydet = QPushButton("Hesap Oluştur")
        btn_kaydet.clicked.connect(self._hesap_olustur)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.reject)
        button_layout.addWidget(btn_iptal)

        self.kullanici_adi_entry.setFocus()

    def _hesap_olustur(self):
        kullanici_adi = self.kullanici_adi_entry.text().strip()
        sifre = self.sifre_entry.text().strip()
        sifre_tekrar = self.sifre_tekrar_entry.text().strip()

        if not (kullanici_adi and sifre and sifre_tekrar):
            QMessageBox.warning(self, "Eksik Bilgi", "Lütfen tüm alanları doldurun.")
            return

        if sifre != sifre_tekrar:
            QMessageBox.warning(self, "Şifre Hatası", "Girilen şifreler eşleşmiyor.")
            return

        try:
            # Yeni API rotasını çağırıyoruz
            data = {"kullanici_adi": kullanici_adi, "sifre": sifre}
            response = requests.post(f"{self.db.api_base_url}/dogrulama/register", json=data)
            response.raise_for_status()
            
            QMessageBox.information(self, "Başarılı", "Yeni kullanıcı hesabınız başarıyla oluşturuldu.")
            self.accept()
        except requests.exceptions.RequestException as e:
            error_detail = "Kullanıcı adı zaten mevcut."
            if e.response is not None and e.response.json():
                error_detail = e.response.json().get('detail', str(e))

            QMessageBox.critical(self, "Kayıt Hatası", f"Kullanıcı oluşturulurken bir hata oluştu:\n{error_detail}")
        except Exception as e:
            QMessageBox.critical(self, "Beklenmeyen Hata", f"Beklenmeyen bir hata oluştu: {e}")

class PersonelYonetimiPenceresi(QDialog):
    def __init__(self, parent, db_manager):
        super().__init__(parent)
        self.db_manager = db_manager
        
        # Ana pencereden firma no bilgisi çekiliyor
        self.firma_no = getattr(parent, 'firma_no', "N/A")
        # Güncelleme sırasında kullanılacak versiyon numarasını saklamak için
        self.current_version = 1 
        
        self.setWindowTitle("Personel Yönetimi")
        self.setMinimumSize(700, 400) 

        self.layout = QVBoxLayout(self)

        # Personel Listesi Tablosu
        self.table_widget = QTableWidget()
        
        self.table_widget.setColumnCount(7) 
        self.table_widget.setHorizontalHeaderLabels([
            "ID", "Ad", "Soyad", "E-posta", "Telefon", "Rol", "Aktif"
        ])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) 
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)

        # FIRMA NO ETİKETİ
        self.firma_no_label = QLabel(f"Firma No: {self.firma_no}")
        self.firma_no_label.setAlignment(Qt.AlignCenter)
        palette = self.firma_no_label.palette()
        palette.setColor(QPalette.WindowText, QColor(255, 0, 0))
        self.firma_no_label.setPalette(palette)

        # Butonlar
        self.btn_yeni_personel = QPushButton("Yeni Personel Ekle")
        self.btn_yenile = QPushButton("Listeyi Yenile")

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.btn_yenile)
        button_layout.addStretch()
        button_layout.addWidget(self.firma_no_label) 
        button_layout.addStretch()
        button_layout.addWidget(self.btn_yeni_personel)

        self.layout.addWidget(self.table_widget)
        self.layout.addLayout(button_layout)

        # Sinyaller
        self.btn_yeni_personel.clicked.connect(lambda: self._personel_formu_dialog(None)) # Yeni form aç
        self.btn_yenile.clicked.connect(self.personel_listesini_yukle)
        self.table_widget.doubleClicked.connect(self._personel_duzenle_trigger) # Düzenleme tetikleyici
        
        self.personel_listesini_yukle()


    def personel_listesini_yukle(self):
        """API veya DB'den personel listesini çeker ve tabloyu günceller."""
        personeller_response, hata = self.db_manager.personel_listesi_getir()
        
        if hata:
            QMessageBox.critical(self, "Hata", f"Personel listesi yüklenemedi:\n{hata}")
            self.table_widget.setRowCount(0) 
            return

        if personeller_response is None or not isinstance(personeller_response, dict) or "items" not in personeller_response:
            QMessageBox.critical(self, "Hata", "Personel listesi alınamadı veya sunucu hatası devam ediyor. Lütfen sunucu loglarını kontrol edin.")
            self.table_widget.setRowCount(0)
            return

        personel_listesi = personeller_response.get("items", [])
        
        self.table_widget.setRowCount(len(personel_listesi))
        self.table_widget.setColumnCount(7) 
        
        for row, personel_data in enumerate(personel_listesi):
            personel = personel_data 
            
            if isinstance(personel_data, str):
                try:
                    personel = json.loads(personel_data)
                except:
                    continue
            
            if not isinstance(personel, dict):
                continue

            # Tabloyu yeni sütun sırasına göre doldur
            self.table_widget.setItem(row, 0, QTableWidgetItem(str(personel.get('id', 'N/A'))))
            self.table_widget.setItem(row, 1, QTableWidgetItem(personel.get('ad', 'N/A')))
            self.table_widget.setItem(row, 2, QTableWidgetItem(personel.get('soyad', 'N/A')))
            self.table_widget.setItem(row, 3, QTableWidgetItem(personel.get('email', 'N/A')))
            # KRİTİK EKLENTİ: Telefon bilgisi doldurma
            self.table_widget.setItem(row, 4, QTableWidgetItem(personel.get('telefon', 'N/A'))) 
            self.table_widget.setItem(row, 5, QTableWidgetItem(personel.get('rol', 'N/A')))
            self.table_widget.setItem(row, 6, QTableWidgetItem("Aktif" if personel.get('aktif') else "Pasif"))

    
    def _personel_detay_getir(self, personel_id: int):
        """db_manager üzerinden API'den tek bir personelin detayını çeker."""
        # API'de /yonetici/personel/{id} rotasının var olduğu varsayılır.
        return self.db_manager.personel_detay_getir(personel_id)


    def _personel_duzenle_trigger(self, index):
        """Tabloda çift tıklanan personelin ID'sini alıp düzenleme dialogunu açar."""
        if index.isValid():
            row = index.row()
            personel_id_item = self.table_widget.item(row, 0)
            if personel_id_item:
                personel_id = int(personel_id_item.text())
                
                # API proxy metodunu çağırarak detayları al
                personel_detay, hata = self._personel_detay_getir(personel_id)
                
                if hata:
                    QMessageBox.critical(self, "Hata", f"Personel detayları çekilemedi: {hata}")
                    return

                if isinstance(personel_detay, dict) and personel_detay.get('id'):
                    self._personel_formu_dialog(personel_detay)
                else:
                    QMessageBox.warning(self, "Hata", "Personel detayları doğru formatta alınamadı.")


    def _personel_formu_dialog(self, personel_data=None):
        """Personel ekleme/düzenleme formunu açar."""
        dialog = QDialog(self)
        is_duzenle = personel_data is not None
        dialog.setWindowTitle("Personel Düzenle" if is_duzenle else "Yeni Personel Oluştur")
        form_layout = QFormLayout(dialog)
        
        # Zorunlu alan etiketi için stil (Kırmızı *)
        zorunlu_stil = "<span style='color: red;'>*</span>"
        
        # 1. Input Alanları
        ad_input = QLineEdit(dialog)
        soyad_input = QLineEdit(dialog)
        email_input = QLineEdit(dialog)
        telefon_input = QLineEdit(dialog) # TELEFON ALANI EKLENDİ
        rol_combo = QComboBox(dialog)
        aktif_check = QCheckBox("Aktif", dialog)
        aktif_check.setChecked(True)
        sifre_input = QLineEdit(dialog)
        sifre_input.setEchoMode(QLineEdit.Password)
        
        # 2. Rol Kısıtlamaları
        personel_id = personel_data.get('id', 0) if is_duzenle else 0
        mevcut_rol = personel_data.get('rol', 'personel').lower() if is_duzenle else 'personel'
        
        # Ana Yönetici (ID=1) rolü değiştirilemez ve sadece kendi rolünü gösterir.
        if personel_id == 1: 
            rol_combo.setDisabled(True)
            rol_combo.addItem(mevcut_rol.upper())
        else:
            # Diğer personeller için sadece personel ve yönetici seçenekleri
            rol_combo.addItems(["personel", "yonetici"])
        
        # 3. Verileri Doldur (Düzenleme Modu)
        if is_duzenle:
            self.current_version = personel_data.get('version', 1) 
            
            ad_input.setText(personel_data.get('ad', ''))
            soyad_input.setText(personel_data.get('soyad', ''))
            email_input.setText(personel_data.get('email', ''))
            telefon_input.setText(personel_data.get('telefon', '')) # TELEFON DOLDURMA
            
            if personel_id != 1:
                 rol_combo.setCurrentText(mevcut_rol)
            
            aktif_check.setChecked(personel_data.get('aktif', True))

        # 4. Form Düzeni (Zorunlu Alan İşaretleri)
        form_layout.addRow("Ad:" + zorunlu_stil, ad_input) 
        form_layout.addRow("Soyad:", soyad_input)
        form_layout.addRow("E-posta:" + zorunlu_stil, email_input) 
        form_layout.addRow("Telefon:", telefon_input) 
        form_layout.addRow("Rol:", rol_combo)
        form_layout.addRow("Durum:", aktif_check)
        form_layout.addRow("Şifre (Değiştirmek için doldurun):", sifre_input) 
        
        # 5. Butonlar ve Bağlantılar
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, Qt.Horizontal, dialog)
        form_layout.addRow(buttons)
        buttons.accepted.connect(lambda: self._personel_kaydet_ve_kapat(
            dialog, is_duzenle, ad_input, soyad_input, email_input, telefon_input, rol_combo, sifre_input, aktif_check, personel_data.get('id') if is_duzenle else None
        ))
        buttons.rejected.connect(dialog.reject)

        dialog.exec()        
    
    def _personel_kaydet_ve_kapat(self, dialog, is_duzenle, ad_input, soyad_input, email_input, telefon_input, rol_combo, sifre_input, aktif_check, personel_id):
        """Personel kaydetme ve güncelleme işlemini yapar, zorunlu alanları kontrol eder."""
        
        # 1. Veri Toplama ve Zorunluluk Kontrolü
        ad = ad_input.text().strip()
        email = email_input.text().strip()
        sifre = sifre_input.text()
        
        # Zorunluluk Kontrolleri
        if not ad:
            QMessageBox.warning(dialog, "Eksik Bilgi", "Ad alanı zorunludur. Lütfen doldurunuz.")
            return
        if not email:
            QMessageBox.warning(dialog, "Eksik Bilgi", "E-posta alanı zorunludur. Lütfen doldurunuz.")
            return
        if not is_duzenle and not sifre:
             QMessageBox.warning(dialog, "Eksik Bilgi", "Yeni personel için şifre zorunludur.")
             return


        # 2. Veri Hazırlama
        personel_data = {
            "ad": ad,
            "soyad": soyad_input.text().strip(),
            "email": email,
            "telefon": telefon_input.text().strip(), # TELEFON DAHİL EDİLDİ
            "aktif": aktif_check.isChecked()
        }
        
        # Rolü sadece değiştirilebilir durumda ise al
        if not rol_combo.isDisabled():
            personel_data["rol"] = rol_combo.currentText().lower()
        
        # Şifre varsa eklenir
        if sifre:
            personel_data["sifre"] = sifre

        # 3. API Çağrısı
        if is_duzenle:
            # Version eklenir
            personel_data["version"] = self.current_version 
            personel_data["rol"] = personel_data.get("rol", rol_combo.currentText().lower()) # Pasif rol alanındaki değeri koru
            
            basarili, mesaj = self.db_manager.personel_guncelle(personel_id, personel_data) 
        else:
            basarili, mesaj = self.db_manager.personel_olustur(personel_data)

        # 4. Sonuç
        if basarili:
            QMessageBox.information(self, "Başarılı", mesaj)
            self.personel_listesini_yukle() 
            dialog.accept() 
        else:
            QMessageBox.critical(self, "Hata", f"İşlem başarısız oldu:\n{mesaj}")

    def _yeni_personel_ekle_dialog(self):
        """Yeni personel ekleme butonundan çağrılır."""
        self._personel_formu_dialog(None)            